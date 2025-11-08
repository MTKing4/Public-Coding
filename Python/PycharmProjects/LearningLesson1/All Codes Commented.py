#outputs a string, can use ' and "#
print("Hello")

# backslash + n to print on a new line with a single print, no space needed after typing it
print("Hello, \nworld")

#Define a Variable called age and give it a value of 20#
age = 20

#outputs a variable#
print(age)

#Code executes from top to bottom, the output will be the last defined variable, 30#
age = 20
age = 30
print(age)

#Can use decimal pointed numbers when defining#
price = 19.25

#Bolean Value (True, False) First letter CAPS - CASE SENSITIVE#
is_online = True

#is a function to get an input from the terminal window and stores it as a string#
input("what is your name?")

#to make the cursor be on the second line to make it look better, add \n
input("what is your name?\n")

#use \ to ignore the symbol after it, if your script has both ' and " use either ' or " to declare your string, and add / before every occurrence of the other symbol, in this example it was /' the ' is ignored here
input('You\'re at a cross road, go "left" or "right"')

#storing the input in name variable and print hello followed by the name# INPUT DISPLAYS ON THE SCREEN ON ITS OWN WITHOUT USING PRINT
name = input("what is your name? ")
print("Hello! " + name)

print("Hello" + "" + "Angela") #another way of adding space between two concatenated strings

# Nested input() inside print(), input will be executed first, then the result of that will be passed to the input function, and the print function will display both the string and the result of the input function
print("Hello! " + input("what is your name? "))

print("*" * 10) #can multiply a string with an integer to display the string 10 times#

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
#Excersice to switch up the variables

a = input("a:")
b = input("b:")

temp_a = a
a = b
b = temp_a

print("a = " + a)
print("b = " + b)

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

#3 types of data in python#
10 #Numbers#
"Mosh" #Strings#
True #Boolean#

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

#Input your birth year and it will show you your age#
birth_year = input("Enter your Birth Year: ") #storing the value as a string in your input() function in a variable#
AgeNumber = int(birth_year) #convert the string number in birth_year variable into an integer and store it in a variable#
age = 2024 - AgeNumber # now the subtraction is integer-integer#
print(age)
print(type(age)) #type function returns the type of the variable#
print(len(str(age))) #len() funtion doesn't work with integers, so we converted it to string

int() #converts a value to an integer#
float() #converts s value to a floating point number#
bool() #converts a value to a boolean#
str() #converts a value to a string#


#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

#Subscript
print("Hello"[0]) #subscripting us using square brackets to print specific letters based on the index number between the brackets (index number = the position/order of the letter in the string)

#for large integers to be easily read, like 1,000,000 in python we type _ instead, and python will ignore it
print(1_000_000)

#type converters
str() # to a string
float() # to a float

#flocats can be mathed with integers
print(1_000_000 + float(10.45))

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
#Exersice to sum two numbers of a two digit number together

integer = input()

digit_adder =  int(integer[0]) + int(integer[1])

print(digit_adder)


#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

#Calculator:#
First = float(input("First: "))
Second = float(input("Second: "))
Sum = First + Second
print("Sum: " + str(Sum)) #to Print an integer with a stirng you have to convert the integer to string as well#

#when a function is part of an object it is referred to as a "Method"#

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

course = 'Python for beginners'
# index   01234567#
print(course.upper()) #it creates a new string, it won't modify the original string#
print(course.lower())
print(course.find('y')) #it will return the index of the first occurence of 'y', it will return '-1' if not found#
print(course.find('for')) #it will return the index of the first letter of 'for' of the first occurence#
print(course.replace('for', 'Loves')) #if look for a character that doesn't exist nothing's going to happen#
print('Python' in course) # looks for the word Python, 'in' is an operator, a special keyword. the result is a boolean value#
print(course[0:3]) #it will print only the letters from index 0 to 3 excluding 3, can also use [1:] to return all the characters til the end of the string  starting from 1, also if you type it like so [:3] python will assume the first index is 0, if you use [:] it will bring everyhting#
print(course[-1]) #index of the last character, -2 index of the second last character etc#
print(course[1:-1]) # will bring from the second index to the last index exluding the last index -1
print(len(course)) # len() function is used to count the characters in your string, can also count the number of items in a list
print(course.title()) # makes all initials Caps in the words of the string

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

#Arthemetic operators#
print(10 + 3) #Addition#
print(10 - 3) # Subtraction#
print(10 * 3) #Multipliction#
print(10 / 3) #Division returning a decimal point, Float Value#
print(10 // 3)# Division returning an ineger.#
print(10 % 3) #Modulous, returning the remainder of a division#
print(10 ** 3) #Exponent, takes precedence over multiplication/divison#
x = 10
x = x + 3
x += 3 #same with other signs, this is called augmented assignement operator,enhanced#
x = 10 + 3 * 2 #Operator Precedence, can use paranthesis (), to change the order#
x = 3 > 2
x = 3 >= 2
x = 3 <= 2
x = 3 == 2
x = 3 != 2
print(x)

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
# Age calculator exersice

age = input("what is your current age? ")

result = 90 - int(age)
days = result * 365
weeks = result * 52
months = result * 12

print(f"you have {days} Days, {weeks} weeks , and {months} months left")

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
#bill calculator exersice

bill = float(input("what bill "))
tip = int(input("what tip "))
people = int(input("people? "))
percentage = tip / 100
bill_tip = bill * percentage
total = bill + bill_tip
split = total / people
# rounded = round(split, 2)             this only rounds the number to a certain digit number if there are MORE DIGITS, if not it will show less, like 1 digit, that's why it was commented out
rounded = "{:.2f}".format(split)        #using the format() function, we're forcing the result to show up with 2 decimal points even if there isn't one, the :.2f means we want 2 digits after the floating point, this function turned the float number into a string
print(rounded)

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

#Logical Operators#
price = 25
print(price > 10 and price > 30) #can use 'or'#
print(not price > 10)

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

#if statment
temperature = 25
if temperature > 30:  # the colon ':' means 'then',#
    print("It's a hot day")  #used " instead of of ' because there's ' already in the string#
    print("drink plenty of water") #there's no Curly braces {} in python, indentation is used instead#
elif temperature > 20:
    print("it's a nice day")
else:
    print("it's cold")
print("Done")

#strings note: use " when you have ' in your string, and use ' when you have " in your string, and use ''' to print a multiple line string

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

#even and odd exersice

number = int(input("your number "))

if number % 2 == 0:
    print("even")
else:
    print("odd")

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

# leap year exersice

year = int(input("enter the year "))

if year % 4 == 0:
    if year % 100 == 0:
        if year % 400 == 0:
            print("leap")
        else:
            print("not leap")
    else:
        print("leap")

else:
    print("not leap")

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

# multiple if conditions
height = int(input("what your height in cm "))
bill = 0

if height >= 120:
    print("you can ride")
    age = int(input("what is your age? "))
    if age < 12:
        bill = 5
        print(f"child ticket is {bill}")
    elif age <= 18:
        bill = 7
        print(f"Youth ticket {bill}")
    else:
        bill = 12
        print(f"adult ticket is {bill} ")

    picture_taken = input("you want picture? type Y or N ").lower()
    if picture_taken == "y":                                        #a second if statement in the same indentation of the first but is independent of the first and both can be true, this will always execute after the first if is done executing, no need for an else statement because we will do nothing if this statement was false
        bill += 3
    print(f"your final bill is {bill}")
else:
    print("you can't ride")

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

#Weight Converter#
Weight = float(input("Weight: ")) #Stored as float to be able to operated on in the converted variable
Unit = input("K(g) or (L)bs: ") #input stores string only
if Unit.upper() == ("K"):
    Converted = Weight / 0.45 #float divided by float
    print("Weight in Lbs " + str(Converted)) #Converted Variable has to be converted to string to in order to concatenate it with a string#
else:
    Converted = Weight * 0.45
    print("Weight in Kgs " + str(Converted))

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

#while Loop#
i = 1
while i <= 5:
    print(i)
    i = i + 1

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

#ex.1#
i = 1
while i <= 10:
    print(i * "*") #can multiply a number with a string, will repeat the string "*" 10 times
    i = i + 1

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

#ex.2: Descending#
i = 10
while i <= 10 and i > 0:
    print(i * "*")
    i = i - 1

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

#Lists#
names = ["John", "Bob", "Mosh", "Sam", "Mary"]
names[0] = "Jon"  # can change an opject at a given index
print(names[2:]) # can get items from a range till the end of the list, ex. here from index 2 till the end of the list
print(names[0:2]) # this will return a sublist from the original list even if it has 1 item, the numbers are the index for the items in the list, Syntax: names(from:to], so from index 0 to index 2, you can use negagive index, -1 will show the first from the end of the list, mary, -2 will show the second of the last and so on
print(names[0:4:2]) # the third number the number of steps to jump to, in this case, it will get every other item
print(names[::2]) # get every other item
print(names[::-1]) # this will reverse the order of items, from the end to the begining
print(names[0]) # this returns an integer, not a sublist (can't be used with for loops after the in clause)
print(names[:]) # this will return all the items from the begining to the end of the list
print(names)  # Will return the original list unchanged#
#NOTE: most/all of these work with tuples as well

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
# nested lists

fruits = ["banana", "manga"]
vegetables = ["potato", "tomato"]
freshs = [fruits, vegetables]

print(freshs)

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

#List Methods#
numbers = [1, 2 ,3 ,4 ,5]
numbers.clear()
numbers.append(6)                   # adds to the list at the end#
numbers.insert(0, -1)               # first number is the index of where you want to insert, the second number is the number you want to insert
numbers.remove(-1)
numbers.pop()                       # it removes an item from the end of the list
numbers.sort()                      # can't be printed, this only sorts the numbers ascendingly, printing it will return: None, which is an object in python that represents the absence of a value
numbers.reverse()                   # reverses the order of the list, also can't be printed
numbers.copy()                      # copies the list, any changes to the original list won't affect it
numbers.extend([34, 54, 65, 65])    # add more than 1 item (a list) to the list
print(numbers.index(6))             # gives you the index number of the first occurance of the element, the number between the parenthesis is the item(integer, String, etc) you are looking for
print(numbers)
print(6 in numbers)                 # looks for 6 in the list, returms a boolean
print(len(numbers))                 # len() returns the number of the elements in the list#
print(numbers.count(6))             # returns the number of occurances this item has in the list, the number in the parenthesis is an item(integer, String, etc)


#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
#treasure map exercise (my version: God Awful, she done it in three lines)


row1 = ["⬜", "⬜", "⬜"]
row2 = ["⬜", "⬜", "⬜"]
row3 = ["⬜", "⬜", "⬜"]

map = [row1, row2, row3]

print(f"{row1}\n{row2}\n{row3}\n")

position = input("where put treasure?")


if position[0] == "1":
    if position[1] == "1":
        map[0][0] = "❤️"
    elif position[1] == "2":
        map[1][0] = "❤️"
    else:
        map[2][0] = "❤️"
elif position[0] == "2":
    if position[1] == "1":
        map[0][1] = "❤️"
    elif position[1] == "2":
        map[1][1] = "❤️"
    else:
        map[2][1] = "❤️"
elif position[0] == "3":
    if position[1] == "1":
        map[0][2] = "❤️"
    elif position[1] == "2":
        map[1][2] = "❤️"
    else:
        map[2][2] = "❤️"

print(f"{row1}\n{row2}\n{row3}\n")


#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
#treasure map exercise (her version: Million times better)


row1 = ["⬜", "⬜", "⬜"]
row2 = ["⬜", "⬜", "⬜"]
row3 = ["⬜", "⬜", "⬜"]

map = [row1, row2, row3]

print(f"{row1}\n{row2}\n{row3}\n")

position = input("where put treasure?")

horizontal = int(position[0])               #horizontal represents columns, position[0] only works with strings, it will take the first digit from the left from the string we entered, and transfer it to integer (so that we can subtract from it to sync it with index numbers next) then store it in the horizontal variable
vertical = int(position[1])                 #vertical represents rows

map[vertical-1][horizontal-1] = "X"         #vertical will check in the parent list, map = [row1, row2, row3] so it will traverse vertically between the three sublists, horizontal will check each element inside the selected list, so it will go horizontally, -1 is to sync the entered number with the index count which starts at 0 instead of 1, whatever was selected will be replaced with "X"



print(f"{row1}\n{row2}\n{row3}\n")


#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

#for loop#
numbers = [1, 2 ,3 ,4 ,5]
for item in numbers: # for loop, it will iterate over all the values in the list, in each iteration it will hold 1 value, 1st iteration 1, 2nd 2, and present them on a separate line each
    print(item)
i = 0
while i < len(numbers): # same result as the for loop above,
    print(numbers[i])
    i = i + 1


#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
#Height Average (my code)

students = [180, 124, 165, 173, 189, 169, 146]

sum = 0
heights = 0

for student in students:
    sum += student              # can also use sum(student_heights) to get the sum total of items in the list (no for loop)
    heights += 1                # can also use len(student_heights) the number of the elements in a list (no for loop)

avg = sum / heights

print(round(avg))

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

#height average, teachers code easy mode (no for loop)

student_heights = input("Input a list of student heights: ") .split()         # items entered will be elements in a list, when no delimiter is specified in spit() function, it will be space by default
for n in range(0, len(student_heights)):                                      # this will loop to whatever is the end of the list (n will be the index number based on the range specifed from 0 to the number of the list, say 5, but 5 won't be included because it's a range, which is good because index numbers start from 0, so the 5th element index will be 4
    student_heights[n] = int(student_heights[n])                              # this line and the line above it is to iterate on all elements in the list and convert them from strings to integers
print(student_heights)

total = sum(student_heights)                                                  # sum(student_heights) to get the sum total of items in the list
number_of_students = len(student_heights)                                     # len(student_heights) the number of the elements in a list
average = round(total / number_of_students)

print(average)

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

#Range() Function#
numbers = range(5,10,2) # can use range(5) to do from 0 to 4, and range(5,10) 5 being the start 10 being the end and excluded, third value can be added to indicate steps, 2 means it will jump two numbers at a time
for range in numbers:
    print(range)

----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
# For Range example to add all the numbers from 1 to 100 together

total = 0

for number in range(1, 101):
    total = total + number

print(total)

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
#exersice to sum all even numbers

total = 0

for number in range(2, 101, 2):
    total = total + number

print(total)

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

#FizBuzz Game
#https://en.wikipedia.org/wiki/Fizz_buzz#:~:text=Fizz%20buzz%20is%20a%20group,with%20the%20word%20%22fizzbuzz%22.

for number in range (1, 101):
    if number % 3 == 0 and number % 5 == 0:
        print("FizzBuzz")
    elif number % 5 == 0:
            print("Buzz")
    elif number % 3 == 0:
                print("Fizz")
    else:
        print(number)

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

#password generator project ( i did it all by myself yay :D )

import random

letters = [ 'a', 'b', 'c', 'd', 'e', 'f', 'g', 'h', 'i', 'j', 'k', 'l', 'm', 'n',
            'o', 'p', 'q', 'r', 's', 't', 'u', 'v', 'w', 'x', 'y', 'z', 'A', 'B',
            'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P',
            'Q,' 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z']

numbers = ['0','1','2','3','4','5','6','7','8','9']
symbols = ['!', '#', '$', '%', '&', '(', ')', '*', '+']


print("welcome to the password Generator!")

nr_letters = int(input("how many letters? \n"))
nr_symbols = int(input("how many symbols? \n"))
nr_numbers = int(input("how many numbers? \n"))

password = ""

for letter in range(nr_letters):
    letter = random.choice(letters)
    password = password + letter

for symbol in range(nr_symbols):
    symbol = random.choice(symbols)
    password = password + symbol

for number in range(nr_numbers):
    number = random.choice(numbers)
    password = password + number

password = ''.join(random.sample(password, len(password)))              #to randomize the order of letters in a string we use random.sample() takes two arguments, the string, and the number of letters, can use len(), this random method returns a list, to convert it back to a string, we use join() method, joining an empty string with our randomized string like so ''.join(random.sample()), the '' can be anything, say if we add '#' it will add # after each letter, also works with lists, tuples, and dictionaries
print(password)

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

#password generator project (her version, kinda dumb tbh)

#-----------------------------------------------(Unrandomized order)----------------------------------------------------
import random

letters = [ 'a', 'b', 'c', 'd', 'e', 'f', 'g', 'h', 'i', 'j', 'k', 'l', 'm', 'n',
            'o', 'p', 'q', 'r', 's', 't', 'u', 'v', 'w', 'x', 'y', 'z', 'A', 'B',
            'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P',
            'Q,' 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z']

numbers = ['0','1','2','3','4','5','6','7','8','9']
symbols = ['!', '#', '$', '%', '&', '(', ')', '*', '+']


print("welcome to the password Generator!")

nr_letters = int(input("how many letters? \n"))
nr_symbols = int(input("how many symbols? \n"))
nr_numbers = int(input("how many numbers? \n"))

password = ""

for char in range(1, nr_letters + 1):
    password += random.choice(letters)

for char in range(1, nr_symbols + 1):
    password += random.choice(symbols)

for char in range(nr_numbers):
    password += random.choice(numbers)

print(password)

#-----------------------------------------------(Randomized order)----------------------------------------------------

#password generator project (her version, kinda dumb tbh) (Unrandomized order)
import random

letters = [ 'a', 'b', 'c', 'd', 'e', 'f', 'g', 'h', 'i', 'j', 'k', 'l', 'm', 'n',
            'o', 'p', 'q', 'r', 's', 't', 'u', 'v', 'w', 'x', 'y', 'z', 'A', 'B',
            'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P',
            'Q,' 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z']

numbers = ['0','1','2','3','4','5','6','7','8','9']
symbols = ['!', '#', '$', '%', '&', '(', ')', '*', '+']


print("welcome to the password Generator!")

nr_letters = int(input("how many letters? \n"))
nr_symbols = int(input("how many symbols? \n"))
nr_numbers = int(input("how many numbers? \n"))

password_list = []

for char in range(1, nr_letters + 1):
    password_list += random.choice(letters)                  # can use append as well, but this also works, appending will look like so password.append(random.choice(letters))

for char in range(1, nr_symbols + 1):
    password_list += random.choice(symbols)

for char in range(nr_numbers):
    password_list += random.choice(numbers)


random.shuffle(password_list)                                # to randomize the order of the elements in the list
password = ""                                                # create a string where the loop will iterate the elements into
for char in password_list:
    password+= char                                          # adds the items from the list to the string, basically converting it

print(password)

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

# hangman Game (mostly written by me, with some hints from Angela :) )

import random

print(r'''
 _                                             
| |                                            
| |__   __ _ _ __   __ _ _ __ ___   __ _ _ __  
| '_ \ / _` | '_ \ / _` | '_ ` _ \ / _` | '_ \ 
| | | | (_| | | | | (_| | | | | | | (_| | | | |
|_| |_|\__,_|_| |_|\__, |_| |_| |_|\__,_|_| |_|
                    __/ |                      
                   |___/    

''')
word_list = ('ant baboon badger bat bear beaver camel cat clam cobra cougar '
             'coyote crow deer dog donkey duck eagle ferret fox frog goat '
             'goose hawk lion lizard llama mole monkey moose mouse mule newt '
             'otter owl panda parrot pigeon python rabbit ram rat raven '
             'rhino salmon seal shark sheep skunk sloth snake spider '
             'stork swan tiger toad trout turkey turtle weasel whale wolf '
             'wombat zebra ').split()
# r'' Means: A raw string tells Python not to interpret backslashes (\) as special characters (like \n for newline or \t for tab). Super handy for things like: ASCII art, Regular expressions, Windows file paths (r"C:\Users\Me")
hangman = [r'''
  +---+
  |   |
  O   |
 /|\  |
 / \  |
      |
=========''', r'''
  +---+
  |   |
  O   |
 /|\  |
 /    |
      |
=========''', r'''
  +---+
  |   |
  O   |
 /|\  |
      |
      |
=========''', r'''
  +---+
  |   |
  O   |
 /|   |
      |
      |
=========''', r'''
  +---+
  |   |
  O   |
  |   |
      |
      |
=========''', r'''
  +---+
  |   |
  O   |
      |
      |
      |
=========''', r'''
  +---+
  |   |
      |
      |
      |
      |
=========''']

chosen_word = random.choice(word_list)
# print(chosen_word) # for testing
display = []
max_length = 1
lives = 6
duplicate = ""

for _ in range(len(chosen_word)):
    display += "_"

end_of_game = False

while not end_of_game:
    guessed_letter = input("Guess a letter: ").lower()[
                     :max_length]                                   # [:max_length] is a way to limit the string to the number of characters you specified, basically cut it down
    duplicate += guessed_letter
    if duplicate.count(guessed_letter) >= 2:
        print(f"you've already tried {guessed_letter}")
    else:
        for position in range(len(chosen_word)):
            if guessed_letter == chosen_word[position]:
                display[position] = guessed_letter
        print(
            f"{"".join(display)}")                                  # can use the join() (that converts lists to strings) in print functions using formatted strings

        if guessed_letter not in chosen_word:

            lives -= 1
            if lives == 0:
                print(f"{hangman[lives]}")
                print("you lost")
                end_of_game = True
            else:
                print(f"{hangman[lives]}")
                print(f"Wrong!, you have {lives} lives left!")

    if "_" not in display:
        end_of_game = True
        print("You win!")

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

# Tuple (similar to list but can't be modified after it was created)#
numbers = (1, 2, 3, 3)
numbers.count(3)  # Counts the number of occurrences of an element#
numbers.index(1)  # returns the index of the first occurence of an element#
print(numbers[2]) #can access individual items similar to lists

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
# convert a tuple to a list:

list = [1, 2, 3]
tuple = tuple(list)

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

#formatted strings#
first = "john"
last = "smith"
msg = f"{first} [{last}] is a coder" # the f"" is the formatted string identifer (prefix), {} can be used to insert variables inside the string
print(msg)

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

#Math Functions#
import math #importing a module

print(math.ceil(2.9)) # ceiling of a number ,type math. and all the available methods/funcions from the math module/Library will show up
print(math.floor(2.9))
x = 2.9
print(round(x)) # Rounds the number
print(round(x, 2)) # Rounds to 2 decimal digits of precision
print(abs(-2.9)) #absolute number without a sign

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

#Paint Area Calculator

import math

def paint_cal(height, width, cover):
    cans = math.ceil((height * width) / cover)              #rounding up a number using math.ceil()
    print(f"you'll need {cans} cans for the de wall")

test_h = int(input("Height of wall: "))
test_w = int(input("Width of wall: "))
coverage = 5
paint_cal(height=test_h, width=test_w, cover=coverage)

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

#If Exercise
house = 1000000
buyer_good = False
if buyer_good:
    deduction = house * 0.1
else:
    deduction = house * 0.2
print(f"down payment is: ${deduction}")

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

# Exercice to check your input, reject it if the characters are less or more than defined, otherwise print it, and make sure to prompt again after a rejected attempt and stop the loop when the characters are within limit
while True:
    name = input("Enter your Name: ")
    if  len(name) < 3:
        print("too low")
    elif len(name) > 15:
        print("too high")
    else:
        print(f"Welcome, {name}!")
        break

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

# exercise: Guessing Game, 3 tries
attempt = 1
while attempt <= 3: # While can have an else statement
    guess = int(input("Guess Number: "))
    if guess == 7:
        print("Correct!")
        break
    else:
        if attempt < 3:
            print("try Again")
        else:
            print("You Lost :(")
    attempt = attempt + 1

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

#Excersice: Car Game, Mosh's Version, Dumber and harder to understand, very stupid and unintutive#
started = False
while True:
    Entry = input("> ").lower() #can use .lower after input() function to make the characters lowercase#
    if Entry == "help":
            print('''start - to start the car
stop - to stop the car
quit - to exit''')
    elif Entry == "start":
        if started: # it is checking if started is True here, but it isn't so this part is skipped
          print("Car Already Started Man!")
        else:
            started = True
            print('Car Started... Ready to go!')
    elif Entry == "stop":
        if not started: # it is checking if started is False here, but it isn't so this part is skipped
            print("Car Stopped Dude!")
        else:
            started = False
            print("Car Stopped.")
    elif Entry =="quit":
        break
    else:
        print("I don't undertand that...")

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

#Excersice: Car Game, my version, Superior and 10 thousand times better and makes sense :)#
started = False
while True:
    Entry = input("> ").lower() #can use .lower after input() function to make the characters lowercase#
    if Entry == "help":
            print('''start - to start the car
stop - to stop the car
quit - to exit''')
    elif Entry == "start":
        if started == False:
            started = True
            print('Car Started... Ready to go!')
        else:
            print("Car Already Started Man!")
    elif Entry == "stop":
        if started == True:
            started = False
            print("Car Stopped.")
        else:
            print("Car Stopped Dude!")
    elif Entry =="quit":
        break
    else:
        print("I don't undertand that...")

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
#love Calculator exersice lmao (teacher is a girl, ofc xD) (my version)

you = input("what is your name? ") .lower()
them = input("what is their name? ") .lower()

t_you = you.count("t")                              #count() method
r_you = you.count("r")
u_you = you.count("u")
e_you = you.count("e")
l_you = you.count("l")
o_you = you.count("o")
v_you = you.count("v")
e_you = you.count("e")


t_them = them.count("t")
r_them = them.count("r")
u_them = them.count("u")
e_them = them.count("e")
l_them = them.count("l")
o_them = them.count("o")
v_them = them.count("v")
e_them = them.count("e")


first_word = t_you + t_them + r_you + r_them + u_you + u_them + e_you + e_them
second_word = l_you + l_them + o_you + o_them + v_you + v_them + e_you + e_them
result = f"{first_word}{second_word}"

if result <= "10" or result >= "90":
    print(f"coke mentos {result}%")

elif result >= "40" and result < "50":
    print(f"alright {result}%")

else:
    print(f"your score is {result}%")

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

#Love Calculator, Angela's code, smarter, combined the strings to write less code

name1 = input("what is your name? \n")
name2 = input("what is their name? \n")

combined_string = name1 + name2
lower_case_string = combined_string.lower()

t = lower_case_string.count("t")
r = lower_case_string.count("r")
u = lower_case_string.count("u")
e = lower_case_string.count("e")

true = t + r + u + e

l = lower_case_string.count("l")
o = lower_case_string.count("o")
v = lower_case_string.count("v")
e = lower_case_string.count("e")

love = l + o + v + e

love_score = int(str(true) + str(love))                     #better than my code because it converted the result to int instead of using strings in the comparison

if (love_score < 10) or (love_score > 90):
    print(f"Your love score is {love_score}, coke mentos")


elif (love_score >= 40) and (love_score <= 50):
    print(f"Your score is {love_score}, alright")

else:
    print(f"Your score is {love_score}")

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

#nested Loops
for x in range (4):
    for y in range(4): #when x = 0, the y loop will start executing all its iterations until range 4, then the y loop ends and the x loop continues when x = 1 and y loop executes all iterations again and hands it back to x and so on.
        print(f"({x}, {y})")

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

#Challenge: Convert a List of numbers to characters displaying the shape of an "F" in nested loops

#My Code
numbers = [5, 2, 5, 2, 2]
From = 0
to = 1

for x in numbers:
    for y in numbers[From:to]: ## first iteration will be numbers[0:1] will display the first element only in a sublist
        y = x * "x"
        From = From + 1
        to = to + 1
        print(y)

#-------------------------------------------------------------------------

# Teacher's code, wow
numbers = [5, 2, 5, 2, 2]
for x in numbers: #this will iterate 5 times, changing the value of x to 5, then 2, then 5, then 2, then 2, where x will affect the amount of inner loops each outer loop has
    output = "" # this is to start off the output with an empty string and to reset it before a new innerloop starts
    for y in range(x): # Code is executed (x = 5) times the first iteration, then 2, then 5, then 2, then 2
        output = output + "x" #each iteration adds 1 "x", so for the first outer loop when x = 5, 5 inner loops will execute, adding 1 "x" each time,i inner loops: 1st: x, 2nd: xx, 3rd: xxx, 4th: xxxx, 5th: xxxxx, then exits
    print(output) #putting the print inside the inner loop will print all the inner loops iterations, putting it outside the loop will only print the last result from the inner loop: output = xxxxx

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
# Find the Maximum number from a list
# my code (Terrible and embaressing, his is much better
numbers = [1, 2, 5, 7, 3]
maximum = 0

for x in numbers:
    for y in numbers:
        if x > y:
            maximum = x
            for z in numbers:
                    if z > maximum:
                        maximum = z

print(f"The biggest number is: {maximum}")

#-------------------------------------------------------------------------
#Teacher's Code, typed my own after the hint of making the first number as the maximum then comparing with it

numbers = [4, 3, 5, 7, 3]
maximum = numbers[0]

for x in numbers:
    if x > maximum:
        maximum = x

print(f"The biggest number is: {maximum}")


#-------------------------------------------------------------------------
#another maximum example, this time i typed it my own :D thanks to thonny

student_scores = input("Input a list of student scores: ") .split()
for n in range(0, len(student_scores)):
    student_scores[n] = int(student_scores[n])
print(student_scores)

maximum = 0

for score in student_scores:
    if score > maximum:
        maximum = score


print(maximum)

#important functions to learn max() and min() do the same as the codes above easily

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
# 2D Lists or Nested lists

matrix = [
    [1, 2, 3],
    [4, 5, 6],
    [7, 8, 9]
]

print(matrix[0][2]) #this is how to access an item, with two brackets one for the list index and another for the sublist index
matrix[0][0] = 20 #can change certain items as so

for row in matrix:
    for item in row:
        print(item) #the nested loop here will iterate all the items in the 2D list

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
#Exercise to remove duplicates from a list
#my code
numbers = [7, 21 ,3 ,4, 6 ,5, 6]


for x in numbers:
    if numbers.count(x) == 2:
        numbers.remove(x)
print(numbers)
#-------------------------------------------------------------------------
#The teacher's code
#Kinda weird tbh, i't not removing from the list, it's making a different list for uniques
numbers = [7, 21 ,3, 3, 4, 6 ,5, 6]
uniques = []

for number in numbers:
    if number not in uniques:
        uniques.append(number)
print(uniques)

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
#Unpacking
coordinates = (1, 2, 3)

x = coordinates[0]
y = coordinates[1]
z = coordinates[2]

x, y, z = coordinates # this line does the same thing as the 3 lines above with much less code, python inepreter, will get the first item in the tupple and assign it to x, then the second item to y, the the third to z and so on, UNPACKNG the tupple to those 3 variables, works with lists too
print(z)

#NOTE: Good Example of this in line 2309

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

#dictionaries (key => value pairs, equivalent of Assosicative Arrays in php)
customer = {
    "name": "John Smith", #keys can't be duplicated, meaning we can't have two keys called "name" for example
    "age": 30,
    "is_verified": True #keys can be strings or numbers, values can be anything, strings, numbers, booleans, lists etc.
}

print(customer["name"]) #to access an item in the dictionary, we type the key in square brackets and the result is the value, will return an error if the key doesn't exist
print(customer.get("age")) #using the get() method works the same way as above, but will return None, if the referenced key doesn't exist
print(customer.get("birthdate", "default Value")) # can also assign a default value so it will return that instead of None
customer["name"] = "hamodie" #can update a value in the dictionary as so
customer["birthdate"] = "12-3-1992" #can update a value in the dictionary as so, or adding a key and a value
print(customer["name"])
print(customer)

#how to loop through the dictionary (only returns keys)
for key in customer:
    print(key)

#how to loop through the dictionary (only returns values)
for value in customer:
    print(customer.get(value))

#how to loop through the dictionary (keys and values)
for key in customer:
    print(key)
    print(customer[key])                # or print(customer.get(key)) both work


#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
#dictionaries exersice (very Fucking hard, i failed) Teacher's Code
phone = (input("Phone: "))
numbers = {
    "1": "One",
    "2": "Two",
    "3": "Three",
    "4": "four"
}

output = ""
for chars in phone:                                 # will iterate over the characters typed (1234)
    output += numbers.get(chars, "!") + "  "        # if the input was 1234, in the first iteration, chars is 1, so numbers.get(1) will bring the value associated with the key 1, and add it to our output, then a space + "  " and the second iteraion will do the same and add it to our output result ing in "one two" the "!" is if the key wasn't found replace it with !
print(output)

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

#Grading Program (with dictionaries) (my code)


student_scores = \                                                      # I didn't know i can put the brackets on the next line by putting \ wow
{
    "Harry" : 81,
    "Ron" : 78,
    "Hermione" : 99,
    "Draco": 74,
    "Neville" : 62
}

student_grades = {}

for student in student_scores:
    student_grades[student] = student_scores[student]                   #a way to transfer all keys and values from one dictionary to the other empty one
    if 91 < student_grades[student] <= 100:
        student_grades[student] = "Outstanding"
    elif 81 <= student_grades[student] <= 90:
        student_grades[student] = "Exceeds Expectations"
    elif 71 <= student_grades[student] <= 80:
        student_grades[student] = "Acceptable"
    elif student_grades[student] < 70:
        student_grades[student] = "Fail"



print(student_scores)
print(student_grades)


#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
#Grading Program (with dictionaries) (her code)

student_scores = \
{
    "Harry" : 81,
    "Ron" : 78,
    "Hermione" : 99,
    "Draco": 74,
    "Neville" : 62
}

student_grades = {}

for student in student_scores:
    score = student_scores[student]                                     # we assign the key value to a variable so that we can compare it in the if statements
    if 91 < score <= 100:
        student_grades[student] = "Outstanding"                         # here we're assigning both the key and value for the new dictionary, Syntax: student_grades[key] = "value"
    elif 81 <= score <= 90:
        student_grades[student] = "Exceeds Expectations"
    elif 71 <= score <= 80:
        student_grades[student] = "Acceptable"
    elif score < 70:
        student_grades[student] = "Fail"


print(student_scores)
print(student_grades)

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

# Dictionary Nesting (Dic in dic & dic in list) :)

travel_log = {
    "France": {"cites_visited": ["Paris","Lille", "Dijon"],             #a dic contains a dic that contains a list and a dic
               "total_visits": 12 },

    "Germany": {"cites_visited": ["Berlin", "Hamburg", "stuttgart"],
                "total_visits": 5}
}


travel_loge = [                                                         # a list containing dics that contains lists,
    {
      "country": "France",
      "cites_visited": ["Paris","Lille", "Dijon"],
      "total_visits": 12
    },

    {

      "country":"Germany",
      "cites_visited": ["Berlin", "Hamburg", "stuttgart"],
      "total_visits": 5
    }
]

print(travel_log)

print(travel_loge)

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

# Nesting Exersice to add to a nested dictionary in a list using a function

travel_log = [                                                         # a list containing dics that contains lists,
    {
      "country": "France",
      "visits": 12,
      "cites": ["Paris","Lille", "Dijon"]

    },

    {

      "country":"Germany",
      "visits": 5,
      "cites": ["Berlin", "Hamburg", "stuttgart"],
    }
]


def add_new_country(country_visited, times_visited, cities_visited):
    new_country = {}
    new_country["country"] = country_visited                                                    # can also be written like so: new_country = {"country": country_visited, "visits": times_visited, "cities": cities_visited} a million times better
    new_country["visits"] = times_visited
    new_country["cities"] = cities_visited
    travel_log.append(new_country)                                                              # add the dictionary to the list

add_new_country("Russia", 2, ["Moscow", "SaintPetersburg"])
print(travel_log)

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

#Secret Auction Exersice (my code, not secret because there's no function to clear the terminal in IDE)

resume = True
auction = {}
winner = 0

while resume:
    name = input("name: ")
    bid = int(input("bid: "))
    auction[name] = bid
    print(auction)
    stop = input("stop?: ")
    if stop == "yes":
        resume = False

    for key in auction:
        if auction[name] > winner:
            winner = auction[name]


for key in auction:
    if auction[key] == winner:
        winner = key

print(f"{winner} wins")

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

#Secret Auction Exersice (her code, better than mine because it used a function where mine is direct, but i got all of the idas right, just need better orginzation) (also not secret)

#from replit import clear                   # this only works on her stupid interactive course

#from art import logo                       # not gonna bother with this shit, it's just art


bids = {}
bidding_finished = False


def find_highest_bidder(bidding_record):
    highest_bid = 0
    winner = ""
    for bidder in bidding_record:
        bid_amount = bidding_record[bidder]                             # she stored the value in a variable so that it uses it  in the comparision instead of the dictionary index iteslf (smart)
        if bid_amount > highest_bid:
            highest_bid = bid_amount                                    # now we can assign the value of the dictionary in a different variable (an integer) without facing a problem, because using the dictionary index here instead of the variable highest_bid will change the value which is not what we want (very fucking smart, i think)
            winner = bidder                                             # winner here is string (smart) yet another variable, so we made 2, one string for the key and one integer for the value ( my stupid code used one variable "winner" and i used two for loops to solve the problem (code was longer and harder to understand)
    print(f"the winner is {winner} with a bid of ${highest_bid}")


while not bidding_finished:
    name = input("what is your name?: ")
    price = int(input("bid: "))
    bids[name] = price
    # print(bids)                               printing it for testing
    should_continue = input("Are there any other bidders? type 'yes' or 'no'. ")
    if should_continue == "no":
        find_highest_bidder(bids)
        bidding_finished = True
    # elif should_continue == "yes":                                      #both of these shit lines don't work outside pf the interactive course
    #     clear()


#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

#blackjack Game (all my code no functions, works but not perfect)

import random

cards = [11, 2, 3, 4, 5, 6, 7, 8, 9, 10 ,10 ,10 ,10]




game_started = True


while game_started:

    your_cards = []
    dealer_cards = []

    your_total = 0
    dealer_total = 0

    start_game = input("press 'y' if you wanna play, 'n' to quit: ").lower()

    if start_game == "y":
        your_cards.append(random.choice(cards))
        dealer_cards.append(random.choice(cards))

        print(
f"""Starting deck:
Your cards: {your_cards}
the dealer's cards: {dealer_cards}""")
        input("press enter to draw a card... ")

        your_cards.append(random.choice(cards))
        dealer_cards.append(random.choice(cards))

        print(
f"""First Draw:
Your cards: {your_cards}
the dealer's cards: {dealer_cards}""")

        your_total = 0
        dealer_total = 0

        for card in your_cards:
            your_total += card

        for card in dealer_cards:
            dealer_total += card


        if your_total < 17:
             input("you have less than 17, press enter to draw another card...")
             your_cards.append(random.choice(cards))
             print(
f"""You drew:
Your cards: {your_cards}
the dealer's cards: {dealer_cards}""")

        if dealer_total < 17:
                dealer_cards.append(random.choice(cards))
                print(
f"""Dealer drew:
Your cards: {your_cards}
the dealer's cards: {dealer_cards}""")

        if len(your_cards) == 2:
                draw_pass = input("draw 'd' or pass 'p' ?")
                if draw_pass == "d":
                    your_cards.append(random.choice(cards))
                    print(
f"""You drew second draw:
Your cards: {your_cards}
the dealer's cards: {dealer_cards}""")


        your_total = 0
        dealer_total = 0

        for card in your_cards:
            your_total += card

        for card in dealer_cards:
            dealer_total += card

        if dealer_total < your_total <= 21 and dealer_total not in [21, 20, 19]:           #[Thanks ChatGPT] can use lists to say a variable does not equal any of the numbers in the list, this line is equivalent to: dealer_total != 19 and dealer_total != 20 and dealer_total != 21
                dealer_cards.append(random.choice(cards))
                print(
f"""Dealer Drew a second draw:
Your cards: {your_cards}
the dealer's cards: {dealer_cards}""")

        your_total = 0
        dealer_total = 0

        for card in your_cards:
            your_total += card

        for card in dealer_cards:
            dealer_total += card


        if your_total > 21:
            for card in range(0, len(your_cards) - 1):
                if your_cards[card] == 11:
                    your_cards[card] = 1
            print(
f"""you got an Ace:
Your cards: {your_cards}
the dealer's cards: {dealer_cards}""")


        if dealer_total > 21:
                for card in range(0, len(dealer_cards) - 1):
                    if dealer_cards[card] == 11:
                        dealer_cards[card] = 1
                print(
f"""Dealer got an Ace:
Your cards: {your_cards}
the dealer's cards: {dealer_cards}""")

        your_total = 0
        dealer_total = 0

        for card in your_cards:
            your_total += card

        for card in dealer_cards:
            dealer_total += card

        if your_total < 17:
            input("you have less than 17, press enter to draw another card...")
            your_cards.append(random.choice(cards))
            print(
f"""You drew:
Your cards: {your_cards}
the dealer's cards: {dealer_cards}""")

        if dealer_total < 17:
            dealer_cards.append(random.choice(cards))
            print(
f"""Dealer drew:
Your cards: {your_cards}
the dealer's cards: {dealer_cards}""")

        if dealer_total > 21 or your_total > 21:
            if your_total > 21 and dealer_total > 21:
                print("you both lose")
            elif your_total > 21:
                print("you went over 21, you lose")
            elif dealer_total > 21:
                print("dealer went over 21, you win")

        if dealer_total < your_total < 21:
            print("you win")

        elif your_total < dealer_total < 21:
            print("dealer wins")

        elif your_total == dealer_total:
            print("draw")

    else:
        break


#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
#Blackjack Game (her Version, a billion times better than mine)

import random


def deal_card():
    """Returns a random card from the deck."""
    cards = [11, 2, 3, 4, 5, 6, 7, 8, 9, 10 ,10 ,10 ,10]
    card = random.choice(cards)
    return card


def calculate_score(cards):
    """Take a list of cards and return the score calculated from the cards"""

    if sum(cards) == 21 and len(cards) == 2:        # if a player gets a black jack (2 cards only, 10 and Ace)
        return 0                                    #score is 0 if a blackjack, for some reason


    if 11 in cards and sum(cards) > 21:             # to replace the ace when over 21
        cards.remove(11)                            #remove() will look for the first occurrence of an element and removes it
        cards.append(1)

    return sum(cards)


def compare(user_score, computer_score):
    if user_score == computer_score:
        return "Draw"
    elif computer_score == 0:
        return "Lose, opponent has Blackjack"
    elif user_score == 0:
        return "win with a Blackjack"
    elif user_score > 21:
        return "You went over. You lose"
    elif computer_score > 21:
        return "Opponent went over. You win"
    elif user_score > computer_score:                               #this gets executed last, so having a blackjack, 0 won't be affected because it gets executed first and exists the if else statements
        return "You win"
    else:
        return "You lose"

def play_game():
    user_cards = []
    computer_cards = []
    is_game_over = False

    for _ in range(2):
        user_cards.append(deal_card())
        computer_cards.append(deal_card())


    while not is_game_over:

        user_score = calculate_score(user_cards)
        computer_score = calculate_score(computer_cards)
        print(f"Your cards: {user_cards}, Current score: {user_score}")
        print(f"the computer's first card: {computer_cards[0]}")


        if user_score == 0 or computer_score == 0 or user_score > 21:
            is_game_over = True
        else:
            user_should_deal = input("type 'y' to get another card, type 'n' to pass: ")
            if user_should_deal == "y".lower():
                user_cards.append(deal_card())
            else:
                is_game_over = True


    while computer_score != 0 and computer_score < 17:
        computer_cards.append(deal_card())
        computer_score = calculate_score(computer_cards)

    print(f"Your Final Hand is: {user_cards}, Final score: {user_score}")
    print(f"the computer's final hand: {computer_cards}, final score {computer_score}")
    print(compare(user_score, computer_score))


while input("Do you want to play a game of Blackjack? Type 'y' or 'n': ") == "y".lower():
    play_game()

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
#Emticon-Emoji Converter
message = input(">")
words = message.split(" ") #split function splits the string by a delimiter, here it's " ", and stores them in a list
emojis = {
    ":)": "🙂",
    ":(": "🙁"
}
output = ""
for word in words:
    output += emojis.get(word, word) + " " # if "word" has a key value pair then bring its value, if not, return the word unchanged to the output

print(output)

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
#Functions, Parameters and Arguments
def greet_user(first_name, second_name): #this is how to define a function, this part of code won't execute unless called, between the parenthisis are parameters, which are placeholders for recieving information, local variable inside of the function, can have multiple parameter-arguments
    print(f"Hello {first_name} {second_name}")
    print("data science")

print("Start")
greet_user("John", "Anthony") #calling the function here, "john" here is an argument, that will be assigned to the parameter
greet_user("mary", "Jane") #when a function has parameter, it's obligatory to pass a value (argument) for that parameter, can have multiple parameter-arguments

#arguments above are like "John" positional arguments their values can change based on their positions, there are also keyword arguments, like so (first_name = "John") their postision doesn't matter
greet_user(first_name="Matt", second_name="Murdok") #keyword Argument, when mixing different arguments, positional arguments must always come before keyword arguments, starting with keyword arguments is not allowed

print("End")

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

#return statement
def square(number):
    return number * number #returns values to the callers of the function
    #print(number * number) # all functions return the value 'None' by Default when not using a return statement, this line will get these results 9, None, due to having two print functions in line 6 resulting like so: print(print(number * number)), None is printed because of the outer print function
result = square(3) #this function now returns a value (result of 3 * 3) similar to how input function works
print(result)

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
# function with inputs and output (return or multiple return statements)

def format_name(f_name, l_name):
    if f_name == "" or l_name == "":
        return "You didn't provide valid inputs"
    f_name = f_name.capitalize()                         # method to capitalize first letter and makes the rest small letter
    l_name = l_name.title()                              # another method useful to capitalize the first letter of each word in a sentence
    # return f_name, l_name                              # this returns a tuple for the strings (parameters) in their last known assignment so for the two lines above to be returned, they have to be assigned to the parameters, because strings are immutable, and the methods above don't affect the original strings, they create new ones
    return f"{f_name} {l_name}"                          # can also return strings (wtf I didn't know)

result = format_name(input("what is your first name? " ), input("what is your last name? "))
# print(type(result)) to check the type of the output
print(result)

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
#Leap year, Days in a month, Nested Functions, and multiple returns Exersice (All my code, I'm Awesome)

def is_leap(year):
    """Checks if the year is a leap year, this is a docstring,
    type it after function declaration and it will show when you hover over the function,
    useful for documenting your functions,
    Awesome stuff"""

    if year % 4 == 0:                               #chatGPT is so fucking smart, it made a much shorter version for this function than mine: return year % 4 == 0 and (year % 100 != 0 or year % 400 == 0)
        if year % 100 == 0:
            if year % 400 == 0:
                return True
            else:
                return False
        else:
            return True
    else:
        return False

def days_in_month(year_, month_):
    month_days = [31, 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]
    if is_leap(year_):
        month_days[1] = 29
    month_ = month_days[month_ - 1]
    return month_


year = int(input("enter a year: "))
month = int(input("Enter a month: "))
days = days_in_month(year, month)
print(days)

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

#calcuator (my code, a million times better than hers, hers is fucking trash and reta- never mind, hers is better) (super fucking easy, I finished it in the first 10 seconds of the video GG EZ :D )

num1 = int(input("first number: "))
num2 = int(input("second number: "))

def add(n1, n2):
    return n1 + n2

def subtract(n1, n2,):
    return n1 - n2

def multiply(n1, n2):
    return n1 * n2

def divide(n1, n2):
    return n1 / n2


operations = {
    "+" : add(num1, num2),
    "-" : subtract(num1, num2),
    "*" : multiply(num1, num2),
    "/" : divide(num1, num2)
}

for op in operations:
    print(op)

operation = input("enter the operation: ")

print(f"{num1} {operation} {num2} = {operations[operation]}")

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

#calcuator (her code, very weird and unnecessarily complex) (better, more modular, more flexiable) actually it's a million times better, mine's trash lol, if i needed to operate on the result with a third number, i can't with mine, because num1, num2 are called in the dictionary itself


def add(n1, n2):
    return n1 + n2

def subtract(n1, n2,):
    return n1 - n2

def multiply(n1, n2):
    return n1 * n2

def divide(n1, n2):
    return n1 / n2


operations = {
    "+" : add,                      # I didn't know you can type the function without the braces and arguments WTF!!!
    "-" : subtract,                 # this called Function storing, so that it can be called later, delayed call, much more efficent than calling them all directly, and more flexiible
    "*" : multiply,                 # this is commonly refered to as: Deferring work (lazy evaluation) thanks to ChatGPT for the explanation
    "/" : divide
}

num1 = int(input("first number: "))

for op in operations:
    print(op)

operation_symbol = input("enter the operation: ")

num2 = int(input("second number: "))

calculation_function = operations[operation_symbol]             #we stored the function in a variable, (we're just accessing the values from the dictionary, super convoluted stuff, but wow)
answer = calculation_function(num1, num2)                       #this is basically like a dynamic function now, we defined the arguments here and took the function name from a variable we stored it in the line above, wow actually.

print(f"{num1} {operation_symbol} {num2} = {answer}")


#-----------------------------------------------------------------------------------------------------------------------
# 2nd Version of hers (adding a third number to the result)

def add(n1, n2):
    return n1 + n2

def subtract(n1, n2,):
    return n1 - n2

def multiply(n1, n2):
    return n1 * n2

def divide(n1, n2):
    return n1 / n2


operations = {
    "+" : add,                      # I didn't know you can type the function without the braces and arguments WTF!!!
    "-" : subtract,                 # this called Function storing, so that it can be called later, delayed call, much more efficent than calling them all directly, and more flexiible
    "*" : multiply,                 # this is commonly referred to as: Deferring work (lazy evaluation) thanks to ChatGPT for the explanation
    "/" : divide
}

is_calculating = True


num1 = int(input("first number: "))

for op in operations:
    print(op)

operation_symbol = input("enter the operation: ")

num2 = int(input("type a number: "))
calculation_function = operations[operation_symbol]                         #we stored the function in a variable, (we're just accessing the values from the dictionary, super convoluted stuff, but wow)
first_answer = calculation_function(num1, num2)                             #this is basically like a dynamic function now, we defined the arguments here and took the function name from a variable we stored it in the line above, wow actually.

print(f"{num1} {operation_symbol} {num2} = {first_answer}")

operation_symbol = input("enter another operation: ")
num3 = int(input("another number: "))

calculation_function = operations[operation_symbol]
second_answer = calculation_function(first_answer, num3)

print(f"{first_answer} {operation_symbol} {num3} = {second_answer}")

#-----------------------------------------------------------------------------------------------------------------------
# 3rd version of hers (infinte operations with the option to stop)

def add(n1, n2):
    return n1 + n2

def subtract(n1, n2,):
    return n1 - n2

def multiply(n1, n2):
    return n1 * n2

def divide(n1, n2):
    return n1 / n2


operations = {
    "+" : add,                      # I didn't know you can type the function without the braces and arguments WTF!!!
    "-" : subtract,                 # this called Function storing, so that it can be called later, delayed call, much more efficent than calling them all directly, and more flexiible
    "*" : multiply,                 # this is commonly referred to as: Deferring work (lazy evaluation) thanks to ChatGPT for the explanation
    "/" : divide
}


num1 = int(input("first number: "))
for op in operations:
    print(op)
is_calculating = True

while is_calculating:
    operation_symbol = input("enter the operation: ")

    num2 = int(input("type a number: "))
    calculation_function = operations[operation_symbol]
    answer = calculation_function(num1, num2)

    print(f"{num1} {operation_symbol} {num2} = {answer}")


    if input("press 'y' to continue and 'n' to stop: ") == "y":
        num1 = answer
    else:
        is_calculating = False

#-----------------------------------------------------------------------------------------------------------------------
# 4th version, with recursion :)

def add(n1, n2):
    return n1 + n2

def subtract(n1, n2,):
    return n1 - n2

def multiply(n1, n2):
    return n1 * n2

def divide(n1, n2):
    return n1 / n2


operations = {
    "+" : add,                      # I didn't know you can type the function without the braces and arguments WTF!!!
    "-" : subtract,                 # this called Function storing, so that it can be called later, delayed call, much more efficent than calling them all directly, and more flexiible
    "*" : multiply,                 # this is commonly referred to as: Deferring work (lazy evaluation) thanks to ChatGPT for the explanation
    "/" : divide
}

def calculator():
    num1 = float(input("first number: "))
    for op in operations:
        print(op)
    is_calculating = True


    while is_calculating:
        operation_symbol = input("enter the operation: ")

        num2 = float(input("type a number: "))
        calculation_function = operations[operation_symbol]
        answer = calculation_function(num1, num2)

        print(f"{num1} {operation_symbol} {num2} = {answer}")

        continue_ = input("press 'y' to continue and 'n' to start a new calculation: ")
        if continue_ == "y":
            num1 = answer
        else:
            calculator()                                            #this is recursion, a function that calls itself to repeat the code

calculator()

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

#Creating a reusable function
def Emoji_converter(message): #this also works without using a parameter, but in that case the function will be less reusable due to it relying on an exteranl variable `message`
    words = message.split(" ")
    emojis = {
        ":)": "🙂",
        ":(": "☹️"
    }

    output = ""
    for word in words:
        output += emojis.get(word,word) + " "
    return output #we need two blank lines after a function definition


message = input(">") #not to inlcude the input because it can be changed to take input from a GUI instead of the terminal, making it unreusable

print(Emoji_converter(message)) #not to include in the function because the output can be treated differently depending on where we'll use it, can work without the argument `message` but that will make the function depend on a global variable

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

#Prime Number Checker function exersice (my code)

def prime_checker(number):
    if number % number == 0 and number % 1 == 0:
        if number == 0 or number == 1:
            print("not a prime")
        elif number > 1:
            if number != 2 and number % 2 == 0:
                print("not a prime")

            elif number != 3 and number % 3 == 0:
                print("not a prime")
            else:
                print("prime")


n = int(input("check number: "))
prime_checker(number=n)

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

#Prime Number Checker function exersice (her code, it's better but it's Absolutely retarded, cuz it was WRONG, i fixed it :) )

def prime_checker(number):
    is_prime = True
    if number == 1 or number == 2:                      # the fix, cuz the range loop below won't check 1 and 2 and will output them as prime which they aren''t
        is_prime = False
    for i in range (2, number):                         # this is basically taking the number you entered, and divides it to all the numbers above than 1 and below than the number itself, and if any of them can divide the number, then it's not a prime
        if number % i == 0:
            is_prime = False
    if is_prime:
        print("Prime number")
    else:
        print("Not a Prime")


n = int(input("check number: "))
prime_checker(number=n)

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

# Caesar Cipher (my version with Angela's help, her version is better and shorter)

alphabet = [ 'a', 'b', 'c', 'd', 'e', 'f', 'g', 'h', 'i', 'j', 'k', 'l', 'm', 'n',
            'o', 'p', 'q', 'r', 's', 't', 'u', 'v', 'w', 'x', 'y', 'z']

def caeser(start_text, shift_amount, cypher_direction):                          # this encodes and decodes at the same time
    cypher_text = ""
    for letter in start_text:
        if letter in alphabet:
            position = alphabet.index(letter)                                        # to get the index number from the list
            new_position = 0
            if cypher_direction == "encode".lower():
                new_position = position + shift_amount
                if new_position > 25:
                    new_position -= 26                                               # this is to prevent an index number out of bound of the list size, like any word with letter z if shifted, it will cause an error, my solution, her dumb solution was to just duplicate the alphabet letters, which was my first thought :), also no need to do it for the decode, since indexes in the minus will loop back anyway
            elif cypher_direction == "decode".lower():
                new_position = position - shift_amount
            new_letter = alphabet[new_position]
            cypher_text += new_letter
        else:
            cypher_text += letter
    print(cypher_text)


start = True

while start:
    direction = input("type 'encode' or 'decode'\n")
    text = input("Message\n").lower()
    shift = int(input("shift:\n"))
    shift = shift % 26

    caeser(start_text=text, shift_amount=shift, cypher_direction=direction)         # you can call use variables as arguments! WTF i didn't know that
    choice = input("continue? 'Yes' 'No'")
    if choice == "no".lower():
        start = False

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

# Caesar Cipher ( Angela's Version)

from art import logo

print(logo)

alphabet = [ 'a', 'b', 'c', 'd', 'e', 'f', 'g', 'h', 'i', 'j', 'k', 'l', 'm', 'n',
            'o', 'p', 'q', 'r', 's', 't', 'u', 'v', 'w', 'x', 'y', 'z',
             'a', 'b', 'c', 'd', 'e', 'f', 'g', 'h', 'i', 'j', 'k', 'l', 'm', 'n',
             'o', 'p', 'q', 'r', 's', 't', 'u', 'v', 'w', 'x', 'y', 'z']                    # I did her dumb solution :)


def caeser(start_text, shift_amount, cypher_direction):
    end_text = ""
    if cypher_direction == "decode".lower():
        shift_amount *= -1                                                                  # to make it position + ( - Shift) for decoding on the commented line below
    for char in start_text:
        if char in alphabet:
            position = alphabet.index(char)
            new_position = position + shift_amount                                          # here
            end_text += alphabet[new_position]
        else:
            end_text += char                                                                # this is for any character not in the alphabet list (symbols, etc) will be added as it is
    print(F"Here's the {direction}d result {end_text}")



start = True

while start:
    direction = input("type 'encode' or 'decode'\n")
    text = input("Message\n").lower()
    shift = int(input("shift:\n"))
    shift = shift % 26                                                                      # if the shift was out of bounds of our alphabet list, using the modulus like so will divide any number larger than 26 by 26 (the number of items in the list) as many times, until the final point where it can't be divided cleanly and end up with a remainder, that remainder will be the correct shift number within the bounds of the list, any number smaller than 26 is not affected, because modulus will return that number if it's smaller than the modulus number


    caeser(start_text=text, shift_amount=shift, cypher_direction=direction)
    choice = input("continue? 'Yes' 'No'")
    if choice == "no".lower():
        start = False


#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
# Namespaces, Local vs Global Scope


enemies = 1

def increase_enemies():
    enemies = 2                                         # this creates an entirely new variable that is separate from the global variable
    print(f"enemies inside function: {enemies}")        #local Scope

increase_enemies()
print(f"enemies outside function: {enemies}")           #Global Scope, this will actually print 1, not 2 because it can't see the variable inside the function


#NOTES:
# variables in if statements, while loops, for loops are global, so defining a variable in them will be global

#-----------------------------------------------------------------------------------------------------------------------
#how to moodify a global variable (bad idea)

enemies = 1

def increase_enemies():
    global enemies                          #this is accessing the global variable whose value was 1
    enemies += 1
    print(f"enemies inside function: {enemies}")

increase_enemies()
print(f"enemies outside function: {enemies}")

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
#a better way to increase the value of a global variable

enemies = 1

def increase_enemies():
    print(f"enemies inside function: {enemies}")
    return  enemies + 1

increase_enemies()
print(f"enemies outside function: {enemies}")


#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

# Python Constants

#there are no constants, SIKE :)
#naming convention to differenciate constants from variables is to type them all in UPPPERCASE
PI = 3.14159
URL = "https://potato.com"
TWITTER_HANDLE = "@3armota"


#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
#Advanced Number Guessing Game with Functions (this was ALL ME WOW :DD )

from random import randint


def guess_checker(chosen_guess, number):
    if chosen_guess == number:
        print(f"Correct!, the number is {number}")
        return ""
    elif chosen_guess > number:
        return "too high"
    elif chosen_guess < number:
        return "too low"


def guess_counter(attempts):
    correct_number = randint(1, 100)
    game_ends = False
    while not game_ends:
        print(f"You have {attempts} attempts")
        player_guess = int(input("Make a guess: "))
        result = guess_checker(player_guess, correct_number)
        print(result)
        attempts -= 1
        if attempts == 0:
            print(f"you lost :( the correct number was: {correct_number}")
            game_ends = True
        elif result == "":
            game_ends = True


def game_start():
    print("Welcome to the Number Game!")
    print("thinking of a number between 1 and 100")
    difficulty = input("choose difficulty: ")
    if difficulty == "easy":
        player_attempts = 10
        guess_counter(player_attempts)

    elif difficulty == "hard":
        player_attempts = 5
        guess_counter(player_attempts)

    restart = input("start over? 'y' 'n': ")
    if restart == 'y':
        game_start()
    else:
        return


game_start()


#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

#Advanced Number Guessing Game with Functions (her version, also pretty good actually)


from random import randint

EASY_LEVEL_TURNS = 10                               #Constant
HARD_LEVEL_TURNS = 5

turns = 0

def check_answer(guess, answer, turns):
    """Checks answer against guess. Returns the number of turns remaining"""
    if guess > answer:
        print("too high")
        return turns - 1
    elif guess < answer:
        print("too low")
        return turns - 1
    else:
        print(f"You got it!, the answer was {answer}.")


def set_difficulty():
    level = input("choose a difficulty. Type 'easy' or 'hard': ")
    if level == "easy":
        return EASY_LEVEL_TURNS
    else:
        return HARD_LEVEL_TURNS


def game():
    print("Welcome to the Number Game!")
    print("thinking of a number between 1 and 100")
    answer = randint(1, 100)
    turns = (set_difficulty())
    guess = 0


    while guess != answer:
        print(f"You have {turns} attempts remaining to guess the number.")
        guess = int(input("Make a guess: "))
        turns = check_answer(guess, answer, turns)
        if turns == 0:
            print("you've run out of guesses, you lose.")
            return
        elif guess != answer:
            print("Guess again.")

game()


#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
#higher lower game (my version, All me)

from random import randint

from art import higher_lower, vs
from higherlower import data


def follower_diff(a, b):
    if a > b:
        return a
    elif b > a:
        return b
    else:
        return "Oops, they are one and the same :)"


def answer_checker(answer, bigger):
    if answer == bigger:
        print("you win")
        print(f"your score is {score + 1}")
        return True
    else:
        print(f"you lost, your score is {score}")
        return False


def score_counter():
    if game_continue:
        return score + 1


def name_fetcher():
    if bigger == a:
        return f"{random_person['name']} is the correct answer"
    else:
        return f"{random_person_2['name']} is the correct answer"

def person_randomizer():
    return data[randint(0, len(data) - 1)]


def name_definer(person):
    return f"{person['name']} a {person['description']} from {person['country']}"

def response_checker(answer):
    global bigger, game_continue, score
    bigger = follower_diff(a, b)
    game_continue = answer_checker(answer, bigger)
    score = score_counter()
    print(name_fetcher())

print(higher_lower)


bigger = ""
score = 0
game_continue = True

random_person = person_randomizer()
random_person_2 = person_randomizer()

while game_continue:
    random_person = random_person_2
    random_person_2 = person_randomizer()
    if random_person == random_person_2:
        continue
    a = random_person['followers']
    b = random_person_2['followers']
    a_name = random_person['name']

    print("Compare: A " + name_definer(random_person))
    print(vs)
    print(f"Against: B " + name_definer(random_person_2))


    while True:
        user_input = input("who has more followers? Type 'A' or 'B'")
        if user_input.lower() == 'a':
            answer = a
            response_checker(answer)
            break
        elif user_input.lower() == 'b':
            answer = b
            response_checker(answer)
            break
        else:
            print("please enter 'A' or 'B'")
            continue


#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
#ChatGPT Version ( i Liked his multi value return fucntion answer_checker() aka Unpacking in lines, 2251, 2263, 2266, 2309, the rest is irrelevent, the next code block is better than this

from random import choice
from art import higher_lower, vs
from higherlower import data

def follower_diff(count_a, count_b):
    """
    Compare two follower counts and return:
      - "A" if count_a > count_b
      - "B" if count_b > count_a
      - "tie" if they are equal
    """
    if count_a > count_b:
        return "A"
    elif count_b > count_a:
        return "B"
    else:
        return "tie"

def answer_checker(user_choice, correct_side, current_score):                       #this Function
    """
    Print win/lose message, update score,
    and return (did_win: bool, new_score: int).
    """
    if correct_side == "tie":
        print("They have exactly the same follower count! No points awarded.")
        return True, current_score  # Let user continue
    elif user_choice.upper() == correct_side:
        new_score = current_score + 1
        print("You win this round!")
        print(f"Your score is now {new_score}")
        return True, new_score                                                      #here
    else:
        print(f"You lost. Your final score was {current_score}")
        return False, current_score                                                 #and Here

def name_definer(person):
    """
    Return a string like "John Doe a singer from USA"
    """
    return f"{person['name']}, a {person['description']} from {person['country']}"

def person_randomizer():
    """Pick one random entry from the data list."""
    return choice(data)

def play_higher_lower():
    print(higher_lower)
    score = 0
    game_continue = True

    # Start by picking two distinct people
    person_a = person_randomizer()
    person_b = person_randomizer()
    while person_b == person_a:
        person_b = person_randomizer()

    while game_continue:
        a_followers = person_a["followers"]
        b_followers = person_b["followers"]

        # Show the two options
        print(f"Compare A: {name_definer(person_a)}")
        print(vs)
        print(f"Against B: {name_definer(person_b)}")

        # Get the user's guess
        guess = ""
        while guess not in ("a", "b"):
            guess = input("Who has more followers? Type 'A' or 'B': ").lower()

        # Determine which side truly has more followers
        correct_side = follower_diff(a_followers, b_followers)

        # Check the answer, update score, and see if we continue
        game_continue, score = answer_checker(guess, correct_side, score)                   #and here, the way he returned all the values to their repective variables using Unpacking
        if not game_continue:
            break

        # Announce who was correct
        if correct_side == "A":
            print(f"{person_a['name']} was correct with {a_followers} followers!")
        elif correct_side == "B":
            print(f"{person_b['name']} was correct with {b_followers} followers!")
        # If it was a tie, we already handled it in answer_checker

        # Prepare for the next round: B becomes the new A, and pick a new B
        person_a = person_b
        person_b = person_randomizer()
        while person_b == person_a:
            person_b = person_randomizer()

if __name__ == "__main__":
    play_higher_lower()


#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
#higher lower game (her version, Trillion Times better than mine + Shorter)

from art import higher_lower, vs
from higherlower import data
import random




def format_data(account):
    """Takes the account data and returns the printable format."""
    account_name = account["name"]
    account_descr = account["description"]
    account_country = account["country"]
    return f"{account_name}, a {account_descr}, from {account_country}"


def check_answer(guess, a_followers, b_followers):
    """Take user guess and follower counts and returns if they got it right."""
    if a_followers > b_followers:
        return guess == "a"                                 #Big Brain Move (if a>b, guess == "a" will return True if you guessed correctly and False if you didn't, same goes with a<b
    else:
        return guess == "b"

print(higher_lower)
score = 0
game_should_continue = True
account_b = random.choice(data)

while game_should_continue:
    account_a = account_b
    account_b = random.choice(data)
    while account_a == account_b:
        account_b = random.choice(data)

    print(f"Compare A: {format_data(account_a)}")
    print(vs)
    print(f"Against B: {format_data(account_b)}")

    guess = input("who has more followers? Type 'A' or 'B': ").lower()

    a_follower_count = account_a["followers"]
    b_follower_count = account_b["followers"]

    is_correct = check_answer(guess, a_follower_count, b_follower_count)

    if is_correct:
        score += 1
        print(f"You're right! current score: {score}.")
    else:
        game_should_continue = False
        print(f"Sorry, that's wrong. Final Score: {score}")

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

# Coffee Machine (my code)

MENU = {
    "espresso": {
        "ingredients": {
            "water": 50,
            "milk": 0,
            "coffee": 18,
        },
        "cost": 1.5,
    },
    "latte": {
        "ingredients": {
            "water": 200,
            "milk": 150,
            "coffee": 24,
        },
        "cost": 2.5,
    },
    "cappuccino": {
        "ingredients": {
            "water": 250,
            "milk": 100,
            "coffee": 24,
        },
        "cost": 3.0,
    }
}

resources = {
    "water": 400,
    "milk": 200,
    "coffee": 100,
    "money" : 0.00
}


def resource_checker(selected_item_f):
    water_f = MENU[selected_item_f]["ingredients"]["water"]
    milk_f = MENU[selected_item_f]["ingredients"]["milk"]
    coffee_f = MENU[selected_item_f]["ingredients"]["coffee"]

    if water_f > resources["water"]:
        print("Sorry, you don't have enough Water")
        return False, water_f, milk_f, coffee_f
    if milk_f > resources["milk"]:
        print("Sorry, you don't have enough Milk")
        return False, water_f, milk_f, coffee_f
    if coffee_f > resources["coffee"]:
        print("Sorry, you don't have enough Coffee")
        return False, water_f, milk_f, coffee_f
    return True, water_f, milk_f, coffee_f


def dispenser(water_f, milk_f, coffee_f):
    resources["water"] = resources["water"] - water_f
    resources["milk"] = resources["milk"] - milk_f
    resources["coffee"] = resources["coffee"] - coffee_f


def process_coins():
    print("insert Coins")
    quarters = float(input("How many quarters? "))
    dimes = float(input("How many dimes? "))
    nickles = float(input("How many nickles? "))
    pennies = float(input("How many pennies? "))
    total = quarters * 0.25 + dimes * 0.10 + nickles * 0.05 + pennies * 0.01

    if total < MENU[selected_item]["cost"]:
        print("Sorry, Not enough Cash, refunding")
    elif total == MENU[selected_item]["cost"]:
        resources["money"] = resources["money"] + total
        print("enjoy your drink!")
    else:
        remainder = round(total - MENU[selected_item]["cost"], 2)
        total = round(total - remainder, 2)
        resources["money"] = resources["money"] + total
        print(f"enjoy your drink!, here's the Change ${remainder}")


is_on = True

while is_on:
    print("Welcome to the Coffee Machine")
    print(f"1: Espresso ${MENU["espresso"]["cost"]}")
    print(f"2: Latte ${MENU["latte"]["cost"]}")
    print(f"3: cappuccino ${MENU["cappuccino"]["cost"]}")
    print("report: print a report")
    print("off: Turn off")
    request = input("What would you like? ")



    if request == "1":
        selected_item = "espresso"
        can_order, water, milk, coffee = resource_checker(selected_item)
        if can_order:
            process_coins()
            dispenser(water, milk, coffee)

    elif request == "2":
        selected_item = "latte"
        can_order, water, milk, coffee = resource_checker(selected_item)
        if can_order:
            process_coins()
            dispenser(water, milk, coffee)

    elif request == "3":
        selected_item = "cappuccino"
        can_order, water, milk, coffee = resource_checker(selected_item)
        if can_order:
            process_coins()
            dispenser(water, milk, coffee)

    elif request.lower() == "report":
        print(f"current resources")
        print(f"Water: {resources["water"]}ml")
        print(f"Milk: {resources["milk"]}ml")
        print(f"Coffee: {resources["coffee"]}g")
        print(f"Money: ${resources["money"]}\n")
        input("Press any button to continue\n")

    elif request.lower() == "off":
        is_on = False
        machine_status = input("machine turned off, to turn back on type ON: ")
        if machine_status.lower() == "on":
            is_on = True

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

#Coffee Machine (her code)

MENU = {
    "espresso": {
        "ingredients": {
            "water": 50,
            "milk": 0,
            "coffee": 18,
        },
        "cost": 1.5,
    },
    "latte": {
        "ingredients": {
            "water": 200,
            "milk": 150,
            "coffee": 24,
        },
        "cost": 2.5,
    },
    "cappuccino": {
        "ingredients": {
            "water": 250,
            "milk": 100,
            "coffee": 24,
        },
        "cost": 3.0,
    }
}

profit = 0
resources = {
    "water": 400,
    "milk": 200,
    "coffee": 100
}

def is_resource_sufficient(order_ingredients):
    """Returns True when order can be made, False if ingredients are insufficient"""
    for item in order_ingredients:
        if order_ingredients[item] >= resources[item]:
            print(f"Sorry there is not enough {item}.")
            return False
    return True

def process_coins():
    """Returns the total calculated from coins inserted."""
    print("Please insert coins. ")
    total = int(input("How many quarters?: ")) * 0.25
    total += int(input("How many dimes?: ")) * 0.1                   #Mind = Blown
    total += int(input("How many nickles?: ")) * 0.05
    total += int(input("How many pennies?: ")) * 0.01
    return total

def is_transaction_successful(money_received, drink_cost):
    """Return True if the payment is accepted, or False if the money is insufficient"""
    if money_received >= drink_cost:
        change = round(money_received - drink_cost, 2)
        print(f"here is ${change} in change.")
        global profit
        profit += drink_cost
        return True
    else:
        print("Sorry that's not enough money. Money refunded.")
        return False


def make_coffee(drink_name, order_ingredients):
    """Deduct the required ingredients from the resources."""
    for item in order_ingredients:
        resources[item] -= order_ingredients[item]
    print(f"Here is your {drink_name} ☕")

is_on = True

while is_on:
    choice = input("what would you like? (espresso/latte/cappuccino): ")
    if choice == "off":
        is_on = False
    elif choice == "report":
        print(f"Water: {resources['water']}ml")
        print(f"Milk: {resources['milk']}ml")
        print(f"Coffee: {resources['coffee']}g")
        print(f"Money: ${profit}")
    else:
        drink = MENU[choice]
        if is_resource_sufficient(drink["ingredients"]):
            payment = process_coins()
            if is_transaction_successful(payment, drink["cost"]):
                make_coffee(choice, drink["ingredients"])



#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

#Exception Handling

#the exit code when a program runs Successfully: Process finished with exit code 0
#the exit code when a program encounters an error: Process finished with exit code 1

try:
    age = int(input("age: "))
    income = 20000
    risk = income / age
    print(age)
except ZeroDivisionError:
    print("Age cannot be zero") #can add multiple except statements to handle multiple exceptions in our code
except ValueError:
    print("Invalid input")

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

# Raise Exceptions

# Example 1

try:
    a_dictionary = {"key": "value"}
    print(a_dictionary["sdogj"])
except KeyError as error_message:
    print(f"The key {error_message} does not exist.")
finally:
    raise KeyError("this an error that I made up.")                              # raise will manually raise an exception even if there is no actual error

# Example 2

height = float(input("Height: "))
weight = int(input("Weight: "))

if height > 3:
    raise ValueError("Human Height should not be over 3 meters.")

bmi = weight / height ** 2

print(bmi)

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

# Exception Exersice

fruits = ["apple", "Pear", "Orange"]


def make_pie(index):
    try:
        fruit = fruits[index]
    except IndexError:
        print("there are only 3 fruits")
    else:
        print(fruit + " pie")

make_pie(4)

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

# NATO alphabet with exception handling

import pandas

nato_alpha = pandas.read_csv("nato_phonetic_alphabet.csv")



phonetic_alphabet_dict =  {row.letter:row.code for (index, row) in  nato_alpha.iterrows()}
print(phonetic_alphabet_dict)

while True:
    try:
        word = input("word: ").upper()
        nato_translation = [phonetic_alphabet_dict[letter] for letter in word]

    except KeyError:
        print("Only Letters Please")
    else:
        print(nato_translation)


#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

#classes

#classes are used to model real complex concepts where normal types like integers, strings, booleans, lists, dictionaries can't model, like a shopping cart for example or drawing a point.

#Naming Convertion for classes. Pascal Naming Convention, Capitlize First Letter, for multiple words we capitalize the Initals of every word instead of using _ betwwen them like in variables and functions
class Point:
    def move(self):
        print("move")

    def draw(self):
        print("draw")

#an object is an instance of a class, the class defines the blueprint or template for creating an object, and the object is an instance based on that blueprint
point1 = Point() #creating an object, typing the name of class (calling it like a function)
point1.x = 10 #these are attributes, variables that belong to a particular object
point1.y = 20
point1.draw() #can use the functions we created earlier as methods to the object
print(point1.x)

point2 = Point()
point2.x = 30 #these attributes are specific to point 2 only
point2.y = 40
print(point2.x)

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
#Example with a premade class called Turtle and how to use its methods

from turtle import Turtle, Screen

timmy = Turtle()
print(timmy)
timmy.shape("turtle")
timmy.color("DarkOrange")
timmy.forward(100)

my_screen = Screen()
print(my_screen.canvheight)
my_screen.exitonclick()

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

#moving the turtle and changing its color

from turtle import Turtle, Screen                 # import * imports everything
import turtle
import random

tim = Turtle()
tim.shape("turtle")
tim.color("red")

# Drawing a Square
for i in range(4):
    tim.forward(100)
    tim.right(90)

# Draw a dashe line
for _ in range(14):
    tim.forward(10)
    tim.penup()
    tim.forward(10)
    tim.pendown()

#------------------------------------------

# Generates a random heroic name
import heroes
print(heroes.gen())


#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
# Turtle Exersice: Draw a Triangle, Square, Pentagon, hexagon, heptagon, octagon, nonagon, and decagon :D
# (All my code, YAY)

from turtle import Turtle, Screen
import turtle
import random

tim = Turtle()
tim.shape("turtle")
tim.color("red")

for sides in range(3, 11):
    angle = 360 / sides
    turtle.colormode(255)
    tim.pencolor(random.randint(0, 255), random.randint(0, 255), random.randint(0, 255))
    for _ in range(sides):
        tim.forward(100)
        tim.right(angle)


screen = Screen()
screen.exitonclick()

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
# Turtle Exersice: Draw a Triangle, Square, Pentagon, hexagon, heptagon, octagon, nonagon, and decagon :D
# (her code, kina weird)

from turtle import Turtle, Screen                 # import * imports everything
import turtle
import random

tim = Turtle()

colours = ['dark orange', 'lawn green', 'dodger blue', 'medium violet red', 'yellow', 'medium slate blue', 'spring green', 'deep pink', 'sky blue']

def draw_shape(num_sides):
    angle = 360 / num_sides
    for _ in range(num_sides):
        tim.forward(100)
        tim.right(angle)

for shape_side_n in range(3, 11):
    tim.color(random.choice(colours))
    draw_shape(shape_side_n)


screen = Screen()
screen.exitonclick()


#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
# Turtle: Random Walk (my code, same as hers)

from random import randint
from turtle import Turtle, Screen
import turtle
import random

tim = Turtle()

colours = ['dark orange', 'lawn green', 'dodger blue', 'medium violet red', 'yellow', 'medium slate blue', 'spring green', 'deep pink', 'sky blue', 'blue', 'olive', 'indian red', 'dark green']

moves = [0, 45, 90, 135, 180, 225, 270, 360]

tim.speed(0)
tim.pensize(10)
for _ in range(10000):
    # tim.setheading(randint(0, 270))           # can use this
    tim.setheading(random.choice(moves))        # or this, but not both
    tim.forward(20)
    tim.color(random.choice(colours))


screen = Screen()
screen.exitonclick()


#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
# Turtle: Draw a Spirograph (Circles shape) (my code)

from random import randint
from turtle import Turtle, Screen
import turtle
import random

tim = Turtle()


heading = 0
tim.speed(0)
for _ in range(75):
    tim.setheading(heading)
    turtle.colormode(255)
    tim.color(random.randint(0, 255), random.randint(0, 255), random.randint(0, 255))
    heading += 5
    tim.circle(100)


screen = Screen()
screen.exitonclick()


#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
# Turtle: Draw a Spirograph (Circles shape) (her code, very nice)

from random import randint
from turtle import Turtle, Screen
import turtle
import random

tim = Turtle()
turtle.colormode(255)

def random_color():
    r = random.randint(0, 255)                                          # another way of adding random colors
    g = random.randint(0, 255)
    b = random.randint(0, 255)
    color = (r, g, b)
    return color


tim.speed(0)

def draw_spirograph(size_of_gap):
    for _ in range(int(360 / size_of_gap)):                              # real smart, making it only draw the needed circles to complete the shape, division returns float so we added int()
        tim.color(random_color())
        tim.circle(100)
        tim.setheading(tim.heading() + size_of_gap)                      # using tim.heading() to get the current heading then increase it by 10

draw_spirograph(5)

screen = Screen()
screen.exitonclick()

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
#Draw a Donut (ChatGPT code) (Super Cool)

import turtle
import random
import math

tim = turtle.Turtle()
turtle.colormode(255)
tim.speed(0)

orbit_radius = 100   # distance from center (controls the "hole" size)
circle_radius = 50   # radius of the drawn circles
angle_step = 5      # how much to rotate each step

for angle in range(0, 360, angle_step):
    # compute position on the orbit
    x = orbit_radius * math.cos(math.radians(angle))
    y = orbit_radius * math.sin(math.radians(angle))

    tim.penup()
    tim.goto(x, y)             # move turtle outward on orbit
    tim.setheading(angle)      # make turtle face outward
    tim.pendown()

    tim.color(random.randint(0, 255), random.randint(0, 255), random.randint(0, 255))
    tim.circle(circle_radius)  # draw a circle at this orbit point

screen = turtle.Screen()
screen.exitonclick()

# Code Explanation:

# https://chatgpt.com/share/68a5aee8-28c8-800b-9395-e06e36b2c74b

# Explanation Text:

# 1. The coordinate plane
# In turtle (and in general math graphics):
#   - The center (0,0) is the origin.
#   - The x-axis goes left (negative) and right (positive).
#   - The y-axis goes down (negative) and up (positive).
# We want the turtle to move in a circle around the origin at some distance orbit_radius.

# 2. The unit circle (radius = 1)
# In trigonometry, the unit circle is a circle centered at (0,0) with radius 1.
# For any angle θ (measured from the positive x-axis):
#   - cos(θ) gives the x-coordinate on that circle.
#   - sin(θ) gives the y-coordinate on that circle.

# So:
#   - At θ = 0° → (cos 0, sin 0) = (1, 0) → point on the right.
#   - At θ = 90° → (cos 90, sin 90) = (0, 1) → point on the top.
#   - At θ = 180° → (cos 180, sin 180) = (-1, 0) → left side.
#   - At θ = 270° → (cos 270, sin 270) = (0, -1) → bottom.

# This is how you can walk around a circle just by changing θ.

# 3. Scaling to any radius
# The unit circle only has radius = 1, but we want a bigger circle (orbit_radius).
# So we scale up the coordinates:
#       x = r * cos(θ),     y = r * sin(θ)

# That makes the circle radius r instead of 1.

# Example with r = 200:
#   - At θ = 0° → (200, 0)
#   - At θ = 90° → (0, 200)
#   - At θ = 180° → (-200, 0)
#   - At θ = 270° → (0, -200)

# Now you’ve got a circle of radius 200 centered at (0,0).

# 4. Why math.radians(angle)?
# In Python’s math module:
# cos() and sin() expect the angle in radians, not degrees.
# 1 radian ≈ 57.3°, and 2π radians = 360°.
# So if you want to work in degrees (like we usually do with turtle), you must convert:

math.radians(angle)   # converts degrees → radians

# Example:
# math.cos(math.radians(90)) = 0
# math.sin(math.radians(90)) = 1

# Exactly what we expect.

# 5. Putting it all together

# This line:
x = orbit_radius * math.cos(math.radians(angle))
y = orbit_radius * math.sin(math.radians(angle))

# means:
# 1. Start with an angle in degrees (like 0°, 10°, 20°…).
# 2. Convert to radians for math functions.
# 3. Compute the (x, y) coordinates on a circle of radius orbit_radius.

# Then tim.goto(x, y) moves the turtle to that position, making it orbit smoothly around the center.

# 💡 Visualization: Imagine a stick of length orbit_radius attached at the center.
# You spin it around by changing the angle. The tip of the stick traces a perfect circle,
# and cos/sin give you the tip’s x and y at every angle.

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

#Extract colors from an image with colorgram (my code)

import colorgram

colors = colorgram.extract('image.jpg', 30)

all_colors = []
for number in range(30):
    color = colors[number]
    rgb = color.rgb
    red = rgb[0]
    green = rgb[1]
    blue = rgb[2]
    colors_tuple = (red, green, blue)
    if number > 2:                              #removes the first 3 colors (white background are the most occurring so they get extracted first)
        all_colors.append(colors_tuple)

print(all_colors)

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

#Extract colors from an image with colorgram (her code)

import colorgram


rgb_colors = []
colors = colorgram.extract('image.jpg', 30)
for color in colors:
    r = color.rgb.r
    g = color.rgb.g
    b = color.rgb.b
    new_color = (r, g, b)
    rgb_colors.append(new_color)

print(rgb_colors)

#

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

# Hirst Painting (my code, NICE)

from turtle import Turtle, Screen
import turtle
import random

color_list = [(198, 13, 32), (248, 236, 25), (40, 76, 188), (244, 247, 253), (39, 216, 69), (238, 227, 5), (227, 159, 49), (29, 40, 154), (212, 76, 15), (17, 153, 17), (241, 36, 161), (195, 16, 12), (223, 21, 120), (68, 10, 31), (61, 15, 8), (223, 141, 206), (11, 97, 62), (219, 159, 11), (54, 209, 229), (19, 21, 49), (238, 157, 216), (79, 74, 212), (10, 228, 238), (73, 212, 168), (93, 233, 198), (65, 231, 239), (217, 88, 51)]

tim = Turtle()
screen = Screen()

turtle.colormode(255)
tim.hideturtle()
tim.penup()
tim.setheading(225)
tim.forward(350)
tim.speed(0)

def create_circle_row():
    for _ in range(10):
        tim.setheading(0)
        tim.color(random.choice(color_list))
        tim.begin_fill()
        tim.circle(20)
        tim.end_fill()
        tim.penup()
        tim.forward(50)


create_circle_row()

for __ in range(9):
    tim.setheading(180)
    tim.forward(500)
    tim.setheading(90)
    tim.forward(50)
    create_circle_row()



screen.exitonclick()

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

# Hirst Painting (her code, interesting)

import turtle as turtle_module
import random

turtle_module.colormode(255)
tim = turtle_module.Turtle()
color_list = [(198, 13, 32), (248, 236, 25), (40, 76, 188), (244, 247, 253), (39, 216, 69), (238, 227, 5), (227, 159, 49), (29, 40, 154), (212, 76, 15), (17, 153, 17), (241, 36, 161), (195, 16, 12), (223, 21, 120), (68, 10, 31), (61, 15, 8), (223, 141, 206), (11, 97, 62), (219, 159, 11), (54, 209, 229), (19, 21, 49), (238, 157, 216), (79, 74, 212), (10, 228, 238), (73, 212, 168), (93, 233, 198), (65, 231, 239), (217, 88, 51)]

tim.penup()
tim.hideturtle()
tim.setheading(225)
tim.forward(300)
tim.setheading(0)
number_of_dots = 100
tim.speed("fastest")

for dot_count in range(1, number_of_dots + 1):
    tim.dot(20, random.choice(color_list))
    tim.forward(50)

    if dot_count % 10 == 0:
        tim.setheading(90)
        tim.forward(50)
        tim.setheading(180)
        tim.forward(500)
        tim.setheading(0)


screen = turtle_module.Screen()
screen.exitonclick()

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

#Making the the turtle move with keys using Event lisnters and higher order functions (my code, mostly, this is also the Etch a Sketch Challeng LOL)


from turtle import Turtle, Screen

tim = Turtle()
screen = Screen()
tim.speed(0)

def forward():
    tim.forward(10)

def back():
    tim.back(10)

def left():
    tim.left(10)

def right():
    tim.right(10)

def clear():
    tim.clear()
    tim.reset()             #can also use tim.home() but it's shittier, needs penup and pendown because it paint's on the way back


screen.listen()
screen.onkeypress(forward, "Up")                #she used onkey() much worse, needs constant tapping becuse holding the button won't trigger repeated movements
screen.onkeypress(back, "Down")
screen.onkeypress(left, "Left")
screen.onkeypress(right, "Right")
screen.onkeypress(clear, "c")

screen.exitonclick()

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
# Turtle Racing Game (my code, ALL MINE, YES!) (no need for her code, mine is better anyway :) )



from turtle import Turtle, Screen
import turtle
import random

screen = Screen()
screen.setup(width=500, height=400)
user_bet = screen.textinput(title="Make your bet", prompt="Which will turtle will win the race? Enter a color:")                #for numbers, we use screen.numinput()


turtle.colormode(255)
turtle_names = ["tim", "tom", "tam", "tem", "tym", "tum", "twm"]
turtle_colors = ['red', 'green', 'blue', 'purple', 'gold', 'magenta', 'cyan']
turtles = {}
y_position = 90
number = 0

for turtle_name in turtle_names:
    turtles[turtle_name] = Turtle(shape="turtle")                               # can define the shape when we create the object
    turtles[turtle_name].penup()
    turtles[turtle_name].color(turtle_colors[number])
    number += 1
    turtles[turtle_name].goto(x=-230, y=y_position)                             # give turtle coordinates to go to
    y_position -= 30

race_is_on = True
while race_is_on:
    for turtle_name in turtle_names:                                               # can dictionary.items() to get the keys and the values of the dictionary printed
        turtles[turtle_name].forward(random.randint(3, 10))
        if turtles[turtle_name].xcor() >= 230.00:                                  # xcor() checks x Coordinates
            race_is_on = False
            winning_color = turtles[turtle_name].pencolor()                        # can use .color() or .pencolor() to reference the turtle with its color, pencolor() will return only the color string but color() will return (pencolor, fillcolor)
            if winning_color == user_bet:
                print(f"{winning_color} Wins, you win!")
            else:
                print(f"{winning_color} Wins, you lost!")

screen.exitonclick()

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
# Turtle Racing Game (her code, yes i decided to add it)

from turtle import Turtle, Screen
import random

is_race_on = False
screen = Screen()
screen.setup(width=500, height=400)
user_bet = screen.textinput(title="Make your bet", prompt="Which turtle will win the race? Enter a color: ")
colors = ["red", "orange", "yellow", "green", "blue", "purple"]
y_positions = [-70, -40, -10, 20, 50, 80]
all_turtles = []

#Create 6 turtles
for turtle_index in range(0, 6):
    new_turtle = Turtle(shape="turtle")
    new_turtle.penup()
    new_turtle.color(colors[turtle_index])
    new_turtle.goto(x=-230, y=y_positions[turtle_index])
    all_turtles.append(new_turtle)

if user_bet:
    is_race_on = True

while is_race_on:
    for turtle in all_turtles:
        #230 is 250 - half the width of the turtle.
        if turtle.xcor() > 230:
            is_race_on = False
            winning_color = turtle.pencolor()
            if winning_color == user_bet:
                print(f"You've won! The {winning_color} turtle is the winner!")
            else:
                print(f"You've lost! The {winning_color} turtle is the winner!")

        #Make each turtle move a random amount.
        rand_distance = random.randint(0, 10)
        turtle.forward(rand_distance)

screen.exitonclick()

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

# Creating a table using the prettytable Package (OOP)

from prettytable import PrettyTable

table = PrettyTable()

table.add_column("Pokemon Name", ["Pikachu","Squirtle", "Charmander"])
table.add_column("Type", ["Electric", "Water", "Fire"])

table.align = "l"

print(table)

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

#Constructors

class Point:
    def __init__(self, x, y): #the method __init__() is a constructor, it is used initialize/construct/create an object, self is a referance to the current object, x and y are the parameters whose values(arguments) will be the attributes of the object (x, y)
        self.x = x # Syntax: object.attribute = argument, this line is the same as point1.x = 10, where self will reference point1, point2, etc, x in self.x is the attribute, it is the same as the x in point1.x, and the x after the "=" is the argument of the parameters, which will be the value of the attribute at the creation of the object like so: point1 = Point(10, 20) where 10, 20 are the arguments
        self.y = y

    def move(self):
        print("move")

    def draw(self):
        print("draw")


point1 = Point(10, 20) #using constructors, attributes can be given directly when calling the class function
point1.x = 11 #can change the values of either attributes after creating the object like so
print(point1.x)

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

# Excersice, crate a Person Type and a talk method
class Person:
    def __init__(self, name, age):
        self.name = name
        self.age = age
    def talk(self):
        print(f"Hi, i'm {self.name} and i'm {self.age} years old")


john = Person("John Kenedy",20)
john.talk()
maria = Person("maria Gonzales", 18)
maria.talk()
print(john.name)

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

# Coffee Machine with OOP (my code) (her code was literally the same as mine, NICE)

from menu import Menu, MenuItem
from coffee_maker import CoffeeMaker
from money_machine import MoneyMachine


coffee_maker = CoffeeMaker()
menu = Menu()
money_machine = MoneyMachine()



is_on = True

while is_on:
    choice = input("what would you like? " + menu.get_items())
    if choice == "off":
        is_on = False
    elif choice == "report":
        coffee_maker.report()
        money_machine.report()
    else:
        drink = menu.find_drink(choice)
        if coffee_maker.is_resource_sufficient(drink):
            if money_machine.make_payment(drink.cost):
                coffee_maker.make_coffee(drink)

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

# Constructers Notes
#NOTE 1:

class User:
    ...                                 #can also use pass

user_1 = User()
user_1.id = "001"                       #you can assign attributes to the object directly, without initializing them in the class
user_1.username = "Angela"

print(user_1.username)

user_2 = User()

#-----------------------------------------------------------------------------------------------------------------------
#NOTE 2:

class User:
    def __init__(self):                                 #special function to initialize attributes, it executes everytime an object is created
        print("new user being created...")

user_1 = User()
user_1.id = "001"                                       #creating attributes to the object manually, (without a constructor) i.e. without initializing them in the class
user_1.username = "Angela"

print(user_1.username)

user_2 = User()

#-----------------------------------------------------------------------------------------------------------------------
#NOTE 3:

class User:
    def __init__(self, user_id, username):
        self.id = user_id                                       # self.attribute = parameter, the convention is to have the same name for the attribute and the parameter, but it's possible for them to have different names
        self.username = username
        self.followers = 0                                      # provide a default value, by setting it to a value instead of a parameter, it means you don't have to pass in a parameter when creating an object (you don't have to provide the follower count at the time of the creation)


user_1 = User("001", "Angela")                                  #creating attributes (using a constructor), this is much better with less code
user_2 = User("002", "Jack")


print(user_1.followers)

#-----------------------------------------------------------------------------------------------------------------------
#NOTE 3:

#Adding methods to a class

class User:
    def __init__(self, user_id, username):
        self.id = user_id                                       # self.attribute = parameter, the convention is to have the same name for the attribute and the parameter, but it's possible for them to have different names
        self.username = username
        self.followers = 0                                      # provide a default value, by setting it to a value instead of a parameter, it means you don't have to pass in a parameter when creating an object (you don't have to provide the follower count at the time of the creation)
        self.following = 0

    def follow(self, user):
        user.followers += 1
        self.following =+ 1


user_1 = User("001", "Angela")                  #creating attributes (using a constructor), this is much better with less code
user_2 = User("002", "Jack")


user_1.follow(user_2)
user_2.follow(user_1)

print(f"User 1 Followers: {user_1.followers}")
print(f"User 1 Following: {user_1.following}")
print(f"User 2 Followers: {user_2.followers}")
print(f"User 2 Following: {user_2.following}")


#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

# Quiz Project OOP (mostly the same between me and hers)
#----------------------------------------------File: data.py------------------------------------------------------------

question_data = [
{"text": "A slug's blood is green.", "answer": "True"},
{"text": "The loudest animal is the African Elephant.", "answer": "False"},
{"text": "Approximately one quarter of human bones are in the feet.", "answer": "True"},
{"text": "The total surface area of a human lungs is the size of a football pitch.", "answer": "True"},
{"text": "In West Virginia, USA, if you accidentally hit an animal with your car, "
         "you are free to take it home to eat.", "answer": "True"},
{"text": "In London, UK, if you happen to die in the House of Parliament, "
         "you are entitled to a state funeral.", "answer": "False"},
{"text": "It is illegal to pee in the Ocean in Portugal.", "answer": "True"},
{"text": "You can lead a cow down stairs but not up stairs.", "answer": "False"},
{"text": "Google was originally called 'Backrub'.", "answer": "True"},
{"text": "Buzz Aldrin's mother's maiden name was 'Moon'.", "answer": "True"},
{"text": "No piece of square dry paper can be folded in half more than 7 times.", "answer": "False"},
{"text": "A few ounces of chocolate can to kill a small dog.", "answer": "True"}
]

#----------------------------------------------File: main.py------------------------------------------------------------

from question_model import Question
from data import question_data
from quiz_brain import QuizBrain

question_bank = []

for question in question_data:
    question_text = question["text"]
    question_answer = question["answer"]
    new_question = Question(question_text, question_answer)
    question_bank.append(new_question)



quiz = QuizBrain(question_bank)
while quiz.still_has_questions():
    quiz.next_question()

print("You've completed the quiz")
print(f"Your final score was: {quiz.score}/{quiz.question_number}")

#----------------------------------------------File: question_model-----------------------------------------------------

class Question:
    def __init__(self, q_text, q_answer):
        self.text = q_text
        self.answer = q_answer

#----------------------------------------------File: quiz_brain---------------------------------------------------------

class QuizBrain:
    def __init__(self, q_list):
        self.question_number = 0
        self.questions_list = q_list
        self.score = 0


    def still_has_questions(self):
        return self.question_number < len(self.questions_list)       #mind blown (this replaces if then True else then False)


    def next_question(self):
        question = self.questions_list[self.question_number]
        self.question_number += 1
        user_answer = input(f"Q.{self.question_number}: {question.text} (True, False): ")
        self.check_answer(user_answer, question.answer)


    def check_answer(self, user_answer, correct_answer):
        if user_answer.lower() == correct_answer.lower():
            print("You got it right!")
            self.score += 1
        else:
            print("That's wrong.")
        print(f" the correct answer was: {correct_answer}")
        print(f"Your current score is {self.score}/{self.question_number}\n")

#-----------------------------------------------------------------------------------------------------------------------
#Notes (Differences between my main.py and hers at the begining):

#mine (complicated)

from question_model import Question
from data import question_data

question_bank = []

for tick in range(0,len(question_data)-1):
    sub_dic = question_data[tick]
    for key in sub_dic:
        text = sub_dic[key]
        answer = key
        question = Question(text,answer)
        question_bank.append(question)

print(question_bank[0].answer)

#-----------------------------------
#hers (much better)

from question_model import Question
from data import question_data

question_bank = []

for question in question_data:
    question_text = question["text"]
    question_answer = question["answer"]
    new_question = Question(question_text, question_answer)
    question_bank.append(new_question)

print(question_bank[0].answer)


#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

#inheritance
# it is mechanism for reusing code

class Mammal:
    def walk(self):
        print("walk")


class Dog(Mammal): #Dog is a child class of the parent class Mammal, meaning it will inherit its methods like the walk() method
    pass #pass will allow you to have an empty class when you have no specific method to add to the child class


class Cat(Mammal):
    def meo(self):
        print("Meow!")


cat1 = Cat()
dog1 = Dog()
cat1.meo()

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

# inheritnace constructors

class Animal:
    def __init__(self):
        self.num_eyes = 2

    def breathe(self):
        print("Inhale, exhale.")



class Fish(Animal):
    def __init__(self):
        super().__init__()                      # this will initialize the parent methods and attributes from the parent class
    def swim(self):
        print("moving in water.")

    def breathe(self):
        super().breathe()                       # this means we're going to do everything that the super class breathe method does and more in the child class
        print("doing this underwater.")

nemo = Fish()
nemo.swim()
nemo.breathe()                                  # now child class objects can access parent class methods and attributes
print(nemo.num_eyes)

juno = Animal()                                 # parent breathe method unaffected
juno.breathe()

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

# The Snake Game (OOP + Inheritence) (my code)

#------------------------------------------File: main.py----------------------------------------------------------------

from turtle import Screen
import time
from scoreboard import Score
from food import Food
from snake import Snake

screen = Screen()
screen.setup(width=600, height=600)
screen.bgcolor("black")
screen.title("Snake Game")

screen.tracer(0)
snake = Snake()
food = Food()
score = Score()

screen.listen()
screen.onkeypress(snake.left, "Up")
screen.onkeypress(snake.right, "Down")
screen.onkeypress(snake.left, "Left")
screen.onkeypress(snake.right, "Right")
screen.onkeypress(snake.clear, "c")


game_is_on = True
while game_is_on:
    screen.update()
    time.sleep(0.1)
    snake.move()
    # print(snake.segments[0].pos())

    if food.distance(snake.segments[0]) <= 20:
        food.change_location()
        snake.extend()
        score.count_score()

    if (snake.segments[0].xcor() > 300 or snake.segments[0].xcor() < -300
        or snake.segments[0].ycor() > 300 or snake.segments[0].ycor() < -300):
        game_is_on = False
        score.game_over()
        print("you lost :(")


    for segment in snake.segments[1:]:                          # using a sublist/slice instead of doing a for loop over all of the original list and excluding the first segment with an if statement, much shorter now
        if snake.segments[0].distance(segment) < 10:
            game_is_on = False
            score.game_over()
            print("you lost :(")

screen.exitonclick()

#------------------------------------------File: snake.py---------------------------------------------------------------

from turtle import Turtle

MOVE_DISTANCE = 20

class Snake():
    def __init__(self):
        self.tails_names = ["tail_1", "tail_2", "tail_3"]
        self.tails = {}
        self.x_position = 0
        self.segments = []

        for tail in self.tails_names:
            self.add_segment(tail)


    def add_segment(self, tail):
        self.tails[tail] = Turtle(shape="square")
        self.tails[tail].color("white")
        self.tails[tail].penup()
        self.tails[tail].goto(self.x_position, y=0)
        self.x_position -= 20
        self.segments.append(self.tails[tail])

    def extend(self):
        self.add_segment(self.segments[-1].position())                     # -1 here means the last segment, in lists, -1 index is the last item in the list


    def move(self):
        for seg_num in range(len(self.segments) -1, 0, -1):
            self.segments[seg_num].goto((self.segments[seg_num - 1].pos()))
        self.segments[0].forward(MOVE_DISTANCE)


    def left(self):
        self.segments[0].left(90)

    def right(self):
        self.segments[0].right(90)

    def clear(self):
        self.segments[0].clear()

#------------------------------------------File: scoreboard.py----------------------------------------------------------

from turtle import Turtle
FONT = ("", 13, "")                 #constant, just a variable written in all caps, there are no constants in python :)

class Score(Turtle):
    def __init__(self):
        super().__init__()
        self.hideturtle()
        self.penup()
        self.color("white")
        self.score = 0
        self.goto(-40, 280)
        self.score_update()

    def score_update(self):
        self.write(f"Score: {self.score}", font= FONT)

    def game_over(self):
        self.goto(-40,0)
        self.write("GAME OVER", font=FONT)

    def count_score(self):
        self.clear()
        self.score += 1
        self.score_update()

#------------------------------------------File: food.py----------------------------------------------------------------

from turtle import Turtle
import random

class Food(Turtle):

    def __init__(self):
        super().__init__()
        self.shape("circle")
        self.color("blue")
        self.penup()
        self.shapesize(stretch_len=0.5, stretch_wid=0.5)
        self.speed("fastest")
        self.change_location()


    def change_location(self):
        self.goto(random.randint(-280, 280), random.randint(-280, 280))

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

# The Snake Game (OOP + Inheritence) (her code)
# many differences, but the same functionality, can't tell which is better, mine or hers :)

#------------------------------------------File: main.py----------------------------------------------------------------

from turtle import Screen
from snake import Snake
from food import Food
from scoreboard import Scoreboard
import time

screen = Screen()
screen.setup(width=600, height=600)
screen.bgcolor("black")
screen.title("My Snake Game")
screen.tracer(0)

snake = Snake()
food = Food()
scoreboard = Scoreboard()

screen.listen()
screen.onkey(snake.up, "Up")
screen.onkey(snake.down, "Down")
screen.onkey(snake.left, "Left")
screen.onkey(snake.right, "Right")

game_is_on = True
while game_is_on:
    screen.update()
    time.sleep(0.1)
    snake.move()

    #Detect collision with food.
    if snake.head.distance(food) < 15:
        food.refresh()
        snake.extend()
        scoreboard.increase_score()

    #Detect collision with wall.
    if snake.head.xcor() > 280 or snake.head.xcor() < -280 or snake.head.ycor() > 280 or snake.head.ycor() < -280:
        game_is_on = False
        scoreboard.game_over()

    #Detect collision with tail.
    for segment in snake.segments:
        if segment == snake.head:
            pass
        elif snake.head.distance(segment) < 10:
            game_is_on = False
            scoreboard.game_over()


screen.exitonclick()


#------------------------------------------File: snake.py---------------------------------------------------------------

from turtle import Turtle
STARTING_POSITIONS = [(0, 0), (-20, 0), (-40, 0)]
MOVE_DISTANCE = 20
UP = 90
DOWN = 270
LEFT = 180
RIGHT = 0


class Snake:

    def __init__(self):
        self.segments = []
        self.create_snake()
        self.head = self.segments[0]

    def create_snake(self):
        for position in STARTING_POSITIONS:
            self.add_segment(position)

    def add_segment(self, position):
        new_segment = Turtle("square")
        new_segment.color("white")
        new_segment.penup()
        new_segment.goto(position)
        self.segments.append(new_segment)

    def extend(self):
        self.add_segment(self.segments[-1].position())

    def move(self):
        for seg_num in range(len(self.segments) - 1, 0, -1):
            new_x = self.segments[seg_num - 1].xcor()
            new_y = self.segments[seg_num - 1].ycor()
            self.segments[seg_num].goto(new_x, new_y)
        self.head.forward(MOVE_DISTANCE)

    def up(self):
        if self.head.heading() != DOWN:
            self.head.setheading(UP)

    def down(self):
        if self.head.heading() != UP:
            self.head.setheading(DOWN)

    def left(self):
        if self.head.heading() != RIGHT:
            self.head.setheading(LEFT)

    def right(self):
        if self.head.heading() != LEFT:
            self.head.setheading(RIGHT)


#------------------------------------------File: scoreboard.py----------------------------------------------------------

from turtle import Turtle
ALIGNMENT = "center"
FONT = ("Courier", 24, "normal")


class Scoreboard(Turtle):

    def __init__(self):
        super().__init__()
        self.score = 0
        self.color("white")
        self.penup()
        self.goto(0, 270)
        self.hideturtle()
        self.update_scoreboard()

    def update_scoreboard(self):
        self.write(f"Score: {self.score}", align=ALIGNMENT, font=FONT)

    def game_over(self):
        self.goto(0, 0)
        self.write("GAME OVER", align=ALIGNMENT, font=FONT)

    def increase_score(self):
        self.score += 1
        self.clear()
        self.update_scoreboard()


#------------------------------------------File: food.py----------------------------------------------------------------

from turtle import Turtle
import random


class Food(Turtle):

    def __init__(self):
        super().__init__()
        self.shape("circle")
        self.penup()
        self.shapesize(stretch_len=0.5, stretch_wid=0.5)
        self.color("blue")
        self.speed("fastest")
        self.refresh()

    def refresh(self):
        random_x = random.randint(-280, 280)
        random_y = random.randint(-280, 280)
        self.goto(random_x, random_y)

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

# Pong Game (my code)
# OOP, Inheritence, lambda
# Ball Speeding

#------------------------------------------File: main.py----------------------------------------------------------------

import time
from turtle import Screen, Turtle

from ball import Ball
from scoreboard import Scoreboard
from paddle import Paddle



screen = Screen()
screen.setup(1200, 800)
screen.bgcolor('black')
screen.tracer(0)

paddle1 = Paddle()
paddle1.paddle_creation(580, 0)

paddle2 = Paddle()
paddle2.paddle_creation(-590, 0)


keys = {"Up": False, "Down": False, "w": False, "s": False}         #keys are being used as toggles, they will be changed to True when the key is pressed, and False when not pressed

def key_press(key):
    keys[key] = True                                                # this is setting the key "up" for example to True, allowing the movement to happen when the value is Evaluated to be True in the main loop


def key_release(key):
    keys[key] = False                                               # this is setting the key "up" for example to False, stopping the movement to happen when the value is Evaluated to be False in the main loop


screen.listen()

screen.onkeypress(lambda: key_press("Up"), "Up")                # lambda in Python creates a tiny unnamed function on the fly. It lets you pass arguments into functions for event handlers, without calling them immediately.
screen.onkeyrelease(lambda: key_release("Up"), "Up")            # what lambda is doing here is basically: “When the Up key is pressed, run this little function that calls key_press("Up").” Python doesn’t execute it right away — it just saves that tiny function to be triggered later.

screen.onkeypress(lambda: key_press("Down"), "Down")            # without lambda for example: ```screen.onkeypress(key_press("Up"), "Up")``` this calls the function immediately when Python reads the line, instead of waiting for the key press. That’s not what we want.
screen.onkeyrelease(lambda: key_release("Down"), "Down")

screen.onkeypress(lambda: key_press("w"), "w")
screen.onkeyrelease(lambda: key_release("w"), "w")

screen.onkeypress(lambda: key_press("s"), "s")
screen.onkeyrelease(lambda: key_release("s"), "s")

scoreboard = Scoreboard()
ball = Ball()

line = Turtle()
line.color("white")
line.speed(0)
line.penup()
line.hideturtle()
line.goto(0, 400)
line.setheading(-90)
for _ in range(40):
    line.pendown()
    line.forward(10)
    line.penup()
    line.forward(10)

game_on = True
while game_on:
    screen.update()
    time.sleep(0.02)
    ball.move()

    if ball.ycor() > 380 or ball.ycor() < -380:
        ball.bounce_y()

    if ball.xcor() > 550 and ball.distance(paddle1.paddles[0]) < 100:
        ball.bounce_x()

    if ball.xcor() < -550 and ball.distance(paddle2.paddles[0]) < 100:
        ball.bounce_x()

    if ball.xcor() > 600:
        ball.goto(0,0)
        ball.reset_ball()
        scoreboard.left_point()

    if ball.xcor() < -600:
        ball.goto(0,0)
        ball.reset_ball()
        scoreboard.right_point()

    if keys["Up"]:                              # here when the key is evaluated to be True, "after it was changed by the key_press function" it will trigger the movement", and once it gets False "after it was changed by the key_release function" it will stop the movement.
        paddle1.up()
    if keys["Down"]:
        paddle1.down()

    if keys["w"]:
        paddle2.up()
    if keys["s"]:
        paddle2.down()

screen.exitonclick()

#------------------------------------------File: paddle.py--------------------------------------------------------------

from turtle import Turtle
import turtle



class Paddle():
    def __init__(self):
        self.paddles = []
        ...


    def paddle_creation(self, x_loc, y_loc):
        paddle = Turtle(shape="square")
        paddle.color("white")
        paddle.penup()
        paddle.shapesize(2,8)
        paddle.setheading(90)
        paddle.speed(0)
        paddle.goto(x_loc, y_loc)
        self.paddles.append(paddle)


    def up(self):
        self.paddles[0].setheading(90)
        self.paddles[0].forward(20)

    def down(self):
        self.paddles[0].setheading(-90)
        self.paddles[0].forward(20)


#------------------------------------------File: ball.py----------------------------------------------------------------

from turtle import Turtle

class Ball(Turtle):
    def __init__(self):
        super().__init__()
        self.pendown()
        self.shape("circle")
        self.goto(0,0)
        self.color("white")
        self.shapesize(2)
        self.penup()
        self.x_move = 10
        self.y_move = 10


    def move(self):
        new_x = self.xcor() + self.x_move
        new_y = self.ycor() + self.y_move
        self.goto(new_x, new_y)

    def bounce_x(self):
        self.x_move *= -1
        self.x_move *= 1.1
        self.y_move *= 1.1

    def bounce_y(self):
        self.y_move *= -1

    def reset_ball(self):
        self.bounce_x()
        self.x_move = 10
        self.y_move = 10

#------------------------------------------File: scoreboard.py----------------------------------------------------------

from turtle import Turtle

class Scoreboard(Turtle):
    def __init__(self):
        super().__init__()
        self.color("white")
        self.penup()
        self.hideturtle()
        self.goto(0, 350)
        self.score_left = 0
        self.score_right = 0


    def update_score(self):
        self.clear()
        self.write(f"{self.score_left}        {self.score_right}", align="center", font=("Courier", 36, "normal"))


    def left_point(self):
        self.score_left += 1
        self.update_score()


    def right_point(self):
        self.score_right += 1
        self.update_score()

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

# Pong Game (her Code)
# Took ball speeding idea from her (but mine is better)

#------------------------------------------File: main.py----------------------------------------------------------------

from turtle import Screen, Turtle
from paddle import Paddle
from ball import Ball
from scoreboard import Scoreboard
import time

screen = Screen()
screen.bgcolor("black")
screen.setup(width=800, height=600)
screen.title("Pong")
screen.tracer(0)

r_paddle = Paddle((350, 0))
l_paddle = Paddle((-350, 0))
ball = Ball()
scoreboard = Scoreboard()

screen.listen()
screen.onkey(r_paddle.go_up, "Up")
screen.onkey(r_paddle.go_down, "Down")
screen.onkey(l_paddle.go_up, "w")
screen.onkey(l_paddle.go_down, "s")

game_is_on = True
while game_is_on:
    time.sleep(ball.move_speed)
    screen.update()
    ball.move()

    #Detect collision with wall
    if ball.ycor() > 280 or ball.ycor() < -280:
        ball.bounce_y()

    #Detect collision with paddle
    if ball.distance(r_paddle) < 50 and ball.xcor() > 320 or ball.distance(l_paddle) < 50 and ball.xcor() < -320:
        ball.bounce_x()

    #Detect R paddle misses
    if ball.xcor() > 380:
        ball.reset_position()
        scoreboard.l_point()

    #Detect L paddle misses:
    if ball.xcor() < -380:
        ball.reset_position()
        scoreboard.r_point()

screen.exitonclick()

#------------------------------------------File: paddle.py--------------------------------------------------------------

from turtle import Turtle


class Paddle(Turtle):

    def __init__(self, position):
        super().__init__()
        self.shape("square")
        self.color("white")
        self.shapesize(stretch_wid=5, stretch_len=1)
        self.penup()
        self.goto(position)

    def go_up(self):
        new_y = self.ycor() + 20
        self.goto(self.xcor(), new_y)

    def go_down(self):
        new_y = self.ycor() - 20
        self.goto(self.xcor(), new_y)


#------------------------------------------File: ball.py----------------------------------------------------------------

from turtle import Turtle


class Ball(Turtle):

    def __init__(self):
        super().__init__()
        self.color("white")
        self.shape("circle")
        self.penup()
        self.x_move = 3
        self.y_move = 3
        self.move_speed = 0.1

    def move(self):
        new_x = self.xcor() + self.x_move
        new_y = self.ycor() + self.y_move
        self.goto(new_x, new_y)

    def bounce_y(self):
        self.y_move *= -1

    def bounce_x(self):
        self.x_move *= -1
        self.move_speed *= 0.9

    def reset_position(self):
        self.goto(0, 0)
        self.move_speed = 0.1
        self.bounce_x()


#------------------------------------------File: scoreboard.py----------------------------------------------------------

from turtle import Turtle


class Scoreboard(Turtle):

    def __init__(self):
        super().__init__()
        self.color("white")
        self.penup()
        self.hideturtle()
        self.l_score = 0
        self.r_score = 0
        self.update_scoreboard()

    def update_scoreboard(self):
        self.clear()
        self.goto(-100, 200)
        self.write(self.l_score, align="center", font=("Courier", 80, "normal"))
        self.goto(100, 200)
        self.write(self.r_score, align="center", font=("Courier", 80, "normal"))

    def l_point(self):
        self.l_score += 1
        self.update_scoreboard()

    def r_point(self):
        self.r_score += 1
        self.update_scoreboard()

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

#Turtle Crossing Game (my code)

#------------------------------------------File: player.py--------------------------------------------------------------

from turtle import Turtle

STARTING_POSITION = (0, -280)
MOVE_DISTANCE = 10
FINISH_LINE_Y = 280


class Player(Turtle):
    def __init__(self):
        super().__init__()
        self.shape("turtle")
        self.penup()
        self.goto(STARTING_POSITION)
        self.setheading(90)


    def up(self):
        self.forward(MOVE_DISTANCE)

#------------------------------------------File: main.py----------------------------------------------------------------

import time
from turtle import Screen
from player import Player, STARTING_POSITION, FINISH_LINE_Y
from car_manager import CarManager
from scoreboard import Scoreboard

screen = Screen()
screen.setup(width=600, height=600)
screen.tracer(0)

player = Player()

screen.listen()
screen.onkeypress(player.up,"Up")

car_manager = CarManager()

scoreboard = Scoreboard()


game_is_on = True
while game_is_on:
    time.sleep(0.1)
    car_manager.create_car()
    car_manager.cars_movement()
    screen.update()
    scoreboard.print_score()

    for car in car_manager.cars:
        if player.distance(car) < 30:
            scoreboard.game_over()
            game_is_on = False

    if player.ycor() > FINISH_LINE_Y:
        player.goto(STARTING_POSITION)
        car_manager.cars_speeder()
        scoreboard.print_score()
        scoreboard.update_score()



screen.exitonclick()

#------------------------------------------File: scoreboard.py----------------------------------------------------------

from turtle import Turtle

LEVEL = 1
FONT = ("Courier", 24, "normal")

screen_text = Turtle()
screen_text.penup()
screen_text.hideturtle()
screen_text.goto(-290, 260)

class Scoreboard:
    def __init__(self):
        self.level = LEVEL

    def game_over(self):
        screen_text.goto(-50,0)
        screen_text.write("Game Over", font=FONT)

    def update_score(self):
        self.level += 1

    def print_score(self):
        screen_text.clear()
        screen_text.write(f"Level {self.level}",font=FONT)

#------------------------------------------File: car_manager.py---------------------------------------------------------

from random import randint
from turtle import Turtle
import random

COLORS = ["red", "orange", "yellow", "green", "blue", "purple"]
STARTING_MOVE_DISTANCE = 5
MOVE_INCREMENT = 10


class CarManager:
    def __init__(self):
        self.cars = []
        self.move_distance = STARTING_MOVE_DISTANCE             #smart way to change a constant in a class in a method: store the current speed as an instance variable inside CarManager, not as a global constant.

    def create_car(self):
        random_chance = randint(1, 6)
        if random_chance == 6:
            square = Turtle()
            square.shape("square")
            square.shapesize(1, 2)
            square.penup()
            square.color(random.choice(COLORS))
            square.y_position = randint(-250, 250)
            square.x_position = randint(300, 500)
            square.goto(square.x_position, square.y_position)
            self.cars.append(square)



    def cars_movement(self):
        for car in range(len(self.cars)):
            new_location_x = self.cars[car].xcor() - self.move_distance
            self.cars[car].goto(new_location_x, self.cars[car].y_position)


    def cars_speeder(self):
        self.move_distance += MOVE_INCREMENT

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

#Turtle Crossing Game (her code)

#------------------------------------------File: player.py--------------------------------------------------------------

from turtle import Turtle

STARTING_POSITION = (0, -280)
MOVE_DISTANCE = 10
FINISH_LINE_Y = 280


class Player(Turtle):

    def __init__(self):
        super().__init__()
        self.shape("turtle")
        self.penup()
        self.go_to_start()
        self.setheading(90)

    def go_up(self):
        self.forward(MOVE_DISTANCE)

    def go_to_start(self):
        self.goto(STARTING_POSITION)

    def is_at_finish_line(self):
        if self.ycor() > FINISH_LINE_Y:
            return True
        else:
            return False


#------------------------------------------File: main.py----------------------------------------------------------------

import time
from turtle import Screen
from player import Player
from car_manager import CarManager
from scoreboard import Scoreboard

screen = Screen()
screen.setup(width=600, height=600)
screen.tracer(0)

player = Player()
car_manager = CarManager()
scoreboard = Scoreboard()

screen.listen()
screen.onkey(player.go_up, "Up")

game_is_on = True
while game_is_on:
    time.sleep(0.1)
    screen.update()

    car_manager.create_car()
    car_manager.move_cars()

    #Detect collision with car
    for car in car_manager.all_cars:
        if car.distance(player) < 20:
            game_is_on = False
            scoreboard.game_over()

    #Detect successful crossing
    if player.is_at_finish_line():
        player.go_to_start()
        car_manager.level_up()
        scoreboard.increase_level()


screen.exitonclick()

#------------------------------------------File: scoreboard.py----------------------------------------------------------

from turtle import Turtle

FONT = ("Courier", 24, "normal")


class Scoreboard(Turtle):

    def __init__(self):
        super().__init__()
        self.level = 1
        self.hideturtle()
        self.penup()
        self.goto(-280, 250)
        self.update_scoreboard()

    def update_scoreboard(self):
        self.clear()
        self.write(f"Level: {self.level}", align="left", font=FONT)

    def increase_level(self):
        self.level += 1
        self.update_scoreboard()

    def game_over(self):
        self.goto(0, 0)
        self.write(f"GAME OVER", align="center", font=FONT)


#------------------------------------------File: car_manager.py---------------------------------------------------------

from turtle import Turtle
import random

COLORS = ["red", "orange", "yellow", "green", "blue", "purple"]
STARTING_MOVE_DISTANCE = 5
MOVE_INCREMENT = 10


class CarManager:

    def __init__(self):
        self.all_cars = []
        self.car_speed = STARTING_MOVE_DISTANCE

    def create_car(self):
        random_chance = random.randint(1, 6)
        if random_chance == 1:
            new_car = Turtle("square")
            new_car.shapesize(stretch_wid=1, stretch_len=2)
            new_car.penup()
            new_car.color(random.choice(COLORS))
            random_y = random.randint(-250, 250)
            new_car.goto(300, random_y)
            self.all_cars.append(new_car)

    def move_cars(self):
        for car in self.all_cars:
            car.backward(self.car_speed)

    def level_up(self):
        self.car_speed += MOVE_INCREMENT


#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
#Modules
#breaking up the code accross multiple files, each file is called a module and it should contain all the related functions and classes, it is good for better code orginizaion and reusabilty

#------------------------------------------File: Converters.py----------------------------------------------------------
def lbs_to_kg(weight):
    return weight * 0.45

def kg_to_lbs(weight):
    return weight / 0.45

#------------------------------------------File: app.py-----------------------------------------------------------------
#two ways to import a module

#1
import converters #imports the entire module
print(converters.kg_to_lbs(70)) #using a module's function with entire module imports, having to prefix it with the name of an object "converters"

#2
from converters import kg_to_lbs #import specific functions from the module (presing Ctrl + Space after import will open a menu of all the functions that module has
print(kg_to_lbs(23)) #you can call the module's function directly when using the from statement in the line above

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
#Modules Excersice: make a find max function and put it in a separate file

#------------------------------------------File: utils.py---------------------------------------------------------------

def find_max(numbers):
    maximum = numbers[0]

    for number in numbers:
        if number > maximum:
            maximum = number
    return maximum

#------------------------------------------File: app.py-----------------------------------------------------------------

from utils import find_max

potato = [4, 3, 5, 7, 3] #can call the list whatever we want

print(f"The biggest number is: {find_max(potato)}") #the function's parameter "numbers" can be passed to an external variable with any name just fine just by calling the function and name the argument with the name of the variable you want to pass it to, in this case, potato

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
#NOTES:
#Ctrl + / automatically comments out a line, hit it again to remove the comment
max() #is a built in function to find the maximum number, assigning a variable to max will overwrite the function
#shift + F6 refactors/renames variables
#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

#Packages
#Package is a container for multiple modules, a folder or a directory
#Create one from the project panel, and create a file inside called __init__.py to so that python intepreter treats it as a package

#------------------------------------------File: shipping.py-----------------------------------------------------------------
def calc_shipping():
    print("Calc_Shipping")

#------------------------------------------File: app.py-----------------------------------------------------------------
#three ways to import from another package

#1
import ecommerce.shipping #importing a module from another package
ecommerce.shipping.calc_shipping() #usuing the function, syntax: package.module.function()

#2
#single function
from ecommerce.shipping import calc_shipping #using from to import only the functio
calc_shipping() #functions can be called directly now

#multiple functions
from ecommerce.shipping import calc_shipping, calc_tax() #typing comma , between function names
calc_shipping()

#3
#can import the entire module using from
from ecommerce import shipping
shipping.calc_shipping()

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

#Generating Random Values
#google: python 3 module index for all the python builtin methods
import random

for i in range(3):
    print(random.random()) # generates a number between 0 and 1 (float)
    print(random.randint(5, 15)) #random whole numbers between two numbers


members = ["John", "mary", "bob", "Mosh"]
print(random.choice(members)) #choice() randomly picks an item from a specified list


#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
#who pays the bill random exercise

import  random

names_string = input("give Names: ")

names = names_string.split(", ")

who_pays= names[random.randint(0,(len(names)-1))]       #this list will now work with however long the list is using len() function, the -1 is because the len() function counts from 1, but the index counts from 0, so to sync them we subtract 1

print(f"{who_pays} who will pay us tonight")


#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
#who pays the bill random exercise (using random.choice() method)

import  random

names_string = input("give Names: ")

names = names_string.split(", ")

who_pays= random.choice(names)                      #random.choice() is a method that will pick an item from a sequence/list randomly
print(f"{who_pays} who will pay us tonight")

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

#Excersice, create a dice class and generate two random numbers and store them in a tuple (HARD, didn't solve it)
import random


class Dice:
    def roll(self):
        first = random.randint(1, 6)
        second = random.randint(1, 6)
        return first, second #when you want to return a tuple from a function you type them like so, without paranthesis


dice1 = Dice() #creating the object
print(dice1.roll()) #calling the method

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

#Opening, Reading and Writing into Files

file = open("my_file.txt")              # opening the file takes some resources from pc, should be closed when done
contents = file.read()                  # filename.read() returns the contents of the file
print(contents)
file.close()                            # closing the file to clear up some resources


# another way of opening the file (this one does not need a close() statement
with open("my_file.txt") as file:
    contents = file.read()
    print(contents)


with open("my_file.txt", mode="a") as file:             # default open mode is ready only: mode="r", and mode="w" means open file in write mode (will replace everything), to add to file without replace use mode="a" which stands for append
    file.write("\nNew text.")


with open("new_file.txt", mode="w") as file: # if new_file.txt doesn't exist, it will create that file for you, only works with write mode "w"
    file.write("\nNew text.")

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

# Snake game with a highscore feature (based on her version)

#------------------------------------------File: snake.py---------------------------------------------------------------

# Unchanged

#------------------------------------------File: food.py----------------------------------------------------------------

# Unchanged

#------------------------------------------File: main.py----------------------------------------------------------------

from turtle import Screen
from snake import Snake
from food import Food
from scoreboard import Scoreboard
import time

screen = Screen()
screen.setup(width=600, height=600)
screen.bgcolor("black")
screen.title("My Snake Game")
screen.tracer(0)

snake = Snake()
food = Food()
scoreboard = Scoreboard()

screen.listen()
screen.onkey(snake.up, "Up")
screen.onkey(snake.down, "Down")
screen.onkey(snake.left, "Left")
screen.onkey(snake.right, "Right")

game_is_on = True
while game_is_on:
    screen.update()
    time.sleep(0.1)
    snake.move()

    #Detect collision with food.
    if snake.head.distance(food) < 15:
        food.refresh()
        snake.extend()
        scoreboard.increase_score()

    #Detect collision with wall.
    if snake.head.xcor() > 280 or snake.head.xcor() < -280 or snake.head.ycor() > 280 or snake.head.ycor() < -280:
        scoreboard.reset()
        snake.reset()

    #Detect collision with tail.
    for segment in snake.segments:
        if segment == snake.head:
            pass
        elif snake.head.distance(segment) < 10:
                scoreboard.reset()
                snake.reset()


screen.exitonclick()


#------------------------------------------File: scoreboard.py----------------------------------------------------------

from turtle import Turtle
ALIGNMENT = "center"
FONT = ("Courier", 24, "normal")



class Scoreboard(Turtle):

    def __init__(self):
        super().__init__()
        self.score = 0
        with open("data.txt") as score_file:                    # reading the data.txt file to access previous highscore from the file
            self.highscore = int(score_file.read())
        self.color("white")
        self.penup()
        self.goto(0, 270)
        self.hideturtle()
        self.update_scoreboard()

    def update_scoreboard(self):
        self.clear()
        self.write(f"Score: {self.score} High Score: {self.highscore}", align=ALIGNMENT, font=FONT)


    def reset(self):
        if  self.score > self.highscore:
            self.highscore = self.score
            with open("data.txt", mode="w") as new_score_file:              # writing the newest highscore to the data.txt file
                new_score_file.write(str(self.highscore))
        self.score = 0
        self.update_scoreboard()

    # def game_over(self):
    #     self.goto(0, 0)
    #     self.write("GAME OVER", align=ALIGNMENT, font=FONT)

    def increase_score(self):
        self.score += 1
        self.update_scoreboard()

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

# Absolute and Relative Paths

with open("C:/Users/ACER/Desktop/new_file.txt") as file:            #Absolute Path
    contents = file.read()
    print(contents)


with open("../../../../../../Art/new_file.txt") as file:            #Relative Path (../ means going backwards once, then going forwards to Art in this example
    contents = file.read()
    print(contents)

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

#Files and Directories
from pathlib import Path
import os # this library is imported to use os.getcwd() to check current working directory

#Absolute Path
#C:\Program Files\Microsoft #Windows Path
#/usr/local/bin #Linux Path

#Relative Path:
path = Path("emails") # not typing an argument between the paranthesis it will reference the current directory
print(path.exists())
print(os.getcwd()) # this is used to check the current working directory
print(path.mkdir()) # to create a new directory
print(path.rmdir()) # to remove the directory
path1 = Path()
for file in path1.glob('*.py'): # this method returns a generator object, with it you can search for the files and directories in the current path, '*' Means everything, '*.*' will get all the files without the directories, '*.py' search all the files with the .py extionsion
    print(file) #this loop will return all .py files

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

# Mail Merge Challenge (check project to work) (my code)
# it reads names from a file, strips extra whitespaces, replaces the placeholder [name], then writes letter.txt files each by their own names

with open("Input/Names/invited_names.txt") as invited_names:
    names_list = invited_names.readlines()                                      #reads the lines and puts each line as an item in a list
    for name in names_list:
        new_name = name.strip()                                                 #rremoves any leading, and trailing whitespaces, including end lines.

        with open("Input/Letters/starting_letter.txt") as starting_letter:
            letter = starting_letter.read()
            replaced_letter = letter.replace("[name]", new_name)

        with open(f"Output/ReadyToSend/{new_name}.txt", mode="w") as mail_merged:
            mail_merged.write(replaced_letter)

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

# Mail Merge Challenge (check project to work) (her code) (./ for pathing didn't not work for me for some reason but here it does)

PLACEHOLDER = "[name]"


with open("./Input/Names/invited_names.txt") as names_file:
    names = names_file.readlines()

with open("./Input/Letters/starting_letter.txt") as letter_file:
    letter_contents = letter_file.read()
    for name in names:
        stripped_name = name.strip()
        new_letter = letter_contents.replace(PLACEHOLDER, stripped_name)
        with open(f"./Output/ReadyToSend/letter_for_{stripped_name}.txt", mode="w") as completed_letter:
            completed_letter.write(new_letter)

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

# three methods to read csv files

# method 1
# can use this way, but it's not good for csv

with open("weather-data.csv") as weather_data:
    data = weather_data.readlines()
    for line in data:
        new_line = line.strip()
        print(new_line)


# method 2

import csv

with open("weather-data.csv") as weather_data:
    data = csv.reader(weather_data)                     #creates an object that can be looped through
    temperatures = []
    for row in data:
        if row[1] != "temp":
            temperatures.append(int(row[1]))
    print(temperatures)


# method 3 (best)

import pandas

data = pandas.read_csv("weather-data.csv")
print(data["temp"])                                     # showing only the temp column ( data["temp"] is a series)

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
# Pandas

import pandas

data = pandas.read_csv("weather-data.csv")
print(type(data))                                             # data is a DataFrame
print(type(data["temp"]))                                     # showing only the temp column ( data["temp"] is a series)

data_dict = data.to_dict()                                      # converting a DataFrame to a dictionary (wow)
print(data_dict)

temp_list = data["temp"].to_list()                              # converting a series to a list
print(temp_list)

average = sum(temp_list) / len(temp_list)                       # simplest way to find the average in a list
print(average)

mean = data["temp"].mean()                                      # finding the average in a series
print(mean)

maxx = data["temp"].max()                                       # finding the maximum in a series
print(maxx)

print(data["condition"])                                        # returns the condition column, pulling columns by the key, like a dictionary
print(data.condition)                                           # returns the condition column, pulling columns by its attribute, like an object, all headings of the table are converted to attributes, which allows such a call to be made

print(data[data["day"] == "Monday"])                              # return the row where the "day" column is equal to "Monday"
print(data[data.day == "Monday"])                                 # this is the same as the above but using the object style instead of the dictionary style

print(data[data.temp == data["temp"].max()])                      # return the row where the temperature was the highest at

monday = data[data.day == "Monday"]                                 # entire row, we saved it in a variable
print(monday.condition)                                           # now we can access each value in the row by the column name the same way we did for the entire table (nice)

#we can also do it directly without using a variable:
print(data[data.day == "Monday"].condition)


# Creating a dataframe from scratch (from a dictionary)
data_dict = {
    "students": ["Amy", "James", "Angela"],
    "scores": [76, 56, 65]
}

data = pandas.DataFrame(data_dict)
data.to_csv("new_data.csv")

#

# Squirrel count (my code)

import pandas

data = pandas.read_csv("Squirrel-Data.csv")
fur = data["Primary Fur Color"]
fur_colors = fur.value_counts()                                         # this function one handedly did all the work that retarded teacher did (get all the unique values and their count)
colors_DF = pandas.DataFrame(fur_colors)
colors_DF.to_csv("squirrel_count.csv")


# Squirrel count (her code) (absolutely retarded, I did it with 1 function)

import pandas

data = pandas.read_csv("Squirrel-Data.csv")
gray_sqrl = len(data[data["Primary Fur Color"] == "Gray"])              #can use len() to get the number of gray squirrels by counting each row (since it's filtered by gray only)
red_sqrl = len(data[data["Primary Fur Color"] == "Cinnamon"])
black_sqrl = len(data[data["Primary Fur Color"] == "Black"])
print(gray_sqrl)
print(red_sqrl)
print(black_sqrl)

data_dict = {
    "fur Color": ["Gray", "Cinnamon", "Black"],
    "Count": [gray_sqrl, red_sqrl, black_sqrl]
}

df = pandas.DataFrame(data_dict)
df.to_csv("squirrel_count.csv")

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

# Squirrel color count (my code)

import pandas

data = pandas.read_csv("Squirrel-Data.csv")
fur = data["Primary Fur Color"]
fur_colors = fur.value_counts()                                         # this function one handedly did all the work that retarded teacher did (get all the unique values and their count)
colors_DF = pandas.DataFrame(fur_colors)
colors_DF.to_csv("squirrel_count.csv")

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

# Squirrel color count (her code) (absolutely retarded, I did it with 1 function)

import pandas

data = pandas.read_csv("Squirrel-Data.csv")
gray_sqrl = len(data[data["Primary Fur Color"] == "Gray"])              #can use len() to get the number of gray squirrels by counting each row (since it's filtered by gray only)
red_sqrl = len(data[data["Primary Fur Color"] == "Cinnamon"])
black_sqrl = len(data[data["Primary Fur Color"] == "Black"])
print(gray_sqrl)
print(red_sqrl)
print(black_sqrl)

data_dict = {
    "fur Color": ["Gray", "Cinnamon", "Black"],
    "Count": [gray_sqrl, red_sqrl, black_sqrl]
}

df = pandas.DataFrame(data_dict)
df.to_csv("squirrel_count.csv")

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

# US states game (my code) (run in project)

import turtle
from turtle import Turtle

import pandas

screen = turtle.Screen()
screen.title("U.S States Game")
image = "blank_states_img.gif"

screen.addshape(image)                          # addshape accepts images

turtle.shape(image)                             # a turtle can be an image
score = 0

#--------------------------------------------------
# this is only executed once to get all the coordinates of the states by clicking on them on the map and taking the coordinates of the clicked area

def get_mouse_click_coor(x, y):
        print(x, y)

turtle.onscreenclick(get_mouse_click_coor)          # onscreenclick is an event listener, it takes a function with coordinates, this will return the coordinates we clicked on the map and prints them
turtle.mainloop()                                   # this stops the screen from closing, without clicking
#--------------------------------------------------

data = pandas.read_csv("50_states.csv")
all_states = data["state"].tolist()
correct_guesses = []


while len(correct_guesses) < 50:
    answer_state = screen.textinput(title=f"{score}/50 States Correct", prompt="What's another state's name?")


    if answer_state is None:                            # Comparison with None performed with equality operators using: is
        break
    else:
        answer_state = answer_state.title()

    if answer_state in all_states:
        state_series = data[data["state"] == answer_state]
        state_name = state_series["state"].item()
        state_x = state_series["x"].item()                           #to get the raw value from the series
        state_y = state_series["y"].item()


        state_finder = Turtle()
        state_finder.penup()
        state_finder.hideturtle()
        state_finder.goto(state_x, state_y)
        state_finder.write(state_name)
        score += 1
        correct_guesses.append(state_name)
        print(correct_guesses)

missing_states = []
for state in all_states:
    if state not in correct_guesses:
        missing_states.append(state)
print(missing_states)

new_data = pandas.DataFrame(missing_states)
new_data.to_csv("missing_states.csv")
turtle.mainloop()

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

# US states game (her code) (run in project)

import turtle
import pandas

screen = turtle.Screen()
screen.title("U.S. States Game")
image = "blank_states_img.gif"
screen.addshape(image)
turtle.shape(image)

data = pandas.read_csv("50_states.csv")
all_states = data.state.to_list()
guessed_states = []

while len(guessed_states) < 50:
    answer_state = screen.textinput(title=f"{len(guessed_states)}/50 States Correct",
                                    prompt="What's another state's name?").title()
    if answer_state == "Exit":
        missing_states = []
        for state in all_states:
            if state not in guessed_states:
                missing_states.append(state)
        new_data = pandas.DataFrame(missing_states)
        new_data.to_csv("states_to_learn.csv")
        break
    if answer_state in all_states:
        guessed_states.append(answer_state)
        t = turtle.Turtle()
        t.hideturtle()
        t.penup()
        state_data = data[data.state == answer_state]
        t.goto(int(state_data.x), int(state_data.y))
        t.write(answer_state)

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

#Pypi and Pip
#openpyxl used to work with excel sheets

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

#Automation (Excel)

import openpyxl as xl # xl here will be used as an alias

wb = xl.load_workbook("transactions.xlsx")                  #load_workbook is a method to load an excel file, wb here is an object
sheet = wb["Sheet1"]                                        # to access a sepcific sheet (this relies on the workbook variable declared above)
cell = sheet["a1"]                                          # to access a specific cell (this relies on the sheet variable declared above)
cell_2 = sheet.cell(1, 1)                       # another way to access a cell by using the cell() method with the sheet object (variable)
cell.value                                                  #value method returns the value of that cell
sheet.max_row                                               # max_row is a method that returns the number of th rows

for row in range(2, sheet.max_row + 1):                     # the +1 is because the range doesn't include the last value, like range(1, 4) will return 1, 2, 3 but not 4, +1 will bring all, range started from 2 to ignore the header
    Cell = sheet.cell(row,3)
    corrected_price = Cell.value * 0.9                      #this will reduce all the prices by 10%
    corrected_price_cell = sheet.cell(row, 4)       #specify where to enter the data
    corrected_price_cell.value = corrected_price           # set the values in the specified cell/s

wb.save("transactions2.xlsx") #save the file


#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

#Automation Example (Same as above but without clutter + adding a chart)

import openpyxl as xl
from openpyxl.chart import BarChart, Reference # from openpyxl package, inside chart module, import two classes, BarChart and Reference

wb = xl.load_workbook("transactions.xlsx")
sheet = wb["Sheet1"]

for row in range(2, sheet.max_row + 1):
    cell = sheet.cell(row,3)
    corrected_price = cell.value * 0.9
    corrected_price_cell = sheet.cell(row, 4)
    corrected_price_cell.value = corrected_price

values = Reference(sheet,               #we're creating an instance of the Reference class, and stored it in a variable
                 min_row = 2,
                 max_row = sheet.max_row,
                 min_col = 4,
                 max_col = 4) #select a range of cells

chart = BarChart() # create an instance of the BarChart Class
chart.add_data(values) #adding the values using the add_data() method from BarChart Class
sheet.add_chart(chart, "e2") # add the chart to the sheet

wb.save("transactions2.xlsx")

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

# #Automation Example (Optimized Code above by Turning it into a function and you type in the name of the file)

import openpyxl as xl
from openpyxl.chart import BarChart, Reference              # from openpyxl package, inside chart module, import two classes, BarChart and Reference

def proccess_workbook(filename):
    wb = xl.load_workbook(filename)
    sheet = wb["Sheet1"]

    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row,3)
        corrected_price = cell.value * 0.9
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price

    values = Reference(sheet,                               # we're creating an instance of the Reference class, and stored it in a variable
                     min_row = 2,
                     max_row = sheet.max_row,
                     min_col = 4,
                     max_col = 4)                           # select a range of cells

    chart = BarChart()                                      # create an instance of the BarChart Class
    chart.add_data(values)                                  #adding the values using the add_data() method from BarChart Class
    sheet.add_chart(chart, "e7")                            # add the chart to the sheet

    wb.save(filename)

proccess_workbook(input())

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
#Jupyter Notebook code (using pandas to read and display the content of a csv file)

import pandas as pd
df = pd.read_csv("vgsales.csv")                             # df means data frame, read_csv() is a pandas built-in function
df_xlsx = pd.read_excel("vgsales.xlsx")                     # read_excel() is a pandas built-in function
df_txt = pd.read_csv('pokemon_data.txt', delimiter = '\t')  # this will read a tab separated text file and display it like a table, '\t' represents the tab key, delimiter can work for any other character separating the columns
df                                                          # typing this will inspect (View) the sheet
df.shape                                                    # returns the number of rows and columns (2D Array)
df.describe()                                               # returns basic information about each column (statistics), (count, mean, std, min, 25%, 50%, 75%, max) whatever any of those mean ¯\_(ツ)_/¯
df.values                                                   # shows the 2D array
df.head(2)                                                  # shows the first number of rows (2) of the file
df.tail(2)                                                  # shows the last number of rows (2) of the file

df.columns                                                  # prints the headers of the table
df['column name'][0:5]                                      # prints all the data in that column, can also specify the index range, for one worded columns, you can get away with ex. df.name
df[['Name', 'Type 1', 'HP']]                                # prints multiple columns in any order
df.iloc[1]                                                  # prints a specific row iloc stands for integer location, a function in the data frame
df.iloc[1:4]                                                # multiple rows also works
df.iloc[3][1]                                               # to get a specific value by [row][column] coordinates, this method is deprecated apparently
df.iloc[3,1]                                                # same as line above (new way of getting a specific value by coordinates)


for index, row in df.iterrows():                            # iterate through each row, (it sucks bcause it has to use print which ruins the format + it prints values of each row at a new line, iterrows() returns data as: (index, Series), series is a full column
    print(index, row)

for index, row in df.iterrows():                            # iterate through each row in Name Column, listing them one by one, seems very useful
    print(index, row['Name'])


df.loc[df['Type 1'] == "Fire"]                                  # Searches for specific values (in this case, it finds all rows that contain 'Fire' in 'Type 1' column)
df.loc[(df['Type 1'] == "Grass") & (df['Type 2'] == 'Poison')]  # Searches for multiple values, separated by () can use & for AND, | for OR
df = df.reset_index()                                           # filtered data keep the old index so you need to reset the index on the dataframe (old index will be added as a column)
df = df.reset_index(drop=True)                                  # will remove the old index column from the table
df.reset_index(drop=True, inplace=True)                     # this will update the dataframe without setting it to a new variable/dataframe


df.sort_values('Name', ascending = False)                   # Sorts the values alphabetically, Descending
df.sort_values(['Type 1', 'HP'], ascending = [1, 0])        # Sorts multiple columns, can also specify ascending/descending values for each in order with this [1, 0] where 1 is ascending for 'Type 1', and 0 is descending for 'HP'



df['Total'] = df['HP'] + df['Attack'] + df['Defence'] + df['Sp. Atk'] + df['Sp. Def'] + df['Speed']                 # adding a new  column to the dataframe called 'Total'

df = df.drop(columns=['Total'])                                                                                     # Dropping the column 'Total' from the dataframe

df['Total'] = df.iloc[:, 4:10].sum(axis=1)                                                                          # another way to add columns, [:, 4:10] the first : means all rows, second : means from column 4 to 10, sum(axis=1) adding values horizontally (values in 1 row), if axis=0 it would add vertically (values in different rows)

df = df[['Total', 'HP', 'Defense']]                                                                                 # show only these columns and in this order

# rearrange the total column and put it in the middle
cols = list(df.columns)                                                                                             # puts all column names in a list
df = df[cols[0:4] + [cols[-1]] + cols[4:12]]                                                                        # combine the columns you like in the orders you prefer, the middle one has [] brackets around it to convert it to a list because a single item will count as string/integer

df.to_csv('modified.csv')                                                                                           # saves the data frame to a csv file
df.to_csv('modified.csv', index = False)                                                                            # index = False removes the first index column from the file
df.to_excel('modified.xlsx', index = False)                                                                         # saves the data frame to an Excel file
df.to_csv('modified.txt', index = False, sep = '\t')                                                                # saves the data frame to a tab separated file, sep means separater, '\t' means tab

df.loc[df['Name'].str.contains('Mega')]                                                                             # looks for a text 'Mega' inside the 'Name' column entries
df.loc[~df['Name'].str.contains('Mega')]                                                                            # looks for a text 'Mega' inside the 'Name' column entries and NOT include it in the result

# Contains() also works with regex
# Examples With Regex
#--------------------------------------------------------------------------

import re

df.loc[df['Type 1'].str.contains('Fire|Grass', regex=True)]                                                         # 'Fire|Grass' means fire or grass
df.loc[df['Type 1'].str.contains('fire|grass', flags = re.I, regex=True)]                                           # flags = re.I means ignore case sensitivity


df.loc[df['Name'].str.contains('^pi[a z]*', flags = re.I, regex=True)]                                              # 'pi[a z]*' means look for anything that starts ^ with pi then 0 to infinite number of the letters a to z
df.loc[df['Name'].str.contains('pi[a z]*', flags = re.I, regex=True)]                                               # not having the ^ will search for all occurrences of pi

#--------------------------------------------------------------------------

df.loc[df['Type 1'] == 'Fire', 'Type 1'] = 'Flamer'                                                                 # replacing the word Fire to Flamer in all the data frame in  'Type 1' Column, .loc[] is label-based indexing, which allows you to, 1. Select rows based on a condition, 2. Select specific columns to work with. syntax: df.loc[df['Column_Name_where_data_exists'] == 'Old_Cell_Value', 'Column_where_the_data_will_be_changed'] = 'New_cell_value'

df.loc[df['Type 1'] == 'Fire', 'Legendary'] = True                                                                  # another example but making the changes on a different column 'Legendary', this makes all Fire Types Legendary

df.loc[df['Total'] > 500, ['Generation', 'Legendary']] = 'Test Value'                                               # Modifying multiple columns based on a condition to equal one value

df.loc[df['Total'] > 500, ['Generation', 'Legendary']] = ['Test 1', 'Test 2']                                       # Modifying multiple columns with corresponding multiple values each


# Groupby() method

df.groupby(['Type 1']).mean(numeric_only = True).sort_values('Defense', ascending=False)                            # this is grouping all the data and finding the average (mean) of all it's numercal values, and sorting by defense descending
df.groupby(['Type 1']).sum(numeric_only=True)                                                                       # Summing values
df.groupby(['Type 1']).count()                                                                                      # counting entries

df['count'] = 1
df.groupby(['Type 1']).count()['count']                                                                             # cleaner way to count by making a count column that equals 1 on each row and counting that

df.groupby(['Type 1', 'Type 2']).count()['count']                                                                   # grouping by multiple parameters, example, Type 1 of Bug, two have Type 2 of Electric, etc

pd.read_csv('modified.csv', chunksize=5)                                                                            # read large dataframes in chunks

# chunks presentation
for df in pd.read_csv('modified.csv', chunksize=5):
                     print("CHUNK DF\n")
                     print(df)


# example of chunks

new_df = pd.DataFrame(columns=df.columns)                                                                           # this code will read the big file by chunks, group the chunks by counting type 1, then concatenating the the chunk with the new dataframe, basically extract information from main df and store only the data you need in the new df
for df in pd.read_csv('modified.csv', chunksize=5):
     results = df.groupby('Type 1').count()
     new_df = pd.concat([new_df, results])

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
#Machine Learning Project (on Jupyter)

import pandas as pd
from sklearn.tree import DecisionTreeClassifier                 #machine learning algorithm (Syntax: from package.module import Class)

music_data = pd.read_csv("music.csv")
X = music_data.drop(columns = ["genre"])                        #drop a column from the dataset (won't modify original dataset, it will create a new one without the genre column, this line is the input set, Capital X is the conventional name for it
y = music_data["genre"]                                         #how to get all the values in a given column, this line is the output set, small y is the conventional name for it

model = DecisionTreeClassifier()                                #creating an instance of this Class
model.fit(X, y)                                                 #training our model with fit() method, takes two datasets, input and output
predictions = model.predict([ [21, 1], [22, 0] ])               #ask the model to make a prediction, predict() method takes 2D arrays
predictions

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
# Same Code a above (in Jupyter) Better version from the teacher's (fixes an error that is returns with the output)

#the Error: C:\Users\ACER\AppData\Local\Programs\Python\Python313\Lib\site-packages\sklearn\utils\validation.py:2739: UserWarning: X does not have valid feature names, but DecisionTreeClassifier was fitted with feature names warnings.warn(
#Feature names are the column names of your input dataset (X) that represent the characteristics (features) used to train the model. When you drop the "genre" column to create X, it still has column names (age, gender). But when you pass [[21, 1], [22, 0]] to model.predict(), Python treats it as a simple list without column names.

import pandas as pd
from sklearn.tree import DecisionTreeClassifier

music_data = pd.read_csv("music.csv")
X = music_data.drop(columns = ["genre"])
y = music_data["genre"]

model = DecisionTreeClassifier()
model.fit(X, y)
input_data = pd.DataFrame([ [21, 1], [22, 0] ], columns = X.columns)    # Instead of passing a raw list like [[21, 1], [22, 0]], convert it into a pandas DataFrame with the same column names, The '.columns' attribute of a Pandas DataFrame provides an index-like object containing the labels (names) of all the columns in that DataFrame. This allows you to retrieve all column names of `X`.

predictions = model.predict(input_data)
predictions

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
#Calculating the Accuracy (Edited on the Improved Code above)

import pandas as pd
from sklearn.tree import DecisionTreeClassifier
from sklearn.model_selection import train_test_split                                        #importing a function to split the dataset into two, one for training the model 80% and one for testing 20%
from sklearn.metrics import accuracy_score                                                  #function used to measure the accuracy of our prediction

music_data = pd.read_csv("music.csv")
X = music_data.drop(columns = ["genre"])
y = music_data["genre"]
X_train, X_test, y_train, y_test = train_test_split(X, y, test_size = 0.2)           #specifying the size of the test dataset as 20%, allocating 20% of the data for testing, this function returns a tuple, so we'll unpack it into four variables, train_test_split() function randomly picks data for training and testing, so the result will be different everytime, if we increase the testing dataset to 80% for example, the accuracy of the prediction will drop significantly, i.e the more training data, the better results

model = DecisionTreeClassifier()
model.fit(X_train, y_train)                                                                  #instead of passing the entire dataset for training, we pass only the training dataset
input_data = pd.DataFrame(X_test, columns = X.columns)                                       #using two samples was not enough for predictions, now we're using the testing dataset
predictions = model.predict(input_data)

score = accuracy_score(y_test, predictions)                                                  #comparing the output from our test dataset to the predicted output, returns an accuracy score between 0 to 1
score

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
#Calculating the Accuracy (Original Code without the improvement) works with no errors

import pandas as pd
from sklearn.tree import DecisionTreeClassifier
from sklearn.model_selection import train_test_split
from sklearn.metrics import accuracy_score

music_data = pd.read_csv("music.csv")
X = music_data.drop(columns = ["genre"])
y = music_data["genre"]
X_train, X_test, y_train, y_test = train_test_split(X, y, test_size = 0.2)

model = DecisionTreeClassifier()
model.fit(X_train, y_train)
predictions = model.predict(X_test)

score = accuracy_score(y_test, predictions)
score

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
#Model Persistence

#----------------------------------------------------Saving the Model---------------------------------------------------
import pandas as pd
from sklearn.tree import DecisionTreeClassifier
import joblib #object that has methods for saving and loading models, so that learned models get saved, and in the next time they get loaded before learning, to avoid relearning things

music_data = pd.read_csv("music.csv")
X = music_data.drop(columns = ["genre"])
y = music_data["genre"]

model = DecisionTreeClassifier()
model.fit(X, y)

joblib.dump(model, 'music-recommender.joblib') #save the learned model, takes two arguments, the name of the model and the name of the file that the model will be stored into

#----------------------------------------------------Loading the Model--------------------------------------------------
#we only need to learn once

import pandas as pd
from sklearn.tree import DecisionTreeClassifier
import joblib


model = joblib.load('music-recommender.joblib')

predictions = model.predict([ [21, 1], [22, 0] ])
predictions

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
#Visualizing a Decision Tree (Jupyter)

import pandas as pd
from sklearn.tree import DecisionTreeClassifier
from sklearn import tree                                        # this object has a method for exporting the decision tree in a graphical format

music_data = pd.read_csv("music.csv")
X = music_data.drop(columns = ["genre"])
y = music_data["genre"]

model = DecisionTreeClassifier()
model.fit(X, y)

tree.export_graphviz(model, #the model
                     out_file = "music-recommender.dot",        # name of the output file, this is using keyword arguments, to selectively pass an argument without worrying about its order, the .dot format is a graph description language
                     feature_names = ["age", "gender"],         # feature_names will be the columns/properties/features of the data, this line is so we can see the rules for each node (whether it was male/female and their age etc)
                     class_names = sorted(y.unique()),          # class_names = will be set to the list of classes/labels we have in our output dataset, the y dataset includes all the genres, but they're repeated, so we'll use .unique method to get the uniqe list without repitions, then sort it alphabetically using sorted() function, this line will display the class (genre) for each node
                     label = "all",                             # so every node has labels (text areas) so we can read them
                     rounded = True,                            # rounded corners for the square nodes
                     filled = True)                             # means each node will be filled with a color

#after executing the code, the graph will be saved to music-recommender.dot file, which you can view in vs code using graphviz (dot) extensions

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

#---------------Django---------------#

#NOTES:
# 1- Ctrl + L: clears the terminal window
# 2 - for the django project Will use: pip install django==2.1 (turns out it sucks, just install the latest version)
# 3 - to create a django project in the new django folder, use this command django-admin startproject pyshop . # django-admin is a django utility/program that can be executed in the commandline, startproject creates the project, pyshop the name of the project, . means in this current folder, the modules created are: __init__.py means it's a package, settings.py is where we define various settings for our application, urls.py is where we'll define what will the user see when he goes to /about /contact /products etc, wsgi.py is webserver gateway interface which the purpose of is to provide a standard interface between applications built with django and web servers, asgi.py application server gateway interface (I assume)
# 4 - type the command: python manage.py runserver #using python interpreter, we are running the manage.py program and passing runserver as an argument, manage.py is the same program as django-admin, except django-admin is used before creating the django project when wr have a project, we work with manage.py instead, which is a model we use to manage our django project, this command will start a localhost web server
# 5 - type the command: python manage.py startapp products # each web app can consist of multiple apps, each app can serve a specific function, like customer management, order management, etc. each having their own functions etc, this command will create a package, which can be reusable for other django projects as well, the modules created are: __init__.py means it's a package, admin.py is used for defining how the administrative panel of this app will look like, apps.py (should be called config.py) is a confusing name but it's used to store various configuration settings for this app, models.py is where we define classes or new types for modeling the concepts in this app (domain) Ex. in products domain you will probably have concepts like products, category, items, etc, tests.py is where we write automated tests, views.py is where we define what should the user see when they navigate to a certain page
# 6 - packages with certain functionalities can be published for other people to use at pypi (meaning i can install some stuff from there without having to create it from scratch, yay :)

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

#View Functions
# is function that gets called by django when the user navigates to a particular page ( ex. when typing /products on the browser, the browser sends an HTTP request to our webserver, django takes that request, inspects the url, then it will call a view function whose job is to return the response to the browser, generating an HTML Markup (HTML code) to return to the client, and it will be shown in the browser

#----------------------------------------------File: Products/views.py--------------------------------------------------


from django.http import HttpResponse        #importing a class called HttpResponse, with this class we can create an http response to return to the client(browser)
from django.shortcuts import render         #from django package, shortcut module, import render function


def index(request):                          #all view functions should always take a parameter, request is HTTP request
    return HttpResponse("Hello World")

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------


#URL Mapping

#NOTES:
# 1 - as soon as you define a new url pattern in the project that default django homepage (the one with the rocket) disappears


#----------------------------------------------File: products/views.py--------------------------------------------------


from django.http import HttpResponse
from django.shortcuts import render


def index(request):
    return HttpResponse("Hello World")


#----------------------------------------------File: products/urls.py---------------------------------------------------


from django.urls import path                    #path function is used to reference urls, i.e. we can map a url to a view function
from . import views                             # . means the current folder, importing views module so that we can map the path to the view, we didn't use "import views" directly because it could conflict with a dependency from another libray with the same name


urlpatterns = [                                 #by convention, this variable needs to be created and named exactly like so: urlpatterns this is where we'll map urls to views
    path('',                                    #the first argument is a string that specifies the url, here we're passing an empty string which represents the root of this app i.e. /products
         views.index)                           #here we're not calling the index() function, only passing a reference to it, and it will get called on runtime by django when the client sends an HTTP request to the server so we shouldn't call it here,  Reminder: when importing a module, it will end up being an object, using the . Operator allows us to access its functions
]


#----------------------------------------------File: pyshop/urls.py---------------------------------------------------


"""
URL configuration for pyshop project.

The `urlpatterns` list routes URLs to views. For more information please see:
    https://docs.djangoproject.com/en/5.1/topics/http/urls/
Examples:
Function views
    1. Add an import:  from my_app import views
    2. Add a URL to urlpatterns:  path('', views.home, name='home')
Class-based views
    1. Add an import:  from other_app.views import Home
    2. Add a URL to urlpatterns:  path('', Home.as_view(), name='home')
Including another URLconf
    1. Import the include() function: from django.urls import include, path
    2. Add a URL to urlpatterns:  path('blog/', include('blog.urls'))
"""
from django.contrib import admin                         #added by default (the comments above too)
from django.urls import path, include                    #added include function, which will be used to include the products/urls.py module in this parent url.py module

urlpatterns = [                                          #this is the parent urls module, any url in the products app won't show unless django recognizes that the product package exists, aso the product's urls.py module needs to be referenced here using include function
    path('admin/', admin.site.urls),                     #here we're telling django that any url that starts with admin, delegate them to the admin app, every django project by default comes with an administrative panel
    path('products/', include('products.urls'))          #here we're making django aware of the urls.py module inside the products package, i.e. we need to reference the urls module in the products app. this needs to be referenced only once, afterwards, you only need to define the urls and view functions in the products app
]


#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
#Excersice: Add another Page: products/new

#Unchanged Files: pyshop/urls.py
#----------------------------------------------File: views.py-----------------------------------------------------------

from django.http import HttpResponse
from django.shortcuts import render


def index(request):
    return HttpResponse("Hello World")


def new(request):
    return HttpResponse("New Products!")


#----------------------------------------------File: products/urls.py---------------------------------------------------

from django.urls import path
from . import views


urlpatterns = [
    path('',views.index),
    path('new/', views.new)
]

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

#Models

# it is a representation of a real world concept, ex. order, customer, shopping cart, product

#----------------------------------------------File: products/models.py-------------------------------------------------


from django.db import models                                       # from django.db package, import models module, which contains the Model Class from django, this module is imported by default


class Product(models.Model):                                       # we will inherit this class from the Model class in django, Model class defines the characteristics and behavior for models in a django application ex. storing models in a database, or read them from a database, or delete them, common operations that we need to perform on model objects, by inheriting them, our Product class will have all these capabilities built in without having to code them
    name = models.CharField(max_length = 255)                      # defining an attribute for our product, setting it to an instance of the CharField class from models module, CharField (Character Field) is a class that represents a field that can contain textual data, will also supply it with a maximum length to prevent malicious user form entering large data, passing a keyword argument (kwarg) max_length
    price = models.FloatField()                                    # field for floating point numbers
    stock = models.IntegerField()                                  # field for Integer numbers
    image_url = models.CharField(max_length = 2083)                # 2083 is the standard maximum length for urls


#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

#Migrations

# NOTES:
# 1- type the command: python manage.py makemigrations # nothing will happen because django doesn't see our products model.py (which is the class we created above) to migrate its data to a database, to fix this, we need to add the products app in the pyshop settings.py file in the INSTALLED_APPS variable contaning a List of the installed apps by typing its class found in products in apps.py which is the configurations file for the products app, in this format package.module.ClassName
# 2- running the command above again will create a migration file to add the table from the model to the database, which will contain an operation specifying all the clumns in that table + the ID primary key
# 3- run the command: python manage.py migrate to start the migration, and add the table (stored in db.sqlite3, which can be viewed in DB Browser for SQLite, a light SQL Language Editor), running this command, django will look at every app and run every pending migration

#----------------------------------------------File: products/apps.py---------------------------------------------------

from django.apps import AppConfig


class ProductsConfig(AppConfig):                            #to add the products app to the installed apps, add the package.module.ClassName of the app in the pyshop settings.py module
    default_auto_field = 'django.db.models.BigAutoField'
    name = 'products'


#----------------------------------------------File: pyshop/settings.py-------------------------------------------------

#snippet of the code


# Application definition

INSTALLED_APPS = [
    'django.contrib.admin',                      # the administrative Panel
    'django.contrib.auth',                       # used to authenticate and authorize users, give roles, and permissions
    'django.contrib.contenttypes',
    'django.contrib.sessions',
    'django.contrib.messages',
    'django.contrib.staticfiles',
    'products.apps.ProductsConfig'               # adding the products app by adding its class from apps.py like so
]

#---------------------------------------File: products/migrations/0001_initial.py---------------------------------------

# this module was generated    after   Command 1: python manage.py makemigrations
# after it was generated, migrate with command 2: python manage.py migrate

# Generated by Django 5.1.7 on 2025-03-16 16:53

from django.db import migrations, models


class Migration(migrations.Migration):

    initial = True

    dependencies = [
    ]

    operations = [                                                          #this operation was automatically created when we made the migration for our products model.py, which is the first step to add it to the database
        migrations.CreateModel(
            name='Product',
            fields=[
                ('id', models.BigAutoField(auto_created=True, primary_key=True, serialize=False, verbose_name='ID')),
                ('name', models.CharField(max_length=255)),
                ('price', models.FloatField()),
                ('stock', models.IntegerField()),
                ('image_url', models.CharField(max_length=2083)),
            ],
        ),
    ]

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------


#Migration Excersice

#----------------------------------------------File: products/models.py-------------------------------------------------

#snippet of the code

from django.db import models

class Offer(models.Model):
    code = models.CharField(max_length = 7)
    description = models.CharField(max_length = 255)
    discount = models.FloatField()


#then run Command 1: python manage.py makemigrations  #creates 0002_offer.py

#----------------------------------------File: products/migrations/0002_offer.py----------------------------------------

# Generated by Django 5.1.7 on 2025-03-16 18:38

from django.db import migrations, models


class Migration(migrations.Migration):

    dependencies = [
        ('products', '0001_initial'),
    ]

    operations = [
        migrations.CreateModel(
            name='Offer',
            fields=[
                ('id', models.BigAutoField(auto_created=True, primary_key=True, serialize=False, verbose_name='ID')),
                ('code', models.CharField(max_length=7)),
                ('description', models.CharField(max_length=255)),
                ('discount', models.FloatField()),
            ],
        ),
    ]

#then run command 2: python manage.py migrate       #and the table is added to the database, yay :)

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

#Admin

#NOTES:
# 1- First we need to create the first user, a superuser, run this command: python manage.py createsuperuser
# 2- admin panel has users and groups settings by default, if we want to control the products app from the admin panel, we have to register it in products/admin.py
# 3- afterwards you can add items easily from the admin panel


#----------------------------------------------File: products/admin.py--------------------------------------------------

from django.contrib import admin
from .models import Product             # .models means the models module in the current folder, import the Product class

admin.site.register(Product)            #this will add the products app settings to the admin panel, syntax: object.attribute.method(Class), admin is an object (imported), site is an object itself but is passed as n attribute, which has a method called register, and we passed the Product class as an argument for that method


#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

#Customizing the Admin

#----------------------------------------------File: products/admin.py--------------------------------------------------

from django.contrib import admin
from .models import Product                                 # .models means the models module in the current folder, import the Product class


class ProductAdmin(admin.ModelAdmin):                       #create new class, naming convention, syntax: ModelnameAdminsuffix, ProductAdmin, class is inherited from django's admin module's class ModelAdmin, which has all the common functionality for managing models in the admin area, we're doing this so that we can make the columns visible in the new added product in the admin panel
    list_display = ('name', 'price', 'stock')               #in the body of this class we can overwrite/change functionalities/attributes, list_display attribute specifies the coluns that should be visible in the table in admin panel, will set it to a tuple


admin.site.register(Product, ProductAdmin)                  # we need to register ProductAdmin now as well so that it shows on admin panel


#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

#Exersice: add offers to admin panel

from django.contrib import admin
from .models import Product, Offer


class ProductAdmin(admin.ModelAdmin):
    list_display = ('name', 'price', 'stock')


class OfferAdmin(admin.ModelAdmin):
    list_display = ('code', 'discount', 'description')


admin.site.register(Product, ProductAdmin)
admin.site.register(Offer, OfferAdmin)          #we register are offer model for administration (link Offer class to OfferAdmin class)

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

#Templates
# adding the content to the web page
# 1-  we will create a Directory (folder) in the products root directory and call it templates (has to be this exact name because django will look for templates in a directory with this exact name
# 2- then create a file called index.html, where we will make an an h1 html header and an unordered list

#----------------------------------------------File: views.py-----------------------------------------------------------


from django.http import HttpResponse
from django.shortcuts import render                         # we use this function to render a template, it was added by default
from .models import Product                                 # we're importing the product so that we can view them in the index function (main web page) after storing the products in the database


def index(request):
    products = Product.objects.all()                        # method that returns all the products we have in the database, one of the useful methods this class inherited from django.db Model class in models module, other useful methods like .filter(), for filtering products ex. only out of stock ones, and .get() to return a single product, and .save() for inserting a product or updating one
    return render(request, 'index.html',                    # make this index view method return the index.html template using the render function, request needs to be put as first argument
                  {'products':products})                    # in the third argument we pass the list of products we loaded from the database, need to pass it as a dictionary, adding one key => value pair, we'll call the key products, but it can be any name it doesn't matter, and the value for that key should be the object(variable) products defined in the line above, this dictionary is called the context, it provides data to be used in a template, in this case we have a property called 'products' (the key), which we can access in the index.html template, in the for loop


def new(request):
    return HttpResponse("New Products!")


#----------------------------------------------File: index.html---------------------------------------------------------

<h1>Products</h1>
<ul>
    {% for product in products %}                                   <!--- Template Tag in django, used to dynamically execute logic, WOW, we can write dynamic code between the markers, like loops, if statements etc. the 'products' here is the property of the dictionary (the key) it's accessible here, through it, we will loop over all the products --->
        <li>{{ product.name }} (${{ product.price }})</li>          <!--- double curly braces, used to dynamically render values in an HTML markup, whenever double curly braces {{ are used in a django template, django will evaluate the expression in between and render it in html, this will bring all the items from the products model dynamically like so, product here is the for variable, will be represent the 'products' key property in the dictionary, which gets translated to the products value, which is the variable of the index view function whose storing the result of Product.objects.all(), which will bring all the products, then the for products variable increments, and translates to the second item in the dictionary and so on, we also added the prices the same way, parenthesis was for ethicist purposes --->
    {% endfor %}                                                    <!---- to end the for block, because django unlike python doesn't rely on indentation to determine the scope of a block --->
</ul>

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

#Adding Bootstrap

# 1- we will create a template for bootstrap that has the basic bootstrap definition, that we will be able to use everywhere, we'll call it base.html

#----------------------------------------------File: base.html----------------------------------------------------------

<!doctype html>
<html lang="en">
  <head>
    <meta charset="utf-8">
    <meta name="viewport" content="width=device-width, initial-scale=1">
    <title>Bootstrap demo</title>
    <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.3/dist/css/bootstrap.min.css" rel="stylesheet" integrity="sha384-QWTKZyjpPEjISv5WaRU9OFeRpok6YctnYmDr5pNlyT2bRjXh0JMhjY6hW+ALEwIH" crossorigin="anonymous">
  </head>
  <body>
    {% block content %}   <!--- we're creating a block/hole in the document called content, this will get filled by defining a block with the same name and has the elements in it while also extending base.html there, in this case it's index.html, what will happen is index.html will load the markup here, and add its content inside this block here, and view it in index.html (basically taking the styles from here --->
    {% endblock %}        <!-- ending the block --->

    <script src="https://cdn.jsdelivr.net/npm/bootstrap@5.3.3/dist/js/bootstrap.bundle.min.js" integrity="sha384-YvpcrYf0tY3lHB60NNkmXc5s9fDVZLESaAA55NDzOxhy9GkcIdslK1eN7N6jIeHz" crossorigin="anonymous"></script>
  </body>
</html>

#----------------------------------------------File: index.html---------------------------------------------------------

{% extends 'base.html' %}                                               <!--- extends statement: extending the base template i.e. we're going to use all the markup defined in base.html here, (similar to the include feature in php) --->

{% block content %}                                                     <!--- this block is also defined in the base.html file with the same name, where the content of the block here will be inserted there, and view it here with the styles defined in the base.html --->
    <h1>Products</h1>
    <ul>
        {% for product in products %}
            <li>{{ product.name }} (${{ product.price }})</li>
        {% endfor %}
    </ul>
{% endblock %}

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

#rendering cards
# we will use the Card Markup from Bootstrap by heading over to their website

#----------------------------------------------File: index.html---------------------------------------------------------

{% extends 'base.html' %}

{% block content %}
<h1>Products</h1>
    <div class="row">                                                                               <!--- created by typing div.row then press tab, the other is div.col and press tab --->
        {% for product in products %}
            <div class="col">                                                                       <!--- bootstrap code pasted for a card item inside this column --->
            <div class="card" style="width: 18rem;">                                                <!--- code was beautified here using Ctrl+Alt+Shift+L or reformat code from Code at the top bar --->
                <img src="{{ product.image_url }}" class="card-img-top" alt="...">                  <!--- you can add the double curly inside tag attributes inside the quotation marks and it works just fine --->
                <div class="card-body">
                    <h5 class="card-title">{{ product.name }}</h5>                                  <!--- details added dynamically --->
                    <p class="card-text"> ${{ product.price }}</p>
                    <a href="#" class="btn btn-primary">Add to Cart</a>
                </div>
            </div>
        </div>
        {% endfor %}
    </div>
{% endblock %}


#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

# Final Touches

# we added the navbar in the templates file
# we moved the templates.html from products, to a new folder in the root directory called templates so that we can use it in all apps not just the products app, but now we need django to look in the correct directory for templates, we will do that in pyshop/settings.py
# we will use a container from bootstrap for the padding, then add the content block inside

#----------------------------------------------File: base.html----------------------------------------------------------

<!doctype html>
<html lang="en">
  <head>
    <meta charset="utf-8">
    <meta name="viewport" content="width=device-width, initial-scale=1">
    <title>Bootstrap demo</title>
    <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.3/dist/css/bootstrap.min.css" rel="stylesheet" integrity="sha384-QWTKZyjpPEjISv5WaRU9OFeRpok6YctnYmDr5pNlyT2bRjXh0JMhjY6hW+ALEwIH" crossorigin="anonymous">
  </head>
  <body>
    <nav class="navbar bg-body-tertiary">                 <!--- pasted Navbar from bootstrap --->
      <div class="container-fluid">
        <a class="navbar-brand" href="#">PyShop</a>
      </div>
    </nav>
    <div class="container">                               <!--- using the container from bootstrap, by typing div.container then pressing tab, will give nice look + padding, --->
    {% block content %}
    {% endblock %}
    </div>

    <script src="https://cdn.jsdelivr.net/npm/bootstrap@5.3.3/dist/js/bootstrap.bundle.min.js" integrity="sha384-YvpcrYf0tY3lHB60NNkmXc5s9fDVZLESaAA55NDzOxhy9GkcIdslK1eN7N6jIeHz" crossorigin="anonymous"></script>
  </body>
</html>

#------------------------------------------File: pyshop/settings.py-----------------------------------------------------

#code snippets

import os                                                                           #it wasn't imported automatically, I imported it

#-----------------------------------------------------------------------------------------------------------------------

BASE_DIR = Path(__file__).resolve().parent.parent                                   # this variable is set to the complete path of where you stored your django project on your disk, we will take this directory and append templates to it

#-----------------------------------------------------------------------------------------------------------------------

TEMPLATES = [
    {
        'BACKEND': 'django.template.backends.django.DjangoTemplates',               # engine that is used to parse our template, including template tags, double curly braces, this is the engine that knows how to interpret those
        'DIRS': [                                                                   # here we need to reference the templates folder in this project
            os.path.join(BASE_DIR, 'templates')                                     # we're using the path object defined in the os module that is imported on the top, the path object has a method called join(), using the join() method we can join multiple parts of a path, taking two parameters, BASE_DIR, which is set to the complete path of our django project, the second parameter is our 'templates' folder, which we will append to the base path
        ],
        'APP_DIRS': True,
        'OPTIONS': {
            'context_processors': [
                'django.template.context_processors.debug',
                'django.template.context_processors.request',
                'django.contrib.auth.context_processors.auth',
                'django.contrib.messages.context_processors.messages',
            ],
        },
    },
]

#course over :)
#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

#100 Days of code

#Notes
# 1.
def botato(hamodi):
    print(f"potato {hamodi}")       #if you highlight a word in a string and press { it will enclose it like so {word} useful with formatted strings, also works with ""

# 2.
# TODO you can make Todo notes like so, putting them all over the code and a access them from the bottom left button called TODO (three dots and lines)

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

# list Comprehension

# useful to create lists from sequences, sequences are: (lists, ranges, tuples, strings) with one line
# superior to making an emtpy list and a for loop with its operation that takes 4 lines

# syntax: new_list = [new_item for item in list]

# Example 1 (without list Comprehension)

numbers = [1, 2, 3]
new_list = []
for n in numbers:
    add_1 = n + 1
    new_list.append(add_1)          #it took 4 Lines to create the new list

# Example 1 (with list Comprehension)

numbers = [1, 2, 3]
new_list = [n + 1 for n in numbers]

# Example 2 (with list Comprehension)

name = "angela"
letters = [letter for letter in name]

# Example 3 (with list Comprehension)

range_sequence = range(1, 5)
new_r = [number * 2 for number in range_sequence]

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

# Conditional list Comprehension

# syntax: new_list = [new_item for item in list if test]

# Example

names = ["Alex", "beth", "Caroline", "Dave", "Eleanor", "Freddie"]

new_names = [name for name in names if len(name) < 5]               # prints only names shorter than 5 letters

upper_long = [name.upper() for name in names if len(name) > 5]      # prints only names longer than 5 letters and makes them all uppercase

#-----------------------------------------------------------------------------------------------------------------------

# Exercise 1 - Square Numbers

numbers = [1, 1, 2, 3, 5, 8, 13, 21, 34, 55]

squared_numbers = [number ** 2 for number in numbers]

print(squared_numbers)

#-----------------------------------------------------------------------------------------------------------------------

# Exercise 2 - Filter Even Numbers

numbers = [1, 1, 2, 3, 5, 8, 13, 21, 34, 55]

even_numbers = [number for number in numbers if number % 2 == 0]

print(even_numbers)

#-----------------------------------------------------------------------------------------------------------------------

# Exercise 3 - find the common numbers between two files and convert them to integers
# (my code, Superior) (her solution is not only retarded but it's also wrong, I'm not even going to include it.)

with open("file1.txt") as file_1, open("file2.txt") as file_2:                      # can access multiple files using one with open()
    numbers_1 = [int(num.strip()) for num in file_1.readlines()]                    # file_1.readlines() is a list so we can use it in a list comprehension directly to to strip endlines and convert to ineger for a cleaner search for commons in the results list
    numbers_2 = [int(num.strip()) for num in file_2.readlines()]
    result = [number_1 for number_1 in numbers_1 if number_1 in numbers_2]
    print(result)


#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
# US States Game line Improvement with list comprehension (based on her code)

# Old lines

        missing_states = []
        for state in all_states:
            if state not in guessed_states:
                missing_states.append(state)

# New line

        missing_states = [state for state in all_states if state not in guessed_states]

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

# Dictionary Comprehensions
# Syntax: new_dict = { new_key : new_value for item in list}                             # to create a dictionary from a list
# Syntax: new_dict = { new_key : new_value for (key, Value) dict.items()}                # to create a dictionary from a dictionary, The dict.items() method in Python returns a view object that displays a LIST of a dictionary's key-value TUPLE pairs.
# Syntax: new_dict = { new_key : new_value for (key, Value) dict.items() if test}

# Example 1

names = ["Alex", "beth", "Caroline", "Dave", "Eleanor", "Freddie"]

from random import randint

students_scores = {student : randint(0, 100) for student in names}

print(students_scores)

# Example 2 - taking the previous dictionary and create a new one with a condition

passed_students = {student:marks for (student, marks) in students_scores.items() if marks > 60 }

# Exercise - 1
# convert a string to a list delimited by space and create a dictionary where the value is the number of letters in the key

sentence = "What is the Airspeed Velocity of an Unladen Swallow"
list_str = sentence.split(" ")
new_dict = {item:len(item) for item in list_str}
print(new_dict)


# Exercise - 2
# change the key of a dictionary temperature from c to f and store it in a new f dictionary

weather_c = {
    "Monday": 12,
    "Tuesday": 14,
    "Wednesday": 15,
    "Thursday": 14,
    "Friday": 21,
    "Saturday": 22,
    "Sunday": 24
}

weather_f = {day:(temp * 9/5) + 32 for (day, temp) in weather_c.items()}

print(weather_f)

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

# iterating through a dataframe with iterrows()

student_dict = {
    "student": ["Angela", "James", "Lily"],
    "score": [56, 76, 98]
}


import pandas

student_data_frame = pandas.DataFrame(student_dict)
# print(student_data_frame)

for (index, row) in student_data_frame.iterrows():                    # iterrows() iterates through rows in pandas
    if row.student == "Angela":                                       # row.student can also be typed row["student"]
        print(row.score)                                              # the row here is a pandas series

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
# Nato Alphabet (my code)

import pandas

nato_alpha = pandas.read_csv("nato_phonetic_alphabet.csv")

phonetic_alphabet_dict =  {row.letter:row.code for (index, row) in  nato_alpha.iterrows()}              # converts data frame to a dictionary where letter column is the key and code is the value
print(phonetic_alphabet_dict)


word = input("word: ").upper()
nato_translation = [phonetic_alphabet_dict[letter] for letter in word]                      # translates your text to the nato alphabet in a list

print(nato_translation)

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

# list() functions


# list()
# list() refers to the built-in constructor used to create a new list object. It can be used in a few ways:

# 1. Creating an empty list.

my_list = list()
print(my_list)

# 2. Converting an iterable to a list: When an iterable (like a string, tuple, or another list) is passed as an argument to list(), it creates a new list containing the elements of that iterable.

# Example 1:

my_string = "hello"
list_from_string = list(my_string)
print(list_from_string)

# Example 2:

my_tuple = (1, 2, 3)
list_from_tuple = list(my_tuple)
print(list_from_tuple)


#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

# set()
#The set() function in Python is a built-in function used to create a set object. A set is an unordered collection of unique and immutable elements

# Key Characteristics of Sets:
# 1. Unordered: Elements in a set do not have a defined order, and you cannot access them by index.
# 2. Unique Elements: Sets automatically remove duplicate elements. If you add multiple instances of the same value, only one will be stored.
# 3. Mutable (the set itself): You can add or remove elements from a set after it's created using methods like add(), remove(), or discard().
# 3. Immutable Elements (contents): The elements within a set must be immutable data types (e.g., numbers, strings, tuples). You cannot store mutable objects like lists or other sets directly as elements of a set, as they are not "hashable." check "Mutability Immutability, Hashability" section below for more details, line: 6217

# Common Uses:
# Removing Duplicates: Efficiently remove duplicate values from a list or other iterable.
# Membership Testing: Check if an element is present in a collection using the in operator, which is highly optimized for sets.
# Mathematical Set Operations: Perform operations like union (union() or |), intersection (intersection() or &), difference (difference() or -), and symmetric difference (symmetric_difference() or ^).

# Example 1

my_empty_set = set()
print(my_empty_set)

# Example 2

my_list = [1, 2, 2, 3, 4, 4, 5]
my_set = set(my_list)  # my_set will be {1, 2, 3, 4, 5}
print(my_set)

# Example 3

my_string = "hello"
char_set = set(my_string)  # char_set will be {'h', 'e', 'l', 'o'} (order may vary)


# Why create an empty set() ?

# 1. Collect unique items dynamically (Removing duplicates),You start with nothing, then add things as you find them:
# Example 4

seen_numbers = set()

for n in [1, 2, 2, 3, 1, 4]:
    seen_numbers.add(n)                 # you can add items to sets with .add

print(seen_numbers)  # {1, 2, 3, 4}

# 2. Fast membership checks, Check if an element is present in a collection using the in operator, which is highly optimized for sets.
# 3. To avoid duplicates automatically

# Example 5

names = set()
names.add("Alice")
names.add("Bob")
names.add("Alice")  # ignored
print(names)        # {'Alice', 'Bob'}


# Example 6 - using list() with set()
# Remove duplicates from list Using set() (Does not preserve order):

my_list = [1, 2, 2, 3, 4, 4, 5]
unique_list = list(set(my_list))            #removed duplicates with set() then converted it back to a list with list()
print(unique_list)                          # Output: [1, 2, 3, 4, 5] (Order may vary)


# Example 7
# Can also create sets manually like so:

my_set = {1, 2, 3, 3, "apple", "banana"}            #notice that I added a duplicate object, 3, but it gets ignored automatically when the line executes
print(my_set)

# Example 8
# You can even see (ignoring of duplicate data) with mixed data types that compare equal:

my_set = {1, True}
print(my_set)       # {1}

# Example 9
# Using the set() constructor: This method is used to create an empty set or to convert an iterable (like a list, tuple, or string) into a set.

empty_set = set()
list_to_set = set([1, 2, 2, 3, 4])          # Duplicates are automatically removed
string_to_set = set("hello")                # Creates a set of individual characters
print(list_to_set, string_to_set)

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

# Dictionary fromkeys() Method

# The fromkeys() method returns a dictionary with the specified keys and the specified value. Syntax: dict.fromkeys(keys, value)
# dict.fromkeys() is a class method in Python that provides a convenient way to create a new dictionary. It takes two arguments:
# 1. keys (required): An iterable (like a list, tuple, or set) containing the elements that will be used as keys in the new dictionary.
# 2. value (optional): An initial value to be assigned to all keys in the new dictionary. If this argument is omitted, the default value assigned to each key will be None.
#How it works:
# The fromkeys() method iterates through the provided keys iterable and creates a new dictionary where each element from keys becomes a key. All these keys are then assigned the same value (either the one you provided or None by default).

# Example 1
# Create a dictionary with default None values

my_list = ['a', 'b', 'c']
my_dict_none = dict.fromkeys(my_list)             # dict.fromkeys() is making each item from the list a key, and assigning None by default as the value to all keys
print(my_dict_none)

# Example 2
# Create a dictionary with a specified value

my_list = ['a', 'b', 'c']
my_dict_value = dict.fromkeys(my_list, 0)     # assigning 0 to all keys
print(my_dict_value)

# Important Considerations: Mutable Objects as Values: If you use a mutable object (like a list or another dictionary) as the value argument, all keys will point to the same instance of that mutable object. Modifying the object through one key will reflect the change when accessing it through any other key. If you need distinct mutable objects for each key, a dictionary comprehension is generally more appropriate.
# what this means if the value was a list or dictionary, all keys will be pointing to the same shared list/dictionary because they are mutable, so making a change on one key will be changed across all other keys as well

# Example 3
# Example with a mutable object (caution advised)

mutable_value = []
my_list = ['a', 'b', 'c']
my_dict_mutable = dict.fromkeys(my_list, mutable_value)
my_dict_mutable['a'].append(1)                              # the list is shared across all keys so changing it here will change it for all keys
print(my_dict_mutable)

# Example 4
# Example with an immutable object

mutable_value = 4
my_list = ['a', 'b', 'c']
my_dict_mutable = dict.fromkeys(my_list, mutable_value)
my_dict_mutable['a'] = 2                                    # this changes only 'a' key because it is an immutable value
print(my_dict_mutable)

# Example 5
# if you want a distinct list/dictionary for each key:
# Use a dictionary comprehension for distinct mutable values
# Note: this does not use the dict.fromkeys() method, we're creating the list manually like so key: []

my_list = ['a', 'b', 'c']
my_dict_distinct_mutable = {key: [] for key in my_list}     #this is dictionary comprehension; creates the dictionary by making the list items into keys with key : and giving them an empty list value []
my_dict_distinct_mutable['a'].append(2)
my_dict_distinct_mutable['c'].append(1)                     # each key now have their own list
print(my_dict_distinct_mutable)

# Example 2 - list(dict.fromkeys())
# Remove duplicates from list Using dict.fromkeys() (Preserves order):

my_list = ["a", "b", "a", "c", "c"]
unique_list = list(dict.fromkeys(my_list))
print(unique_list)

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

# Dictionary items() Method

# The dict.items() method in Python returns a view object that displays a list of a dictionary's key-value tuple pairs.

# main characteristics:

# 1. Returns a View Object: It does not return a new list of tuples, but rather a dynamic view of the dictionary's items. This means that if the dictionary is modified after items() is called, the view object will reflect those changes.
# 2. Contains Key-Value Tuples: Each element in the view object is a tuple, where the first element of the tuple is a key from the dictionary and the second element is its corresponding value.
# 3. No Parameters: The items() method does not take any arguments.
# 4. Usage: It is commonly used for iterating through a dictionary to access both its keys and values simultaneously.

# Example

my_dict = {
    'name': 'Alice',
    'age': 30,
    'city': 'New York'
}

print(my_dict.items())                        #output: dict_items([('name', 'Alice'), ('age', 30), ('city', 'New York')])


for key, value in my_dict.items():              # iterating through the items like so: key for key and value for value
    print(f"Key: {key}, Value: {value}")        # output: Key: name, Value: Alice, etc

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

# zip()

# it iterates the first with the first, 2nd with 2nd of each list, and omits extra items (when strict=False) if one of the lists contains more items (i.e. different lengths), in the example below, 'd' was omitted because list_1 doesn't have a 4th item
# syntax zip(*iterables, strict=False)
# *iterables: One or more iterables (such as lists, tuples, or strings) to be zipped together. This is a required argument.
# strict: (Optional) A boolean value. If set to True, it raises a ValueError if the iterables (like in the example below) have different lengths. The default value is False.

# Example

list1 =[1, 2, 3]
list2 =['a', 'b', 'c', 'd']
result =[(x, y) for x, y in zip(list1, list2)]              # or zip(list1, list2, strict=False)
print(result)


#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

# id() function
# The id() function in Python is a built-in function that returns the "identity" of an object. This identity is an integer that is guaranteed to be unique and constant for that object during its lifetime. In CPython, the most common implementation of Python, this identity corresponds to the object's memory address.
# Key characteristics of id():
# Unique and Constant: Each object has a unique identity, and this identity remains constant as long as the object exists in memory.
# Memory Address (CPython): In CPython, the returned integer is essentially the memory address where the object is stored.
# Object Lifetime: While an object's id is constant during its lifetime, the same id value can be reused by a different object after the original object has been garbage collected.
# Syntax: The function takes a single argument, which is the object whose identity you want to retrieve: id(object).

# Example:

name = "mohammed"
print(id(name))

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

# hash() function

# The built-in hash() function in Python returns an integer hash value for a given object. This hash value serves as a "digital fingerprint" for the object and is primarily used for efficient storage and retrieval in hash-based data structures like dictionaries and sets.
# Purpose: The primary purpose of hash() is to enable fast lookups in hash tables, which are the underlying implementation of Python dictionaries and sets. When you add an item to a dictionary or set, its hash value is used to determine where to store it, allowing for near-constant time average-case lookup performance.
# Input: It takes a single argument, object, which must be hashable.
# Output: It returns an integer representing the hash value of the object.
# Hashable Objects: Only immutable objects like numbers (integers, floats), strings, and tuples (containing only hashable elements) can be hashed. Mutable objects like lists and dictionaries are not hashable because their content can change, which would lead to inconsistent hash values.
# Consistency: For a given hashable object, hash() will return the same hash value within a single Python session. However, the exact hash value may differ across different Python sessions due to security considerations (hash randomization).
# Equality and Hashing: A fundamental principle is that if two objects are equal (i.e., obj1 == obj2 is True), then their hash values must also be equal (i.e., hash(obj1) == hash(obj2) must be True). The reverse is not necessarily true; different objects can, in rare cases, have the same hash value (known as a hash collision).
# Custom Objects: You can define how custom objects are hashed by implementing the __hash__() method within the class definition. This method should return an integer hash value based on the object's immutable attributes.

# Example
# Hashing various immutable objects

print(hash(123))
print(hash("hello"))
print(hash(3.14))
print(hash((1, 2, 3)))

# Attempting to hash a mutable object (will raise an error)
print(hash([1, 2, 3]))

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

# hashing Integers and Range() hashing and equality

# 1.
# why is print(hash(123)) is returning the hash as 123 every time?

# it’s completely intentional. Normally, different objects produce different hash values, and hash values are consistent for the lifetime of the program.

# Special case for integers
# For integers (int type), the hash of an integer is the integer itself.

hash(n) == n

# So:

print(hash(123))      # 123
print(hash(-5))       # -5
print(hash(0))        # 0

# This is by design in Python.

# Why?
# Integers are already perfectly good hash values:
# - They’re unique (each distinct integer has a unique numeric value)
# - They’re immutable (can’t change)
# - They’re already fast to compare and store

# this works for float as well

print(hash(123.0))  # 123
print(hash(123))    # 123

# 2.
# Range Equality
# is range mutable or immutable
# range objects are immutable.
# Two range objects are considered equal if they represent the same sequence of numbers — even if they’re different objects.

# Example

range(1, 5) == range(1, 5)   #  True
range(1, 5) == range(0, 4)   #  False

# And since they’re equal, they have the same hash:
# hash(range(1, 5)) == hash(range(1, 5))   # True

# that means:

my_set = {range(1, 5), range(1, 5)}
print(my_set)                           # {range(1, 5)}  ← duplicates removed (same hash & equality)


#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

# Mutability Immutability, Hashability

# 1.
# immutable
# An immutable data type is something that cannot be changed after it’s created.
# Once you make it, its contents are fixed forever — any “change” actually creates a new object in memory.

# Mutables: list, dict, set
# Immutables: tuple, str, int, float, bool, range



# Example: immutable string

name = "John"
print(id(name))        # e.g., 140123456
name += " Doe"         # creates a *new* string
print(id(name))        # different memory address

#Even though you “changed” the string, Python actually created a new one — the original "John" never changed.

# Example: mutable list

numbers = [1, 2, 3]
print(id(numbers))
numbers.append(4)      # modifies in place
print(id(numbers))     # same memory address

# Lists can be modified in place (by adding, removing, or changing elements). That’s what “mutable” means.

# 2.
# Why must set elements be immutable?
# Sets in Python are built on hash tables — a data structure that uses hashing for very fast lookups (in checks, adding, removing).
# To work correctly, a set needs each element to have a fixed hash value.
# If an object could change after being added, its hash could change — and the set would lose track of where it’s stored.


# 3.
# What does hashable mean?
# An object is hashable if it has:
# - A valid, unchanging __hash__() value
# - And can be compared to others with __eq__()
# Basically, being hashable means: “It can serve as a reliable key or element in a set or dictionary.”

hash(42)          # works
hash("hello")     # works
hash((1, 2, 3))   # works
hash([1, 2, 3])   # TypeError: unhashable type: 'list'

# 4.
# Why lists and sets are not hashable
# Because they’re mutable — their contents can change and when the content changes, the hash also changes, therefore it's forbidden

# 5.
# So what can go inside a set?
# Only immutable, hashable things
# int, str, tuple, bool, range

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

# All Python data types

# Python has around 40 data types, depending on how you count them — but the core built-in types are 14 major ones, grouped into numeric, sequence, set, and mapping categories.

#1. Basic (Scalar / Primitive) Data Types

# Type          # Example               Description
# int           42, -5                  Integer numbers (arbitrary precision)
# float         3.14, -2.7              Decimal (floating-point) numbers
# complex       2 + 3j                  Complex numbers (real + imaginary)
# bool          True, False             Boolean values (subclass of int)
# NoneType      None                    Represents “no value” or null


# 2. Sequence Data Types

# Type          # Example               Description
# str           "Hello"                 Immutable text sequence
# list          [1, 2, 3]               Mutable ordered collection
# tuple         (1, 2, 3)               Immutable ordered collection
# range         range(5)                Immutable sequence of numbers
# bytes         b"abc"                  Immutable sequence of bytes
# bytearray     bytearray(b"abc")       Mutable sequence of bytes
# memoryview    memoryview(b"abc")      Memory-efficient view of byte data


# 3. Set Data Types

# Type          # Example               Description
# set           {1, 2, 3}               Mutable unordered collection of unique elements
# frozenset     frozenset({1, 2, 3})    Immutable version of a set


# 4. Mapping Data Types

# Type          # Example               Description
# dict          {"a": 1, "b": 2}        Key–value mapping


# 5. Callable / Functional Data Types

# Type                          # Example               Description
# function                      def f(): ...            User-defined function
# builtin_function_or_method    len, print              Built-in C-implemented functions
# lambda                        lambda x: x*2           Anonymous function
# method                        "abc".upper             Bound method object


# 6. Iterator / Generator / Coroutine Types

# Type                          # Example                   Description
# generator                     (x for x in range(3))       Lazily evaluated iterator
# iterator                      iter([1,2,3])               Object implementing __next__()
# async_generator               async def gen(): yield      Asynchronous generator
# coroutine                     async def f(): ...          Asynchronous coroutine
# async_iterator                aiter(obj)                  Async object implementing __anext__()


# 7. Specialized / Utility Data Types

# Type                                  # Example                                   Description
# type                                  type(123)                                   The class of all classes
# object                                object()                                    The root base type of all objects
# ellipsis                              ...                                         Used in slicing and annotations
# NotImplementedType                    NotImplemented                              Signals unsupported operation
# slice                                 slice(0, 5, 2)                              Represents slice object
# code                                  Compiled code objects (e.g. compile(...))   Represents Python bytecode
# module                                import math                                 Module object
# class                                 class A: pass                               User-defined class
#classmethod, staticmethod, property    —                                           Special method wrappers


# 8. From types and collections Modules (standard library core)
# These are still considered part of Python’s data type system, but live in standard modules.

# Type                                  # Example               Description
# SimpleNamespace                       types                   Namespace container
# MappingProxyType                      types                   Read-only dict view
# deque                                 collections             Double-ended queue
# defaultdict                           collections             Dict with default factory
# Counter                               collections             Multiset of elements
# OrderedDict                           collections             Dict that remembers order (now redundant in 3.7+)
# ChainMap                              collections             Combine multiple dicts logically
# namedtuple                            collections             Tuple subclass with named fields


# Immutable types:
# int, float, complex, bool, str, tuple, frozenset, bytes, range, NoneType

# Mutable types:
# list, dict, set, bytearray


# in summary:

# | Category                  | Count (approx.)              |
# | ------------------------- | ---------------------------- |
# | Basic (scalar)            | 5                            |
# | Sequence                  | 7                            |
# | Set                       | 2                            |
# | Mapping                   | 1                            |
# | Callable                  | 4                            |
# | Iterator/Generator        | 5                            |
# | Special/Utility           | 8                            |
# | Collections/Types module  | 8                            |
# | **Total (core + stdlib)** | **≈ 40 built-in data types** |


#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

# tkinter

import tkinter

window = tkinter.Tk()
window.title("GUI Program")
window.minsize(500, 300)


my_label = tkinter.Label(text="Label", font=("Arial", 24, "bold"))      # creating a label on the screen
my_label.pack(side="bottom", expand=True)                               # will make the label show on the screen, centered by default, side will allow you to choose where to show it, expand will allow it to take up all the area in the screen (allowing it to be bang in the center)


window.mainloop()

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

# Functions with default Arguments (Optional Arguments)

def function(a=1, b=2, c=3)     # parameter values were declared here by default, meaning if its get called without specifying the arguments, it will take the default values specifed here

function()                  # i can call the function normally because all all its arguments are optional and are already set in its declaration

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

# Unlimited Positional Arguments

def add(*args):
    for n in args:
        print(n)


# Example

def add(*args):                 #puts all arguments into a tuple, * is all we need to define it, args is the default naming for the arugments, but we can call it anything else
    total = 0
    for n in args:              # since it's a tuple, we can use indexes like args[0]
        total += n
        return total


print(add(2, 8, 9))

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

# Unlimited Keyword Arguments

def calculate(n, **kwargs):
    n += kwargs["add"]                      # this is saying when the argument is called "add" add it with n
    n *= kwargs["multiply"]                 # basically we're looking if there is an argument called multiply defined when the function was called (will return a key error if it was not defined in the function call)

calculate(2, add=3, multiply=5)          # we're defining the arguments in the function call itself, kwargs is basically now a dictionary, type dict, now n is added to 3 then n=5 is mutiplied with 5


# Example with a class

class Car:
    def __init__(self, **kwargs):
        self.make = kwargs["make"]
        self.model = kwargs["model"]
        self.type = kwargs["type"]              # this will cause a key error because "type" is not defined in the function call
        self.year = kwargs.get("year")          # use get() as an alternative way to retrieve the value, will not cause an error if it wasn't defined in function call, returns None if we try to print it

my_car = Car(make="Toyota", model="Corolla")

print(my_car.model)

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

# tkinter, a button to change the label text based on the input in a textbox

from tkinter import *

window = Tk()
window.title("GUI Program")
window.minsize(500, 300)


my_label = Label(text="Label", font=("Arial", 24, "bold"))
my_label.pack()

my_label["text"] = "New Text"           #changing the label text (using dictionary syntax because this is a multi keyword argument)
my_label.config(text="New Text")        # can also change it with config()


def button_clicked():
    text_entered = entry.get()                      # get() will get the text typed in the textbox
    my_label.config(text=text_entered)              # text entered will change the label

button = Button(text="Click Me", command=button_clicked)                # command argument can execute a function when the button is pressed
button.pack()



entry = Entry(width=20)
entry.pack()


window.mainloop()

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

# all other tkinter elements

from tkinter import *

#Creating a new window and configurations
window = Tk()
window.title("Widget Examples")
window.minsize(width=500, height=500)

#Labels
label = Label(text="This is old text")
label.config(text="This is new text")
label.pack()

#Buttons
def action():
    print("Do something")

#calls action() when pressed
button = Button(text="Click Me", command=action)
button.pack()

#Entries
entry = Entry(width=30)

#Add some text to begin with
entry.insert(END, string="Some text to begin with.")

#Gets text in entry
print(entry.get())
entry.pack()

#Text
text = Text(height=5, width=30)

#Puts cursor in textbox.
text.focus()

#Adds some text to begin with.
text.insert(END, "Example of multi-line text entry.")        # END means: Insert the string 'Example of multi-line text entry.' at the end of whatever text is already inside the Text widget.

#Gets current value in textbox at line 1, character 0
print(text.get("1.0", END))                                 # gets text from: "1.0" line one, character 0, to the END of the string
text.pack()

#Spinbox
def spinbox_used():
    #gets the current value in spinbox.
    print(spinbox.get())
spinbox = Spinbox(from_=0, to=10, width=5, command=spinbox_used)                # from_ and to are the minimum and maximum allowed numbers in the spinbox
spinbox.pack()

#Scale
#Called with current scale value.
def scale_used(value):
    print(value)
scale = Scale(from_=0, to=100, command=scale_used)                              # same as spinbox
scale.pack()

#Checkbutton
def checkbutton_used():
    #Prints 1 if On button checked, otherwise 0.
    print(checked_state.get())

#variable to hold on to checked state, 0 is off, 1 is on.
checked_state = IntVar()                                                                        # is a special variable class that holds an integer value and can be linked to a widget, like a checkbox, so that the widget and variable stay in sync automatically.
checkbutton = Checkbutton(text="Is On?", variable=checked_state, command=checkbutton_used)      # variable → Connects that variable to the Checkbutton.
checked_state.get()                                                                             # When the box is unchecked, checked_state.get() returns 0. When the box is checked, it returns 1.
checkbutton.pack()

#Radiobutton
def radio_used():
    print(radio_state.get())

#Variable to hold on to which radio button value is checked.
radio_state = IntVar()
radiobutton1 = Radiobutton(text="Option1", value=1, variable=radio_state, command=radio_used)       # value is the ID of the grouped radio buttons
radiobutton2 = Radiobutton(text="Option2", value=2, variable=radio_state, command=radio_used)
radiobutton1.pack()
radiobutton2.pack()


#Listbox
def listbox_used(event):                                             # Tkinter will call listbox_used(event) and pass an event object. Inside listbox_used you can access: event.widget → the widget that triggered the event (here, the listbox) Or directly use the listbox variable from outer scope.
    # Gets current selection from listbox
    print(listbox.get(listbox.curselection()))                       # curselection() = current selection, It returns a tuple of the index numbers of the currently selected items in the listbox., ex. (1,)

listbox = Listbox(height=4)
fruits = ["Apple", "Pear", "Orange", "Banana"]
for item in fruits:
    listbox.insert(fruits.index(item), item)               # listbox.insert(index, item) inserts item at the given index (0-based), fruits.index(item) fruits.index(item) returns the first index of item in the fruits list
listbox.bind("<<ListboxSelect>>", listbox_used)                     # Binds the virtual event <<ListboxSelect>> to the function listbox_used, <<ListboxSelect>> fires whenever the selection in the listbox changes (user click, arrow keys, programmatic change).
listbox.pack()
window.mainloop()

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

# index() function

# The index() method of a list returns the position (index number) of the first occurrence of a given value inside the list.

# Basic syntax:

list.index(value, start, end)

# value - The item you’re searching for
# start (optional) - Where to start searching
# end (optional) - Where to stop searching

# If start and end aren’t given, it searches the whole list.

# Example -

fruits = ["apple", "banana", "cherry", "banana"]

print(fruits.index("banana"))      # → 1
print(fruits.index("banana", 2))   # → 3 (starts searching at index 2)

# Note: If the value isn’t found, Python raises a ValueError:

colors = ["red", "green", "blue"]
print(colors.index("yellow"))  # ValueError: 'yellow' is not in list

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

# pack(), place(), and grid()

from tkinter import *


def button_clicked():
    text_entered = entry.get()
    my_label.config(text=text_entered)


window = Tk()
window.title("GUI Program")
window.minsize(500, 300)
window.config(padx=20, pady=20)                                  # adding padding, can be added to any widget


my_label = Label(text="Label", font=("Arial", 24, "bold"))
my_label.config(text="New Text")
my_label.place(x=100, y=50)                                     # place is precise, uses coordinates

button = Button(text="Click Me", command=button_clicked)
button.grid(column=0, row=0)                                    # grid can't be mixed with pack, it will cause a slave error

entry = Entry(width=20)
entry.grid(column=1, row=1)                                     # grid is relative to other widgets, you can't put something at row 2 if there's no widget at row 1 for example, it will take row 1


window.mainloop()

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

# Mile to Km Converter GUI Program (All me)

from tkinter import *

def button_calculate():
    mile_value = int(textbox.get())
    km_value = mile_value * 1.6
    label_km_int.config(text=km_value)

window = Tk()
window.title("Mile to Km Converter")
window.minsize(200, 200)
window.config(padx=20, pady=20)

textbox = Entry(width=20)
textbox.grid(column=1, row=0)

label_miles = Label(text="Miles")
label_miles.grid(column=2, row=0)

label_equal = Label(text="is equal to")
label_equal.grid(column=0, row=1)

label_km_int = Label(text="0")
label_km_int.grid(column=1, row=1)

label_km = Label(text="Km")
label_km.grid(column=2, row=1)

button_calc = Button(text="Calculate", command=button_calculate)
button_calc.grid(column=1, row=2)

window.mainloop()

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

# my countdown timer

import time, math

time_ = 300
minutes = time_ // 60
seconds = time_ % 5
while time_ >= 0:
    if seconds < 10:
        print(f"{minutes}:0{seconds}")
    else:
        print(f"{minutes}:{seconds}")
    time.sleep(1)
    time_ -= 1
    seconds-= 1

    if seconds < 0:
        minutes -= 1
        seconds = 59

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

# Pomodoro (my code)

from tkinter import *

# ---------------------------- CONSTANTS ------------------------------- #
PINK = "#e2979c"
RED = "#e7305b"
GREEN = "#9bdeac"
YELLOW = "#f7f5dd"
FONT_NAME = "Courier"
WORK_MIN = 25
SHORT_BREAK_MIN = 5
LONG_BREAK_MIN = 30
reps = 0
check = ""
start_button_timer = None                       # we added this so that we can access it from count_down() function and use it in reset_timer() function, we assigned it to None for now

# ---------------------------- TIMER RESET ------------------------------- #

def reset_timer():
    global start_button_timer, reps
    window.after_cancel(start_button_timer)             # widget.after_cancel() is used to stop widget.after() function in tkinter
    canvas.itemconfig(timer_text, text="00:00")            # to change canvas item you assign the item to a variable then say canvas_name.itemconfig(item_variable, *args)
    label_timer.config(text="Timer")
    label_check.config(text="")
    reps = 0

# ---------------------------- TIMER MECHANISM ------------------------------- #

def start_timer():
    global reps
    work_sec = WORK_MIN * 60
    short_break_sec = SHORT_BREAK_MIN * 60
    long_break_sec = LONG_BREAK_MIN * 60

    if reps in (0, 2, 4, 6):
        label_timer.config(text="Work", fg=GREEN)
        count_down(work_sec)
        reps += 1
    elif reps in (1, 3, 5):
        label_timer.config(text="Break", fg=PINK)
        count_down(short_break_sec)
        reps += 1
    elif reps == 7:
        label_timer.config(text="Break", fg=RED)
        count_down(long_break_sec)
        reps = 0


# ---------------------------- COUNTDOWN MECHANISM ------------------------------- #

def count_down(count_):                                                 # doing a loop with a recursive function to not overwrite the tkinter mainloop
    global check, start_button_timer
    count_min = count_ // 60                                            # retarded teacher used math.floor(), Booooo
    count_sec = count_ % 60
    if count_sec < 10:
        count_sec = f"0{count_sec}"
    if count_min < 10:
        count_min = f"0{count_min}"

    canvas.itemconfig(timer_text, text=f"{count_min}:{count_sec}")            # to change canvas item you assign the item to a variable then say canvas_name.itemconfig(item_variable, *args)
    if count_ >= 0:
        start_button_timer = window.after(1, count_down, count_ - 1)                  # counting down after each 1000 ms (1 second), widget.after() is a method that executes a command (function) after a time delay
    else:
        start_timer()                                           # go to the next time period after time runs out
        if reps % 2 == 0:
            check += "✓"
            label_check.config(text=check)

# ---------------------------- UI SETUP ------------------------------- #

window = Tk()
window.title("Pomodoro")
window.config(padx=100, pady=50, bg=YELLOW)        # bg= is background color
window.minsize(410, 334)
canvas = Canvas(width=200, height=224, bg=YELLOW, highlightthickness=0)                 # Canvas is used to add images and other graphics to the window, highlightthickness is for the stroke/border of the canvas
image = PhotoImage(file="tomato.png")                                                   # PhotoImage is how you access image files with tkinter
canvas.create_image(100, 112, image=image)                                        # create_image to add the image, takes x and y coordinates and the image file (takes a variable)
timer_text = canvas.create_text(100, 130, text="00:00", fill="white", font=(FONT_NAME, 35, "bold"))      # creating text in canvas
canvas.grid(column=1, row= 1)
label_timer = Label(text="Timer", font=(FONT_NAME, 35, "bold"), fg=GREEN, bg=YELLOW)
label_timer.grid(column=1, row=0)

button_start = Button(text="Start", command=start_timer)
button_start.grid(column=0, row=2)

button_Reset = Button(text="Reset", command=reset_timer)
button_Reset.grid(column=2, row=2)


label_check = Label(font=(FONT_NAME, 12, "bold"), fg=GREEN, bg=YELLOW)
label_check.grid(column=1, row=3)

window.mainloop()

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

# Pomodoro (her code)

from tkinter import *
import math
# ---------------------------- CONSTANTS ------------------------------- #
PINK = "#e2979c"
RED = "#e7305b"
GREEN = "#9bdeac"
YELLOW = "#f7f5dd"
FONT_NAME = "Courier"
WORK_MIN = 1
SHORT_BREAK_MIN = 5
LONG_BREAK_MIN = 20
reps = 0
timer = None

# ---------------------------- TIMER RESET ------------------------------- #

def reset_timer():
    window.after_cancel(timer)
    canvas.itemconfig(timer_text, text="00:00")
    title_label.config(text="Timer")
    check_marks.config(text="")
    global reps
    reps = 0


# ---------------------------- TIMER MECHANISM ------------------------------- #

def start_timer():
    global reps
    reps += 1

    work_sec = WORK_MIN * 60
    short_break_sec = SHORT_BREAK_MIN * 60
    long_break_sec = LONG_BREAK_MIN * 60

    if reps % 8 == 0:
        count_down(long_break_sec)
        title_label.config(text="Break", fg=RED)
    elif reps % 2 == 0:
        count_down(short_break_sec)
        title_label.config(text="Break", fg=PINK)
    else:
        count_down(work_sec)
        title_label.config(text="Work", fg=GREEN)


# ---------------------------- COUNTDOWN MECHANISM ------------------------------- #
def count_down(count):

    count_min = math.floor(count / 60)
    count_sec = count % 60
    if count_sec < 10:
        count_sec = f"0{count_sec}"

    canvas.itemconfig(timer_text, text=f"{count_min}:{count_sec}")
    if count > 0:
        global timer
        timer = window.after(1000, count_down, count - 1)
    else:
        start_timer()
        marks = ""
        work_sessions = math.floor(reps/2)
        for _ in range(work_sessions):
            marks += "✔"
        check_marks.config(text=marks)


# ---------------------------- UI SETUP ------------------------------- #
window = Tk()
window.title("Pomodoro")
window.config(padx=100, pady=50, bg=YELLOW)


title_label = Label(text="Timer", fg=GREEN, bg=YELLOW, font=(FONT_NAME, 50))
title_label.grid(column=1, row=0)

canvas = Canvas(width=200, height=224, bg=YELLOW, highlightthickness=0)
tomato_img = PhotoImage(file="tomato.png")
canvas.create_image(100, 112, image=tomato_img)
timer_text = canvas.create_text(100, 130, text="00:00", fill="white", font=(FONT_NAME, 35, "bold"))
canvas.grid(column=1, row=1)

start_button = Button(text="Start", highlightthickness=0, command=start_timer)
start_button.grid(column=0, row=2)

reset_button = Button(text="Reset", highlightthickness=0, command=reset_timer)
reset_button.grid(column=2, row=2)

check_marks = Label(fg=GREEN, bg=YELLOW)
check_marks.grid(column=1, row=3)

window.mainloop()

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

# Password Manager GUI (my code)

from tkinter import messagebox
import pyperclip

# ---------------------------- PASSWORD GENERATOR ------------------------------- #

import random

letters = [ 'a', 'b', 'c', 'd', 'e', 'f', 'g', 'h', 'i', 'j', 'k', 'l', 'm', 'n',
            'o', 'p', 'q', 'r', 's', 't', 'u', 'v', 'w', 'x', 'y', 'z', 'A', 'B',
            'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P',
            'Q,' 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z']

numbers = ['0','1','2','3','4','5','6','7','8','9']
symbols = ['!', '#', '$', '%', '&', '(', ')', '*', '+']

def generate_pass():
    password = ""
    for letter in range(4):
        letter = random.choice(letters)
        password += letter

    for symbol in range(4):
        symbol = random.choice(symbols)
        password += symbol

    for number in range(4):
        number = random.choice(numbers)
        password += number

    password = ''.join(random.sample(password, len(password)))
    pyperclip.copy(password)                                            # this will copy password to clipboard, NEAT!
    password_field.delete(0, END)
    password_field.insert(0, password)

# ---------------------------- SAVE PASSWORD ------------------------------- #

def save():
    website_field_text = website_field.get()
    user_name_field_text = user_name_field.get()
    password_field_text = password_field.get()


    if website_field_text == "" or user_name_field_text == "" or password_field_text == "":
        messagebox.showinfo(title="Empty Fields", message="Please fill out all the fields!")                        # create popups with messagebox

    else:
        is_ok = messagebox.askokcancel(title=website_field_text, message=f" These are the details entered \n"
                                                                 f"Email {user_name_field_text}\n"
                                                                 f"Password: {password_field_text}\n"
                                                                 f"Is it ok to save?")

        if is_ok:
            with open("data.txt", mode= "a") as file_1:
                file_1.write(f"{website_field_text} | {user_name_field_text} | {password_field_text}\n")
                website_field.delete(0, END)                            # delete contents of text field from start 0 to finish END
                user_name_field.delete(0, END)
                password_field.delete(0, END)

# ---------------------------- UI SETUP ------------------------------- #

from tkinter import *                   # imports all Classes and Constants but not modules

window = Tk()
window.title("Password Manager")
window.config(padx=50, pady=50)

canvas = Canvas(width=200, height=200)
image = PhotoImage(file="logo.png")
canvas.create_image(100, 100, image=image)
canvas.grid(column=1, row=0)

website = Label(text="website: ")
website.grid(column=0, row=1)

user_name = Label(text="Email/Username: ")
user_name.grid(column=0, row=2)

password = Label(text="Password: ")
password.grid(column=0, row=3)

website_field = Entry(width=45)
website_field.grid(column=1,row= 1, columnspan=2)
website_field.focus()                                                       # lets the typing cursor focus this text field on app launch

user_name_field = Entry(width=45)
user_name_field.grid(column=1,row= 2, columnspan=2)
user_name_field.insert(0, "mohammad@gmail.com")                 # inserts text to textfield, syntax: insert(index, string) index can also be END (tkinter constant representing the end of the string)

password_field = Entry(width=27)
password_field.grid(column=1,row= 3, columnspan=1)


generate_password = Button(text="Generate Password", width=14, command=generate_pass)
generate_password.grid(column=2, row=3)

add_pass = Button(text="Add", width=38, command=save)
add_pass.grid(column=1, row=4, columnspan=2)


window.mainloop()

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

# Password Manager GUI (her code)

from tkinter import *
from tkinter import messagebox
from random import choice, randint, shuffle
import pyperclip

# ---------------------------- PASSWORD GENERATOR ------------------------------- #

#Password Generator Project
def generate_password():
    letters = ['a', 'b', 'c', 'd', 'e', 'f', 'g', 'h', 'i', 'j', 'k', 'l', 'm', 'n', 'o', 'p', 'q', 'r', 's', 't', 'u', 'v', 'w', 'x', 'y', 'z', 'A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z']
    numbers = ['0', '1', '2', '3', '4', '5', '6', '7', '8', '9']
    symbols = ['!', '#', '$', '%', '&', '(', ')', '*', '+']

    password_letters = [choice(letters) for _ in range(randint(8, 10))]
    password_symbols = [choice(symbols) for _ in range(randint(2, 4))]
    password_numbers = [choice(numbers) for _ in range(randint(2, 4))]

    password_list = password_letters + password_symbols + password_numbers
    shuffle(password_list)

    password = "".join(password_list)
    password_entry.insert(0, password)
    pyperclip.copy(password)

# ---------------------------- SAVE PASSWORD ------------------------------- #
def save():

    website = website_entry.get()
    email = email_entry.get()
    password = password_entry.get()

    if len(website) == 0 or len(password) == 0:
        messagebox.showinfo(title="Oops", message="Please make sure you haven't left any fields empty.")
    else:
        is_ok = messagebox.askokcancel(title=website, message=f"These are the details entered: \nEmail: {email} "
                                                      f"\nPassword: {password} \nIs it ok to save?")
        if is_ok:
            with open("data.txt", "a") as data_file:
                data_file.write(f"{website} | {email} | {password}\n")
                website_entry.delete(0, END)
                password_entry.delete(0, END)


# ---------------------------- UI SETUP ------------------------------- #

window = Tk()
window.title("Password Manager")
window.config(padx=50, pady=50)

canvas = Canvas(height=200, width=200)
logo_img = PhotoImage(file="logo.png")
canvas.create_image(100, 100, image=logo_img)
canvas.grid(row=0, column=1)

#Labels
website_label = Label(text="Website:")
website_label.grid(row=1, column=0)
email_label = Label(text="Email/Username:")
email_label.grid(row=2, column=0)
password_label = Label(text="Password:")
password_label.grid(row=3, column=0)

#Entries
website_entry = Entry(width=35)
website_entry.grid(row=1, column=1, columnspan=2)
website_entry.focus()
email_entry = Entry(width=35)
email_entry.grid(row=2, column=1, columnspan=2)
email_entry.insert(0, "angela@gmail.com")
password_entry = Entry(width=21)
password_entry.grid(row=3, column=1)

# Buttons
generate_password_button = Button(text="Generate Password", command=generate_password)
generate_password_button.grid(row=3, column=2)
add_button = Button(text="Add", width=36, command=save)
add_button.grid(row=4, column=1, columnspan=2)

window.mainloop()


#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
# Password Manager GUI Improved with JSON, Exception Handling and Search password function (my code)

from tkinter import messagebox
import pyperclip
import json

# ---------------------------- PASSWORD GENERATOR ------------------------------- #

import random

letters = [ 'a', 'b', 'c', 'd', 'e', 'f', 'g', 'h', 'i', 'j', 'k', 'l', 'm', 'n',
            'o', 'p', 'q', 'r', 's', 't', 'u', 'v', 'w', 'x', 'y', 'z', 'A', 'B',
            'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P',
            'Q,' 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z']

numbers = ['0','1','2','3','4','5','6','7','8','9']
symbols = ['!', '#', '$', '%', '&', '(', ')', '*', '+']

def generate_pass():
    password = ""
    for letter in range(4):
        letter = random.choice(letters)
        password += letter

    for symbol in range(4):
        symbol = random.choice(symbols)
        password += symbol

    for number in range(4):
        number = random.choice(numbers)
        password += number

    password = ''.join(random.sample(password, len(password)))
    pyperclip.copy(password)
    password_field.delete(0, END)
    password_field.insert(0, password)

# ---------------------------- SAVE PASSWORD ------------------------------- #

def save():
    website_field_text = website_field.get()
    user_name_field_text = user_name_field.get()
    password_field_text = password_field.get()
    new_data = \
        {
            website_field_text:
                {
                    "email": user_name_field_text,
                    "password": password_field_text
                }
        }


    if website_field_text == "" or user_name_field_text == "" or password_field_text == "":
        messagebox.showinfo(title="Empty Fields", message="Please fill out all the fields!")

    else:
        is_ok = messagebox.askokcancel(title=website_field_text, message=f" These are the details entered \n"
                                                                 f"Email {user_name_field_text}\n"
                                                                 f"Password: {password_field_text}\n"
                                                                 f"Is it ok to save?")

        if is_ok:
            try:
                with open("data.json", mode= "r") as file_1:
                    data = json.load(file_1)                                  # reading old data - reading json files (it converts it to a python dictionary)
                    data.update(new_data)                                     # updating old data - this is how to update json files, new_data is a dictionary, it will append to the old data, not replace them
            except FileNotFoundError:
                messagebox.showinfo(title="File not found", message="File not found, creating one now.")
                with open("data.json", mode="w") as file_1:
                    json.dump(new_data, file_1, indent=4)                     # Creating the file the first time and fill it with data
            else:
                with open("data.json", mode="w") as file_1:
                    json.dump(data, file_1, indent=4)                         # saving the updated data - json.dump writes to the json file, indent beautifies the code
            finally:
                website_field.delete(0, END)
                password_field.delete(0, END)


# ---------------------------- SEARCH WEBSITE ------------------------------ #

def find_password():
    website_field_text = website_field.get()
    try:
        with open("data.json", mode="r") as file_1:
            data = json.load(file_1)
            messagebox.showinfo(title=f"{website_field_text} Credentials", message=f"Email: {data[website_field_text]["email"]}\nPassword: {data[website_field_text]["password"]}")
    except KeyError:
        messagebox.showerror(title="Entry not found.", message="Entry not found.")
    except FileNotFoundError:
        messagebox.showerror(title="File not found.", message="File not found, please enter data first.")

# ---------------------------- UI SETUP ------------------------------- #

from tkinter import *

window = Tk()
window.title("Password Manager")
window.config(padx=50, pady=50)

canvas = Canvas(width=200, height=200)
image = PhotoImage(file="logo.png")
canvas.create_image(100, 100, image=image)
canvas.grid(column=1, row=0)

website = Label(text="website: ")
website.grid(column=0, row=1)

user_name = Label(text="Email/Username: ")
user_name.grid(column=0, row=2)

password = Label(text="Password: ")
password.grid(column=0, row=3)

website_field = Entry(width=27)
website_field.grid(column=1,row= 1)
website_field.focus()

user_name_field = Entry(width=45)
user_name_field.grid(column=1,row= 2, columnspan=2)
user_name_field.insert(0, "mohammad@gmail.com")

password_field = Entry(width=27)
password_field.grid(column=1,row= 3, columnspan=1)


generate_password = Button(text="Generate Password", width=14, command=generate_pass)
generate_password.grid(column=2, row=3)

add_pass = Button(text="Add", width=38, command=save)
add_pass.grid(column=1, row=4, columnspan=2)

search_pass = Button(text="Search", width=14, command=find_password)
search_pass.grid(column=2, row=1)

window.mainloop()

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
# Password Manager GUI Improved with JSON, Exception Handling and Search password function (her code)

from tkinter import *
from tkinter import messagebox
from random import choice, randint, shuffle
import pyperclip
import json

# ---------------------------- PASSWORD GENERATOR ------------------------------- #

#Password Generator Project
def generate_password():
    letters = ['a', 'b', 'c', 'd', 'e', 'f', 'g', 'h', 'i', 'j', 'k', 'l', 'm', 'n', 'o', 'p', 'q', 'r', 's', 't', 'u', 'v', 'w', 'x', 'y', 'z', 'A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z']
    numbers = ['0', '1', '2', '3', '4', '5', '6', '7', '8', '9']
    symbols = ['!', '#', '$', '%', '&', '(', ')', '*', '+']

    password_letters = [choice(letters) for _ in range(randint(8, 10))]
    password_symbols = [choice(symbols) for _ in range(randint(2, 4))]
    password_numbers = [choice(numbers) for _ in range(randint(2, 4))]

    password_list = password_letters + password_symbols + password_numbers
    shuffle(password_list)

    password = "".join(password_list)
    password_entry.insert(0, password)
    pyperclip.copy(password)

# ---------------------------- SAVE PASSWORD ------------------------------- #
def save():

    website = website_entry.get()
    email = email_entry.get()
    password = password_entry.get()
    new_data = {
        website: {
            "email": email,
            "password": password,
        }
    }

    if len(website) == 0 or len(password) == 0:
        messagebox.showinfo(title="Oops", message="Please make sure you haven't left any fields empty.")
    else:
        try:
            with open("data.json", "r") as data_file:
                #Reading old data
                data = json.load(data_file)
        except FileNotFoundError:
            with open("data.json", "w") as data_file:
                json.dump(new_data, data_file, indent=4)
        else:
            #Updating old data with new data
            data.update(new_data)

            with open("data.json", "w") as data_file:
                #Saving updated data
                json.dump(data, data_file, indent=4)
        finally:
            website_entry.delete(0, END)
            password_entry.delete(0, END)


# ---------------------------- FIND PASSWORD ------------------------------- #
def find_password():
    website = website_entry.get()
    try:
        with open("data.json") as data_file:
            data = json.load(data_file)
    except FileNotFoundError:
        messagebox.showinfo(title="Error", message="No Data File Found.")
    else:
        if website in data:
            email = data[website]["email"]
            password = data[website]["password"]
            messagebox.showinfo(title=website, message=f"Email: {email}\nPassword: {password}")
        else:
            messagebox.showinfo(title="Error", message=f"No details for {website} exists.")


# ---------------------------- UI SETUP ------------------------------- #

window = Tk()
window.title("Password Manager")
window.config(padx=50, pady=50)

canvas = Canvas(height=200, width=200)
logo_img = PhotoImage(file="logo.png")
canvas.create_image(100, 100, image=logo_img)
canvas.grid(row=0, column=1)

#Labels
website_label = Label(text="Website:")
website_label.grid(row=1, column=0)
email_label = Label(text="Email/Username:")
email_label.grid(row=2, column=0)
password_label = Label(text="Password:")
password_label.grid(row=3, column=0)

#Entries
website_entry = Entry(width=21)
website_entry.grid(row=1, column=1)
website_entry.focus()
email_entry = Entry(width=35)
email_entry.grid(row=2, column=1, columnspan=2)
email_entry.insert(0, "angela@gmail.com")
password_entry = Entry(width=21)
password_entry.grid(row=3, column=1)

# Buttons
search_button = Button(text="Search", width=13, command=find_password)
search_button.grid(row=1, column=2)
generate_password_button = Button(text="Generate Password", command=generate_password)
generate_password_button.grid(row=3, column=2)
add_button = Button(text="Add", width=36, command=save)
add_button.grid(row=4, column=1, columnspan=2)

window.mainloop()

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

# Flash Card GUI App (my code) (run in respective project)

import random
from tkinter import *
import pandas

BACKGROUND_COLOR = "#B1DDC6"

try:
    pd = pandas.read_csv("data/to_learn_words.csv")
except FileNotFoundError:
    pd = pandas.read_csv("data/french_words.csv")
pd_records = pd.to_dict(orient="records")                            # this will turn the dataframe to a list of dictionaries, each item in the list is a row, contained with it the headers from each column and the corresponding item for each column as the value for that key, in our example it looks like so: [{Column1: row1, column2: row1}, {Column1: row2, column2: row2}, etc
chosen_word = {}

def next_card():
    global chosen_word, flip_timer
    window.after_cancel(flip_timer)                     # cancel the previous timer when card is flipped
    chosen_word = random.choice(pd_records)
    french_word = chosen_word["French"]
    canvas.itemconfig(word, text=french_word, fill="black")
    canvas.itemconfig(title, text="French", fill="black")
    canvas.itemconfig(card_image, image=card_front_img)
    flip_timer = window.after(3000, func=flip_card)

def flip_card():
    canvas.itemconfig(card_image, image=card_back_img)
    english_word = chosen_word["English"]
    canvas.itemconfig(word, text=english_word, fill="white")
    canvas.itemconfig(title, text="English", fill="white")

def guessed_correct():
    pd_records.remove(chosen_word)
    print(len(pd_records))
    to_learn_words = pandas.DataFrame(pd_records)
    to_learn_words.to_csv("data/to_learn_words.csv", index= False)                  # index = False means don't add the index to the file, without this it will keep adding indexes as new columns causing issues
    next_card()

window = Tk()
window.title("Flashy")
window.config(padx=50, pady=50, bg=BACKGROUND_COLOR)

flip_timer = window.after(3000, func=flip_card)                      # after certain time execute this function


canvas = Canvas(width=800, height=526, bg=BACKGROUND_COLOR, highlightthickness=0)
card_front_img = PhotoImage(file="images/card_front.png")
card_back_img = PhotoImage(file="images/card_back.png")
right_img = PhotoImage(file="images/right.png")
wrong_img = PhotoImage(file="images/wrong.png")

card_image = canvas.create_image(410, 263, image=card_front_img)
title = canvas.create_text(400, 150,text="French", font=("Arial", 40, "italic"))
word = canvas.create_text(400, 263,text="word", font=("Arial", 60, "bold"))
canvas.grid(column=0, row=0, columnspan=2)


wrong_btn = Button(image=wrong_img, highlightthickness=0, command=next_card)
wrong_btn.grid(column=0, row=1)


right_btn = Button(image=right_img, highlightthickness=0, command=guessed_correct)
right_btn.grid(column=1, row=1)

next_card()

window.mainloop()

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

# Flash Card GUI App (her code) (run in respective project)

from tkinter import *
import pandas
import random

BACKGROUND_COLOR = "#B1DDC6"
current_card = {}
to_learn = {}

try:
    data = pandas.read_csv("data/words_to_learn.csv")
except FileNotFoundError:
    original_data = pandas.read_csv("data/french_words.csv")
    print(original_data)
    to_learn = original_data.to_dict(orient="records")
else:
    to_learn = data.to_dict(orient="records")


def next_card():
    global current_card, flip_timer
    window.after_cancel(flip_timer)
    current_card = random.choice(to_learn)
    canvas.itemconfig(card_title, text="French", fill="black")
    canvas.itemconfig(card_word, text=current_card["French"], fill="black")
    canvas.itemconfig(card_background, image=card_front_img)
    flip_timer = window.after(3000, func=flip_card)


def flip_card():
    canvas.itemconfig(card_title, text="English", fill="white")
    canvas.itemconfig(card_word, text=current_card["English"], fill="white")
    canvas.itemconfig(card_background, image=card_back_img)


def is_known():
    to_learn.remove(current_card)
    print(len(to_learn))
    data = pandas.DataFrame(to_learn)
    data.to_csv("data/words_to_learn.csv", index=False)
    next_card()


window = Tk()
window.title("Flashy")
window.config(padx=50, pady=50, bg=BACKGROUND_COLOR)

flip_timer = window.after(3000, func=flip_card)

canvas = Canvas(width=800, height=526)
card_front_img = PhotoImage(file="images/card_front.png")
card_back_img = PhotoImage(file="images/card_back.png")
card_background = canvas.create_image(400, 263, image=card_front_img)
card_title = canvas.create_text(400, 150, text="", font=("Ariel", 40, "italic"))
card_word = canvas.create_text(400, 263, text="", font=("Ariel", 60, "bold"))
canvas.config(bg=BACKGROUND_COLOR, highlightthickness=0)
canvas.grid(row=0, column=0, columnspan=2)

cross_image = PhotoImage(file="images/wrong.png")
unknown_button = Button(image=cross_image, highlightthickness=0, command=next_card)
unknown_button.grid(row=1, column=0)

check_image = PhotoImage(file="images/right.png")
known_button = Button(image=check_image, highlightthickness=0, command=is_known)
known_button.grid(row=1, column=1)

next_card()

window.mainloop()



